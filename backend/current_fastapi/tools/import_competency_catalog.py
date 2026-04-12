from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


from app.repositories.competency_repository import CompetencyRepository, build_catalog_source_key


DEFAULT_SHEET_NAME = "Liste des essais"


def normalize_cell(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    return text or None


def import_catalog(workbook_path: Path, sheet_name: str, deactivate_missing: bool) -> tuple[int, int, list[int]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[sheet_name]
    repository = CompetencyRepository()
    imported_source_keys: set[str] = set()
    imported_rows = 0
    skipped_rows: list[int] = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
        domain = normalize_cell(row[0])
        context_type = normalize_cell(row[1])
        label = normalize_cell(row[2])

        if not (domain and context_type and label):
            if any(cell not in (None, "") for cell in row[:4]):
                skipped_rows.append(row_index)
            continue

        reference = normalize_cell(row[3])
        publication_date = normalize_cell(row[4])
        simplified_protocol = normalize_cell(row[5])
        certification = normalize_cell(row[6])
        standard_referent = normalize_cell(row[7])
        standard_update_impact = normalize_cell(row[8])
        trainer_name = normalize_cell(row[9])

        source_key = build_catalog_source_key(domain, context_type, label, reference)
        repository.upsert_catalog_entry(
            source_key=source_key,
            domain=domain,
            context_type=context_type,
            label=label,
            reference=reference,
            publication_date=publication_date,
            simplified_protocol=simplified_protocol,
            certification=certification,
            standard_referent=standard_referent,
            standard_update_impact=standard_update_impact,
            trainer_name=trainer_name,
            is_active=True,
        )
        imported_source_keys.add(source_key)
        imported_rows += 1

    deactivated_rows = 0
    if deactivate_missing:
        deactivated_rows = repository.deactivate_missing_catalog_entries(imported_source_keys)

    return imported_rows, deactivated_rows, skipped_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Import competency catalog from the official workbook.")
    parser.add_argument("workbook_path", type=Path, help="Path to the Excel workbook.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help=f"Sheet to import (default: {DEFAULT_SHEET_NAME!r}).")
    parser.add_argument(
        "--deactivate-missing",
        action="store_true",
        help="Mark existing catalog entries inactive when they are absent from the imported sheet.",
    )
    args = parser.parse_args()

    workbook_path = args.workbook_path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    imported_rows, deactivated_rows, skipped_rows = import_catalog(
        workbook_path=workbook_path,
        sheet_name=args.sheet,
        deactivate_missing=args.deactivate_missing,
    )
    print(f"[OK] catalog imported: {imported_rows} row(s)")
    print(f"[OK] catalog deactivated: {deactivated_rows} row(s)")
    if skipped_rows:
        print(f"[WARN] skipped malformed rows: {skipped_rows}")


if __name__ == "__main__":
    main()