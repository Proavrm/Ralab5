from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from api.import_essais_de import _build_resultats_payload, _extract_sheet_row
from app.core.database import get_db_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Reimport existing DE essais from their source Excel file/sheet.")
    parser.add_argument(
        "--db",
        type=Path,
        default=get_db_path(),
        help="SQLite database path (default: app/core/database configured path)",
    )
    parser.add_argument(
        "--storage-root",
        action="append",
        default=[],
        help="Root folder to scan for Excel files (can be repeated)",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Apply updates. By default script runs in dry-run mode.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Process at most N DE essais (0 = no limit)",
    )
    parser.add_argument(
        "--essai-id",
        type=int,
        action="append",
        default=[],
        help="Restrict processing to specific essai id (can be repeated)",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Show skipped details and per-row updates.",
    )
    return parser.parse_args()


def connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn


def read_de_rows(conn: sqlite3.Connection, target_ids: set[int]) -> list[sqlite3.Row]:
    rows = conn.execute(
        """
        SELECT id, intervention_id, source_signature, resultats
        FROM essais
        WHERE essai_code = 'DE'
        ORDER BY id
        """
    ).fetchall()
    if target_ids:
        rows = [row for row in rows if int(row["id"]) in target_ids]
    return rows


def parse_payload(raw_payload: Any) -> dict[str, Any] | None:
    if not raw_payload:
        return None
    try:
        parsed = json.loads(raw_payload)
    except Exception:
        return None
    return parsed if isinstance(parsed, dict) else None


def build_file_index(storage_roots: list[Path]) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = defaultdict(list)
    for root in storage_roots:
        if not root.exists() or not root.is_dir():
            continue
        for ext in ("*.xlsx", "*.xlsm"):
            for path in root.rglob(ext):
                index[path.name.lower()].append(path)
    return index


def compute_sha1(path: Path) -> str:
    return hashlib.sha1(path.read_bytes()).hexdigest()


def resolve_workbook_path(
    file_name: str,
    expected_hash: str,
    sheet_name: str,
    file_index: dict[str, list[Path]],
    hash_cache: dict[Path, str],
    workbook_cache: dict[Path, Any],
) -> tuple[Path | None, str, str]:
    candidates = file_index.get(file_name.lower(), [])
    if not candidates:
        return None, "", "file_not_found"

    matched: list[tuple[Path, str]] = []
    fallback: list[tuple[Path, str]] = []
    for path in candidates:
        workbook = workbook_cache.get(path)
        if workbook is None:
            try:
                workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
                workbook_cache[path] = workbook
            except Exception:
                continue

        if sheet_name not in workbook.sheetnames:
            continue

        file_hash = hash_cache.get(path)
        if file_hash is None:
            try:
                file_hash = compute_sha1(path)
            except Exception:
                continue
            hash_cache[path] = file_hash

        fallback.append((path, file_hash))
        if expected_hash and file_hash == expected_hash:
            matched.append((path, file_hash))

    if matched:
        return matched[0][0], matched[0][1], "ok"
    if len(fallback) == 1:
        return fallback[0][0], fallback[0][1], "ok_hash_mismatch"
    if len(fallback) > 1:
        return None, "", "ambiguous_file"
    return None, "", "sheet_not_found"


def build_updated_row_payload(
    workbook_path: Path,
    workbook_hash: str,
    sheet_name: str,
) -> tuple[dict[str, Any] | None, str]:
    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:
        return None, f"workbook_read_error:{exc}"
    if sheet_name not in workbook.sheetnames:
        return None, "sheet_not_found"

    ws = workbook[sheet_name]
    sheet_row = _extract_sheet_row(ws)
    payload = _build_resultats_payload(sheet_row, workbook_path.name, workbook_hash)
    return payload, "ok"


def main() -> None:
    args = parse_args()
    db_path = Path(args.db)
    storage_roots = [Path(path) for path in args.storage_root] or [ROOT_DIR.parent.parent / "storage"]

    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")

    file_index = build_file_index(storage_roots)
    hash_cache: dict[Path, str] = {}
    workbook_cache: dict[Path, Any] = {}
    counters = Counter()
    skipped_samples: list[dict[str, Any]] = []
    updated_samples: list[dict[str, Any]] = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with connect(db_path) as conn:
        rows = read_de_rows(conn, set(args.essai_id))
        if args.limit and args.limit > 0:
            rows = rows[: args.limit]

        counters["scanned_de_essais"] = len(rows)

        for row in rows:
            essai_id = int(row["id"])
            payload = parse_payload(row["resultats"])
            if not payload:
                counters["skipped_invalid_payload"] += 1
                skipped_samples.append({"essai_id": essai_id, "reason": "invalid_payload_json"})
                continue
            if payload.get("source") != "de_excel_import":
                counters["skipped_non_excel_import"] += 1
                continue

            file_name = str(payload.get("file_name") or "").strip()
            file_hash = str(payload.get("file_hash") or "").strip()
            sheet_name = str(payload.get("sheet") or "").strip()
            if not file_name or not sheet_name:
                counters["skipped_missing_file_or_sheet"] += 1
                skipped_samples.append({"essai_id": essai_id, "reason": "missing_file_or_sheet"})
                continue

            workbook_path, workbook_hash, resolve_status = resolve_workbook_path(
                file_name=file_name,
                expected_hash=file_hash,
                sheet_name=sheet_name,
                file_index=file_index,
                hash_cache=hash_cache,
                workbook_cache=workbook_cache,
            )
            if not workbook_path:
                counters[f"skipped_{resolve_status}"] += 1
                skipped_samples.append(
                    {
                        "essai_id": essai_id,
                        "file_name": file_name,
                        "sheet": sheet_name,
                        "reason": resolve_status,
                    }
                )
                continue

            updated_payload, build_status = build_updated_row_payload(
                workbook_path=workbook_path,
                workbook_hash=workbook_hash,
                sheet_name=sheet_name,
            )
            if not updated_payload:
                counters[f"skipped_{build_status}"] += 1
                skipped_samples.append(
                    {
                        "essai_id": essai_id,
                        "file_name": file_name,
                        "sheet": sheet_name,
                        "reason": build_status,
                    }
                )
                continue

            if resolve_status == "ok_hash_mismatch":
                counters["reimported_from_hash_mismatch_file"] += 1

            moyenne_comp = updated_payload.get("resume", {}).get("moyenne_compacite_pct")
            try:
                resultat_principal = float(moyenne_comp) if moyenne_comp is not None else None
            except Exception:
                resultat_principal = None
            resultat_label = ""
            if resultat_principal is not None:
                resultat_label = f"Compacite moyenne = {resultat_principal:.2f} %"

            counters["ready_to_update"] += 1
            if args.apply:
                conn.execute(
                    """
                    UPDATE essais
                    SET resultats = ?,
                        date_debut = COALESCE(?, date_debut),
                        date_fin = COALESCE(?, date_fin),
                        operateur = COALESCE(?, operateur),
                        resultat_principal = ?,
                        resultat_unite = '%',
                        resultat_label = ?,
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        json.dumps(updated_payload, ensure_ascii=False),
                        updated_payload.get("meta", {}).get("date_essai") or None,
                        updated_payload.get("meta", {}).get("date_redaction") or None,
                        updated_payload.get("meta", {}).get("operateur") or None,
                        resultat_principal,
                        resultat_label,
                        now,
                        essai_id,
                    ),
                )
                counters["updated"] += 1

            if len(updated_samples) < 20:
                updated_samples.append(
                    {
                        "essai_id": essai_id,
                        "file": workbook_path.name,
                        "sheet": sheet_name,
                        "points": updated_payload.get("resume", {}).get("points"),
                        "conformite_pct": updated_payload.get("resume", {}).get("conformite_pct"),
                        "criteria_definition": updated_payload.get("meta", {}).get("criteria_definition"),
                    }
                )

        if args.apply:
            conn.commit()

    summary = {
        "mode": "apply" if args.apply else "dry-run",
        "db_path": str(db_path),
        "storage_roots": [str(path) for path in storage_roots],
        "counts": dict(counters),
        "updated_preview": updated_samples,
        "skipped_preview": skipped_samples[:30],
    }
    print(json.dumps(summary, indent=2, ensure_ascii=False))

    if args.verbose and skipped_samples:
        for item in skipped_samples:
            print(item)


if __name__ == "__main__":
    main()
