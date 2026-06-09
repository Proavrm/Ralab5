"""Import CODIR REX content from the enriched Excel register into qsse_rex_drafts.

Designed for repeatable daily runs:
- idempotent upsert on qsse_record_id
- row-level mapping via "N° ligne registre" -> qsse_records.row_index
- fallback to document reference or title only when row mapping is unavailable
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sqlite3
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT_DIR.parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.core.database import get_qsse_db_path


DEFAULT_EXCEL_GLOB = "FNC_2026_REX_provisoires_CODIR_RaLab5*.xlsx"
DEFAULT_EXCEL_DIR = PROJECT_ROOT / "storage" / "qsse"
DEFAULT_SHEET_NAME = "Registre enrichi"
DEFAULT_SOURCE_YEAR = 2026
DEFAULT_PROVIDER = "codir-xlsx-import"
DEFAULT_PROMPT_VERSION = "codir-rex-import-v2"


@dataclass(slots=True)
class ExcelRow:
    line_no: int | None
    reference_document: str
    constat_resume: str
    rex_text: str
    cause_root: str
    action_structurante: str
    preuves: str
    missing_fields: str
    family: str
    subfamily: str
    criticite: str
    score_rex_10: Any


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import CODIR REX rows from an Excel sheet into qsse_rex_drafts.",
    )
    parser.add_argument("--excel-path", type=Path, help="Explicit path to the Excel file.")
    parser.add_argument(
        "--excel-dir",
        type=Path,
        default=DEFAULT_EXCEL_DIR,
        help="Directory used to resolve the latest file when --excel-path is omitted.",
    )
    parser.add_argument(
        "--excel-glob",
        default=DEFAULT_EXCEL_GLOB,
        help="Glob pattern used to pick the latest Excel in --excel-dir.",
    )
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME, help="Sheet name to import.")
    parser.add_argument("--source-year", type=int, default=DEFAULT_SOURCE_YEAR, help="Target source_year in qsse_records.")
    parser.add_argument("--preview", action="store_true", help="Only validate and report mapping, without writing DB changes.")
    parser.add_argument("--backup", action="store_true", help="Backup qsse.db before applying changes.")
    parser.add_argument("--db-path", type=Path, help="Override SQLite DB path.")
    return parser.parse_args()


def _normalize(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def _split_list(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    parts = re.split(r"\s*(?:;|\||/|\n|•)\s*", text)
    return [item.strip() for item in parts if item and item.strip()][:8]


def _resolve_excel_path(args: argparse.Namespace) -> Path:
    if args.excel_path:
        return Path(args.excel_path)
    excel_dir = Path(args.excel_dir)
    candidates = sorted(excel_dir.glob(args.excel_glob), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"No Excel files found in {excel_dir} matching {args.excel_glob}")
    return candidates[0]


def _read_excel_rows(excel_path: Path, sheet_name: str) -> list[ExcelRow]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        headers = [str(cell.value or "").strip() for cell in ws[3]]
        index = {header: i for i, header in enumerate(headers) if header}

        required_columns = {"N° ligne registre", "Référence document", "Constat résumé"}
        missing = sorted(col for col in required_columns if col not in index)
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        rows: list[ExcelRow] = []
        for values in ws.iter_rows(min_row=4, values_only=True):
            ref = values[index["Référence document"]] if index["Référence document"] < len(values) else None
            constat = values[index["Constat résumé"]] if index["Constat résumé"] < len(values) else None
            line_raw = values[index["N° ligne registre"]] if index["N° ligne registre"] < len(values) else None

            if not str(ref or "").strip() and not str(constat or "").strip() and not str(line_raw or "").strip():
                continue

            try:
                line_no = int(float(line_raw)) if str(line_raw or "").strip() else None
            except Exception:
                line_no = None

            def val(name: str) -> str:
                pos = index.get(name)
                if pos is None or pos >= len(values):
                    return ""
                return str(values[pos] or "").strip()

            rows.append(
                ExcelRow(
                    line_no=line_no,
                    reference_document=val("Référence document"),
                    constat_resume=val("Constat résumé"),
                    rex_text=val("REX provisoire proposé"),
                    cause_root=val("Cause racine proposée") or val("Cause initiale"),
                    action_structurante=val("Action structurante proposée") or val("Action corrective initiale"),
                    preuves=val("Preuves / pièces à récupérer"),
                    missing_fields=val("Champs à compléter"),
                    family=val("Famille REX proposée"),
                    subfamily=val("Sous-famille REX proposée"),
                    criticite=val("Criticité REX"),
                    score_rex_10=(values[index["Score REX /10"]] if "Score REX /10" in index and index["Score REX /10"] < len(values) else None),
                )
            )
        return rows
    finally:
        wb.close()


def _build_mapping_indexes(conn: sqlite3.Connection, source_year: int) -> tuple[dict[int, list[sqlite3.Row]], dict[str, list[int]], dict[str, list[int]]]:
    db_rows = conn.execute(
        """
        SELECT id, row_index, sheet_name, document_reference, title
        FROM qsse_records
        WHERE register_code = 'FNC' AND record_kind = 'event' AND source_year = ?
        """,
        (int(source_year),),
    ).fetchall()

    by_row: dict[int, list[sqlite3.Row]] = {}
    by_ref: dict[str, list[int]] = {}
    by_title: dict[str, list[int]] = {}

    for row in db_rows:
        by_row.setdefault(int(row["row_index"] or 0), []).append(row)
        by_ref.setdefault(_normalize(row["document_reference"]), []).append(int(row["id"]))
        by_title.setdefault(_normalize(row["title"]), []).append(int(row["id"]))

    return by_row, by_ref, by_title


def _pick_record_id(
    item: ExcelRow,
    by_row: dict[int, list[sqlite3.Row]],
    by_ref: dict[str, list[int]],
    by_title: dict[str, list[int]],
) -> tuple[int | None, str]:
    if item.line_no:
        candidates = by_row.get(int(item.line_no), [])
        if candidates:
            preferred = [c for c in candidates if "registre fnc" in str(c["sheet_name"] or "").lower()]
            chosen = preferred[0] if preferred else candidates[0]
            return int(chosen["id"]), "row_index"

    ref_candidates = by_ref.get(_normalize(item.reference_document), [])
    if len(ref_candidates) == 1:
        return ref_candidates[0], "reference"

    title_candidates = by_title.get(_normalize(item.constat_resume), [])
    if len(title_candidates) == 1:
        return title_candidates[0], "title"

    return None, "unmatched"


def _confidence_score(score_rex_10: Any) -> int:
    try:
        value = float(score_rex_10)
    except Exception:
        return 72
    return max(0, min(100, int(round(value * 10))))


def _build_draft_payload(item: ExcelRow, record_id: int) -> tuple[dict[str, Any], dict[str, Any], int]:
    draft = {
        "headline": item.constat_resume or item.reference_document or f"FNC-{record_id}",
        "summary": item.rex_text or item.constat_resume or item.reference_document,
        "lesson_learned": item.cause_root or "Capitaliser le cas et renforcer les controles de preparation.",
        "root_cause_synthesis": item.cause_root or "Cause racine non renseignee dans le fichier CODIR.",
        "preventive_action": item.action_structurante or "Definir une action structurante avec pilote, delai et controle de verification.",
        "diffusion_message": f"Partager ce REX en revue QSSE avec les equipes concernees ({item.family or 'FNC'}, {item.subfamily or 'n/a'}).",
        "target_audience": ["Encadrement chantier", "Qualite / QSSE", "Prevention"],
        "missing_information": _split_list(item.missing_fields),
        "evidence": [
            evidence
            for evidence in (
                f"Reference: {item.reference_document}" if item.reference_document else "",
                f"Criticite: {item.criticite}" if item.criticite else "",
                item.preuves,
            )
            if evidence
        ],
    }

    payload = {
        "reference_document": item.reference_document,
        "constat_resume": item.constat_resume,
        "line_no": item.line_no,
        "famille_rex": item.family,
        "sous_famille_rex": item.subfamily,
        "criticite_rex": item.criticite,
    }
    return draft, payload, _confidence_score(item.score_rex_10)


def _backup_db(db_path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = db_path.with_name(f"{db_path.stem}.backup_rex_daily_sync_{stamp}{db_path.suffix}")
    shutil.copy2(db_path, target)
    return target


def main() -> None:
    args = parse_args()

    excel_path = _resolve_excel_path(args)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    db_path = Path(args.db_path) if args.db_path else get_qsse_db_path()
    if not db_path.exists():
        raise SystemExit(f"QSSE database not found: {db_path}")

    excel_rows = _read_excel_rows(excel_path, args.sheet_name)

    backup_path = None
    if args.backup and not args.preview:
        backup_path = _backup_db(db_path)

    summary = {
        "mode": "preview" if args.preview else "apply",
        "excel_path": str(excel_path),
        "db_path": str(db_path),
        "source_year": int(args.source_year),
        "excel_rows": len(excel_rows),
        "backup_path": str(backup_path) if backup_path else None,
        "matched_row_index": 0,
        "matched_reference": 0,
        "matched_title": 0,
        "unmatched": 0,
        "imported": 0,
        "unique_records_touched": 0,
        "provider_status_counts": [],
    }

    touched_ids: set[int] = set()

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        by_row, by_ref, by_title = _build_mapping_indexes(conn, int(args.source_year))

        for item in excel_rows:
            record_id, strategy = _pick_record_id(item, by_row, by_ref, by_title)
            if strategy == "row_index":
                summary["matched_row_index"] += 1
            elif strategy == "reference":
                summary["matched_reference"] += 1
            elif strategy == "title":
                summary["matched_title"] += 1
            else:
                summary["unmatched"] += 1
                continue

            if args.preview:
                touched_ids.add(int(record_id))
                continue

            draft, payload, confidence = _build_draft_payload(item, int(record_id))
            conn.execute(
                """
                INSERT INTO qsse_rex_drafts (
                    qsse_record_id,
                    provider,
                    prompt_version,
                    status,
                    confidence_score,
                    source_payload_json,
                    draft_json,
                    generated_at,
                    reviewed_at,
                    updated_at
                ) VALUES (?, ?, ?, 'reviewed', ?, ?, ?, datetime('now'), datetime('now'), datetime('now'))
                ON CONFLICT(qsse_record_id) DO UPDATE SET
                    provider = excluded.provider,
                    prompt_version = excluded.prompt_version,
                    status = 'reviewed',
                    confidence_score = excluded.confidence_score,
                    source_payload_json = excluded.source_payload_json,
                    draft_json = excluded.draft_json,
                    generated_at = datetime('now'),
                    reviewed_at = datetime('now'),
                    updated_at = datetime('now')
                """,
                (
                    int(record_id),
                    DEFAULT_PROVIDER,
                    DEFAULT_PROMPT_VERSION,
                    int(confidence),
                    json.dumps(payload, ensure_ascii=False),
                    json.dumps(draft, ensure_ascii=False),
                ),
            )
            summary["imported"] += 1
            touched_ids.add(int(record_id))

        if not args.preview:
            conn.commit()

        provider_counts = conn.execute(
            """
            SELECT d.provider, d.status, COUNT(*) AS n
            FROM qsse_rex_drafts d
            JOIN qsse_records r ON r.id = d.qsse_record_id
            WHERE r.register_code='FNC' AND r.record_kind='event' AND r.source_year = ?
            GROUP BY d.provider, d.status
            ORDER BY n DESC
            """,
            (int(args.source_year),),
        ).fetchall()

    summary["unique_records_touched"] = len(touched_ids)
    summary["provider_status_counts"] = [dict(row) for row in provider_counts]

    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
