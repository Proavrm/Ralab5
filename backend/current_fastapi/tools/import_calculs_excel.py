"""
Import Excel RaLab_compilation_normalisee → tables ref_* (Calculs).

Usage:
  python -m tools.import_calculs_excel
  python -m tools.import_calculs_excel --xlsx path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from app.core.database import ensure_ralab5_schema, get_db_path


DEFAULT_XLSX = (
    Path(__file__).resolve().parents[3]
    / "storage"
    / "documents"
    / "RaLab_compilation_normalisee(1).xlsx"
)


def _cell(row: dict, *keys):
    for key in keys:
        if key in row and row[key] is not None:
            return row[key]
    return None


def _num(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _rows(ws):
    headers = [str(c.value).strip() if c.value is not None else f"col{i}" for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        yield {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}


def import_excel(xlsx_path: Path, db_path: Path | None = None) -> dict:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(path)
    db = db_path or get_db_path()
    ensure_ralab5_schema(db)
    wb = load_workbook(path, read_only=True, data_only=True)
    counts = {}

    with sqlite3.connect(db) as conn:
        conn.row_factory = sqlite3.Row

        if "ALIZE_etudes" in wb.sheetnames:
            conn.execute("DELETE FROM ref_alize_etudes")
            n = 0
            for r in _rows(wb["ALIZE_etudes"]):
                conn.execute(
                    """
                    INSERT INTO ref_alize_etudes (
                        source_id, document, source_ref, famille, projet, structure, ep_bit_cm,
                        plateforme, module_pf_MPa, trafic_PL, MJA_PL, croissance_pct, duree_ans,
                        materiau_critique, module_crit_MPa, CAM, NE, risque_pct,
                        epsT_adm, epsT_calc, marge_fatigue, conso_fatigue,
                        epsZ_adm, epsZ_calc, marge_pf, conso_pf, sigmaT_MPa, sigmaZ_MPa,
                        conclusion, statut_extraction, is_primary, reference_only
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,1)
                    """,
                    (
                        str(_cell(r, "id") or ""),
                        str(_cell(r, "document") or ""),
                        str(_cell(r, "source_ref") or ""),
                        str(_cell(r, "famille") or ""),
                        str(_cell(r, "projet") or ""),
                        str(_cell(r, "structure") or ""),
                        _num(_cell(r, "ep_bit_cm")),
                        str(_cell(r, "plateforme") or ""),
                        _num(_cell(r, "module_pf_MPa")),
                        str(_cell(r, "trafic_PL") or ""),
                        _num(_cell(r, "MJA_PL")),
                        _num(_cell(r, "croissance_pct")),
                        _num(_cell(r, "duree_ans")),
                        str(_cell(r, "materiau_critique") or ""),
                        _num(_cell(r, "module_crit_MPa")),
                        _num(_cell(r, "CAM")),
                        _num(_cell(r, "NE")),
                        _num(_cell(r, "risque_pct")),
                        _num(_cell(r, "epsT_adm")),
                        _num(_cell(r, "epsT_calc")),
                        _num(_cell(r, "marge_fatigue")),
                        _num(_cell(r, "conso_fatigue")),
                        _num(_cell(r, "epsZ_adm")),
                        _num(_cell(r, "epsZ_calc")),
                        _num(_cell(r, "marge_pf")),
                        _num(_cell(r, "conso_pf")),
                        _num(_cell(r, "sigmaT_MPa")),
                        _num(_cell(r, "sigmaZ_MPa")),
                        str(_cell(r, "conclusion") or ""),
                        str(_cell(r, "statut_extraction") or ""),
                    ),
                )
                n += 1
            counts["ALIZE_etudes"] = n

        if "ALIZE_couches" in wb.sheetnames:
            conn.execute("DELETE FROM ref_alize_couches")
            n = 0
            for r in _rows(wb["ALIZE_couches"]):
                conn.execute(
                    """
                    INSERT INTO ref_alize_couches (
                        id_etude, document, source_ref, ordre, materiau, epaisseur_cm, module_MPa, plateforme
                    ) VALUES (?,?,?,?,?,?,?,?)
                    """,
                    (
                        int(_cell(r, "id_etude") or 0) or None,
                        str(_cell(r, "document") or ""),
                        str(_cell(r, "source_ref") or ""),
                        int(_cell(r, "ordre") or 0) or None,
                        str(_cell(r, "materiau") or ""),
                        _num(_cell(r, "epaisseur_cm")),
                        _num(_cell(r, "module_MPa_si_connu", "module_MPa")),
                        str(_cell(r, "plateforme") or ""),
                    ),
                )
                n += 1
            counts["ALIZE_couches"] = n

        if "ALIZE_criteres" in wb.sheetnames:
            conn.execute("DELETE FROM ref_alize_criteres")
            n = 0
            for r in _rows(wb["ALIZE_criteres"]):
                conn.execute(
                    """
                    INSERT INTO ref_alize_criteres (
                        id_etude, document, source_ref, critere, materiau,
                        admissible_microdef, calcule_microdef, marge_microdef,
                        consommation_pct, sigma_MPa, statut
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        int(_cell(r, "id_etude") or 0) or None,
                        str(_cell(r, "document") or ""),
                        str(_cell(r, "source_ref") or ""),
                        str(_cell(r, "critere") or ""),
                        str(_cell(r, "materiau") or ""),
                        _num(_cell(r, "admissible_microdef")),
                        _num(_cell(r, "calcule_microdef")),
                        _num(_cell(r, "marge_microdef")),
                        _num(_cell(r, "consommation_pct")),
                        _num(_cell(r, "sigma_MPa")),
                        str(_cell(r, "statut") or ""),
                    ),
                )
                n += 1
            counts["ALIZE_criteres"] = n

        if "GEL_degel" in wb.sheetnames:
            conn.execute("DELETE FROM ref_gel_degel")
            n = 0
            for r in _rows(wb["GEL_degel"]):
                conn.execute(
                    """
                    INSERT INTO ref_gel_degel (
                        document, source_ref, projet, structure, station, hiver,
                        Ir_Cj, Ia_Cj, marge_Cj, Qng, Qg, Qm, Qpf, temps_jours, Zgel_m,
                        conclusion, statut, reference_only
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)
                    """,
                    (
                        str(_cell(r, "document") or ""),
                        str(_cell(r, "source_ref") or ""),
                        str(_cell(r, "projet") or ""),
                        str(_cell(r, "structure") or ""),
                        str(_cell(r, "station") or ""),
                        str(_cell(r, "hiver") or ""),
                        _num(_cell(r, "Ir_Cj")),
                        _num(_cell(r, "Ia_Cj")),
                        _num(_cell(r, "marge_Cj")),
                        _num(_cell(r, "Qng")),
                        _num(_cell(r, "Qg")),
                        _num(_cell(r, "Qm")),
                        _num(_cell(r, "Qpf")),
                        _num(_cell(r, "temps_jours")),
                        _num(_cell(r, "Zgel_m")),
                        str(_cell(r, "conclusion") or ""),
                        str(_cell(r, "statut") or ""),
                    ),
                )
                n += 1
            counts["GEL_degel"] = n

        if "MATERIAUX_labo_refs" in wb.sheetnames:
            conn.execute("DELETE FROM ref_materiaux_labo")
            n = 0
            for r in _rows(wb["MATERIAUX_labo_refs"]):
                conn.execute(
                    """
                    INSERT INTO ref_materiaux_labo (
                        document, source_ref, famille, produit_ou_reference, norme_source, formule,
                        bitume, granulats, TL_pct, TLmin_corrigee_pct, MVRg, K, module_E_MPa,
                        eps6_microdef, vides_fatigue_pct, ITSR_pct, Duriez_iC_pct,
                        orniere_30000_pct, vides_orniere_pct, PCG_80_girations_vides_pct, commentaire
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        str(_cell(r, "document") or ""),
                        str(_cell(r, "source_ref") or ""),
                        str(_cell(r, "famille") or ""),
                        str(_cell(r, "produit_ou_reference") or ""),
                        str(_cell(r, "norme_source") or ""),
                        str(_cell(r, "formule") or ""),
                        str(_cell(r, "bitume") or ""),
                        str(_cell(r, "granulats") or ""),
                        _num(_cell(r, "TL_pct")),
                        _num(_cell(r, "TLmin_corrigee_pct")),
                        _num(_cell(r, "MVRg")),
                        _num(_cell(r, "K")),
                        _num(_cell(r, "module_E_MPa")),
                        _num(_cell(r, "eps6_microdef")),
                        _num(_cell(r, "vides_fatigue_pct")),
                        _num(_cell(r, "ITSR_pct")),
                        _num(_cell(r, "Duriez_iC_pct")),
                        _num(_cell(r, "orniere_30000_pct")),
                        _num(_cell(r, "vides_orniere_pct")),
                        _num(_cell(r, "PCG_80_girations_vides_pct")),
                        str(_cell(r, "commentaire") or ""),
                    ),
                )
                n += 1
            counts["MATERIAUX_labo_refs"] = n

        if "GEOTECH_TALREN_images" in wb.sheetnames:
            conn.execute("DELETE FROM ref_talren_images")
            n = 0
            for r in _rows(wb["GEOTECH_TALREN_images"]):
                conn.execute(
                    """
                    INSERT INTO ref_talren_images (
                        document, source_ref, famille, coupe, type_ouvrage, hauteur_m, largeur_m,
                        rapport_BH, alt_sommet, alt_pied, niveau_terrassement,
                        Fmin_default, Fmin_seisme, Fmin_crue, Fmin_decrue, cas_critique,
                        drainage_ou_particularite, reference_only
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)
                    """,
                    (
                        str(_cell(r, "document") or ""),
                        str(_cell(r, "source_ref") or ""),
                        str(_cell(r, "famille") or ""),
                        str(_cell(r, "coupe") or ""),
                        str(_cell(r, "type_ouvrage") or ""),
                        _num(_cell(r, "hauteur_m")),
                        _num(_cell(r, "largeur_m")),
                        _num(_cell(r, "rapport_BH")),
                        _num(_cell(r, "alt_sommet")),
                        _num(_cell(r, "alt_pied")),
                        str(_cell(r, "niveau_terrassement") or ""),
                        _num(_cell(r, "Fmin_default")),
                        _num(_cell(r, "Fmin_seisme")),
                        _num(_cell(r, "Fmin_crue")),
                        _num(_cell(r, "Fmin_decrue")),
                        str(_cell(r, "cas_critique") or ""),
                        str(_cell(r, "drainage_ou_particularite") or ""),
                    ),
                )
                n += 1
            counts["GEOTECH_TALREN_images"] = n

        if "DOCUMENTS_index" in wb.sheetnames:
            conn.execute("DELETE FROM ref_documents_index")
            n = 0
            for r in _rows(wb["DOCUMENTS_index"]):
                conn.execute(
                    "INSERT INTO ref_documents_index (document, source_ref, famille, statut, note) VALUES (?,?,?,?,?)",
                    (
                        str(_cell(r, "document") or ""),
                        str(_cell(r, "source_ref") or ""),
                        str(_cell(r, "famille") or ""),
                        str(_cell(r, "statut") or ""),
                        str(_cell(r, "note") or ""),
                    ),
                )
                n += 1
            counts["DOCUMENTS_index"] = n

        if "DOUBLONS_controle" in wb.sheetnames:
            conn.execute("DELETE FROM ref_doublons_controle")
            n = 0
            for r in _rows(wb["DOUBLONS_controle"]):
                conn.execute(
                    "INSERT INTO ref_doublons_controle (groupe, document_A, document_B, traitement) VALUES (?,?,?,?)",
                    (
                        str(_cell(r, "groupe") or ""),
                        str(_cell(r, "document_A") or ""),
                        str(_cell(r, "document_B") or ""),
                        str(_cell(r, "traitement") or ""),
                    ),
                )
                n += 1
            counts["DOUBLONS_controle"] = n

        conn.commit()

    return counts


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--db", type=Path, default=None)
    args = parser.parse_args()
    counts = import_excel(args.xlsx, args.db)
    print("Import OK:", counts)


if __name__ == "__main__":
    main()
