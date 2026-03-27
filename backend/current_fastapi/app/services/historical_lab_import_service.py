
"""
historical_lab_import_service.py
One-shot historical laboratory Excel import service for RaLab4.
"""
from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


KNOWN_CODES = {
    "WE": "Teneur en eau",
    "ID": "Identification",
    "LCP": "Limites d'Atterberg",
    "IPI - PR": "Proctor / IPI",
    "IM": "CBRi",
    "CFE": "Contrôle de fabrication enrobés",
    "SC": "Coupe de sondage carotté",
    "SO": "Coupes de sondages",
    "DE": "Densités enrobés",
    "DF": "Déflexion",
    "MVA": "Masse volumique des enrobés",
    "PMT": "Mesure de la profondeur de macrotexture",
    "PLD": "Portances des plates-formes Dynaplaque",
    "SOL": "Analyses pollution",
    "FTP": "Fiche technique produit",
}

SUPPORTED_CODES = {
    code: KNOWN_CODES[code]
    for code in ["WE", "ID", "LCP", "IPI - PR", "IM", "MVA", "DE", "CFE", "PLD", "DF", "PMT", "SC", "SO", "SOL", "FTP"]
}

CODE_TOKEN_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("IPI - PR", re.compile(r"\bIPI\s*[-/]?\s*PR\b", re.IGNORECASE)),
    ("LCP", re.compile(r"\bLCP\b")),
    ("MVA", re.compile(r"\bMVA\b")),
    ("MVA", re.compile(r"\bMVE\b")),
    ("PMT", re.compile(r"\bPMT\b")),
    ("PLD", re.compile(r"\bPLD\b")),
    ("CFE", re.compile(r"\bCFE\b")),
    ("SOL", re.compile(r"\bSOL\b")),
    ("FTP", re.compile(r"\bFTP\b")),
    ("SC", re.compile(r"\bSC\b")),
    ("SO", re.compile(r"\bSO\b")),
    ("DF", re.compile(r"\bDF\b")),
    ("DE", re.compile(r"\bDE(?:\s*N[°ºO]?\s*\d*)?\b")),
    ("WE", re.compile(r"\bWE\b")),
    ("ID", re.compile(r"\bID\b")),
    ("IM", re.compile(r"\bIM\b")),
]

TITLE_KEYWORD_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("DF", re.compile(r"déf?lexion", re.IGNORECASE)),
    ("PMT", re.compile(r"macrotexture", re.IGNORECASE)),
    ("MVA", re.compile(r"masse\s+volumique\s+des\s+enrob", re.IGNORECASE)),
    ("PLD", re.compile(r"dynaplaque|plates?-formes", re.IGNORECASE)),
    ("CFE", re.compile(r"contr[oô]le\s+de\s+fabrication\s+enrob", re.IGNORECASE)),
    ("SC", re.compile(r"coupe\s+de\s+sondage\s+carott", re.IGNORECASE)),
    ("SO", re.compile(r"coupes?\s+de\s+sondages?", re.IGNORECASE)),
    ("SOL", re.compile(r"pollution", re.IGNORECASE)),
    ("FTP", re.compile(r"fiche\s+technique\s+produit", re.IGNORECASE)),
]

REFERENCE_AFFAIRE_COLUMNS = [
    "n°affaire",
    "gsa",
    "ehtp",
    "nge_routes",
    "nge_gc",
    "lyaudet",
    "nge_e.s.",
    "nge_transitions",
]


@dataclass(slots=True)
class ImportPaths:
    target_db_path: Path
    affaires_db_path: Path


class HistoricalLabImportService:
    """Service used for one-shot historical import of soil and asphalt Excel files."""

    def __init__(self, target_db_path: Path, affaires_db_path: Path) -> None:
        self.paths = ImportPaths(
            target_db_path=target_db_path,
            affaires_db_path=affaires_db_path,
        )

    def status(self) -> dict[str, Any]:
        self.ensure_import_schema()
        report = self.report_unmatched_imported_affaires(limit=20)
        return {
            "target_db_path": str(self.paths.target_db_path),
            "affaires_db_path": str(self.paths.affaires_db_path),
            "target_db_exists": self.paths.target_db_path.exists(),
            "affaires_db_exists": self.paths.affaires_db_path.exists(),
            "supported_codes": sorted(SUPPORTED_CODES.keys()),
            "known_codes": {code: KNOWN_CODES[code] for code in sorted(KNOWN_CODES.keys())},
            "unmatched_imported_affaires_count": report["count"],
        }


    def preview_folder(self, folder_path: Path, limit: int = 300) -> dict[str, Any]:
        files = self._scan_supported_files(folder_path)
        candidates: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        workbook_rows: list[dict[str, Any]] = []
        total_sheet_count = 0

        for file_path in files:
            try:
                parsed = self._parse_workbook(file_path)
                total_sheet_count += parsed["sheet_count"]
                workbook_rows.append(
                    {
                        "file_name": file_path.name,
                        "sheet_count": parsed["sheet_count"],
                        "supported_sheet_count": len(parsed["supported_candidates"]),
                        "skipped_sheet_count": len(parsed["skipped_sheets"]),
                    }
                )
                if parsed["supported_candidates"]:
                    candidates.extend(parsed["supported_candidates"])
                if parsed["skipped_sheets"]:
                    skipped.extend(parsed["skipped_sheets"])
            except Exception as exc:
                skipped.append(
                    {
                        "file_name": file_path.name,
                        "sheet_name": "",
                        "reason": f"Workbook read error: {exc}",
                    }
                )

        preview_rows = []
        for item in candidates[:limit]:
            preview_rows.append(
                {
                    "file_name": item["file_name"],
                    "sheet_name": item["sheet_name"],
                    "essai_code": item["essai_code"],
                    "essai_label": item["essai_label"],
                    "affaire_nge": item["affaire_nge"],
                    "affaire_nge_normalized": self._normalize_affaire_key(item["affaire_nge"]),
                    "sample_local_ref": item["sample_local_ref"],
                    "date_essai": item["date_essai"],
                    "date_prelevement": item["date_prelevement"],
                    "operator": item["operator"],
                    "import_mode": item.get("import_mode", "simple"),
                    "target_entity": item.get("target_entity", "echantillon"),
                    "import_family": item.get("import_family", "sols"),
                    "subtest_count": len(item.get("composite_subtests", [])),
                }
            )

        return {
            "folder_path": str(folder_path),
            "xlsx_files_found": len(files),
            "sheet_count": total_sheet_count,
            "supported_candidate_count": len(candidates),
            "skipped_count": len(skipped),
            "preview_rows": preview_rows,
            "skipped_rows": skipped[:limit],
            "workbook_rows": workbook_rows[:limit],
        }



    def run_import(self, folder_path: Path, dry_run: bool = False) -> dict[str, Any]:
        self.ensure_import_schema()
        files = self._scan_supported_files(folder_path)

        candidates: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        total_sheet_count = 0
        for file_path in files:
            try:
                parsed = self._parse_workbook(file_path)
                total_sheet_count += parsed["sheet_count"]
                candidates.extend(parsed["supported_candidates"])
                skipped.extend(parsed["skipped_sheets"])
            except Exception as exc:
                skipped.append(
                    {
                        "file_name": file_path.name,
                        "sheet_name": "",
                        "reason": f"Workbook read error: {exc}",
                    }
                )

        if dry_run:
            dry_run_essais = 0
            dry_run_interventions = 0
            for candidate in candidates:
                if candidate.get("target_entity") == "intervention":
                    dry_run_interventions += 1
                else:
                    dry_run_essais += max(1, len(candidate.get("composite_subtests", [])))
            return {
                "dry_run": True,
                "folder_path": str(folder_path),
                "xlsx_files_found": len(files),
                "sheet_count": total_sheet_count,
                "supported_candidate_count": len(candidates),
                "skipped_count": len(skipped),
                "created_affaires": 0,
                "created_demandes": 0,
                "created_echantillons": 0,
                "created_interventions": dry_run_interventions,
                "created_essais": dry_run_essais,
                "linked_existing_affaires": 0,
                "linked_existing_demandes": 0,
                "linked_existing_echantillons": 0,
                "linked_existing_interventions": 0,
                "skipped_rows": skipped,
            }

        created_affaires = 0
        created_demandes = 0
        created_echantillons = 0
        created_interventions = 0
        created_essais = 0
        linked_existing_affaires = 0
        linked_existing_demandes = 0
        linked_existing_echantillons = 0
        linked_existing_interventions = 0

        with sqlite3.connect(self.paths.target_db_path) as conn:
            conn.row_factory = sqlite3.Row
            batch_id = self._create_batch(conn, folder_path, len(candidates), dry_run=False)

            for candidate in candidates:
                try:
                    affaire_id, affaire_created = self._find_or_create_affaire(conn, candidate)
                    if affaire_created:
                        created_affaires += 1
                    else:
                        linked_existing_affaires += 1

                    demande_id, demande_created = self._find_or_create_demande(conn, affaire_id, candidate)
                    if demande_created:
                        created_demandes += 1
                    else:
                        linked_existing_demandes += 1

                    if candidate.get("target_entity") == "intervention":
                        intervention_id, intervention_created = self._find_or_create_intervention(conn, demande_id, candidate)
                        if intervention_created:
                            created_interventions += 1
                        else:
                            linked_existing_interventions += 1

                        created_essais += self._store_intervention_payload_if_missing(conn, intervention_id, candidate)
                    else:
                        echantillon_id, echantillon_created = self._find_or_create_echantillon(conn, demande_id, candidate)
                        if echantillon_created:
                            created_echantillons += 1
                        else:
                            linked_existing_echantillons += 1

                        created_essais += self._create_essai_if_missing(conn, echantillon_id, candidate)

                    self._create_file_log(conn, batch_id, candidate, "imported", "")
                except Exception as exc:
                    self._create_file_log(conn, batch_id, candidate, "error", str(exc))

            self._finalize_batch(
                conn=conn,
                batch_id=batch_id,
                created_affaires=created_affaires,
                created_demandes=created_demandes,
                created_echantillons=created_echantillons,
                created_essais=created_essais,
                status="done",
                notes=f"Skipped worksheets: {len(skipped)}",
            )

        unmatched_report = self.report_unmatched_imported_affaires(limit=50)
        return {
            "dry_run": False,
            "folder_path": str(folder_path),
            "xlsx_files_found": len(files),
            "sheet_count": total_sheet_count,
            "supported_candidate_count": len(candidates),
            "skipped_count": len(skipped),
            "created_affaires": created_affaires,
            "created_demandes": created_demandes,
            "created_echantillons": created_echantillons,
            "created_interventions": created_interventions,
            "created_essais": created_essais,
            "linked_existing_affaires": linked_existing_affaires,
            "linked_existing_demandes": linked_existing_demandes,
            "linked_existing_echantillons": linked_existing_echantillons,
            "linked_existing_interventions": linked_existing_interventions,
            "unmatched_imported_affaires_count": unmatched_report["count"],
            "skipped_rows": skipped[:200],
        }


    def report_unmatched_imported_affaires(self, limit: int = 200) -> dict[str, Any]:
        self.ensure_import_schema()
        rows: list[dict[str, Any]] = []

        with sqlite3.connect(self.paths.target_db_path) as conn:
            conn.row_factory = sqlite3.Row
            imported_affaires = conn.execute(
                """
                SELECT id, reference, affaire_nge, chantier, titulaire, statut, responsable
                FROM affaires_rst
                WHERE COALESCE(statut, '') = 'Importée'
                  AND COALESCE(responsable, '') = 'Import historique'
                ORDER BY id ASC
                """
            ).fetchall()
            target_rows = conn.execute(
                """
                SELECT id, reference, affaire_nge, chantier, titulaire, statut, responsable
                FROM affaires_rst
                ORDER BY id ASC
                """
            ).fetchall()

            external_rows = self._load_reference_affaires()

            for imported in imported_affaires:
                imported_key = self._normalize_affaire_key(self._stringify(imported["affaire_nge"]))
                if not imported_key:
                    rows.append(
                        {
                            "imported_affaire_id": int(imported["id"]),
                            "imported_reference": self._stringify(imported["reference"]),
                            "affaire_nge_raw": self._stringify(imported["affaire_nge"]),
                            "affaire_nge_normalized": "",
                            "chantier": self._stringify(imported["chantier"]),
                            "reason": "Numéro d’affaire vide",
                            "target_match_reference": "",
                            "external_match_code": "",
                        }
                    )
                    continue

                target_match = self._find_best_target_match(target_rows, imported_key, exclude_id=int(imported["id"]))
                if target_match is not None:
                    continue

                external_match = self._find_reference_match(external_rows, imported_key)
                rows.append(
                    {
                        "imported_affaire_id": int(imported["id"]),
                        "imported_reference": self._stringify(imported["reference"]),
                        "affaire_nge_raw": self._stringify(imported["affaire_nge"]),
                        "affaire_nge_normalized": imported_key,
                        "chantier": self._stringify(imported["chantier"]),
                        "reason": "Aucun rattachement affaires_rst après normalisation",
                        "target_match_reference": "",
                        "external_match_code": external_match["matched_code_raw"] if external_match else "",
                        "external_match_column": external_match["matched_column"] if external_match else "",
                    }
                )

        return {
            "count": len(rows),
            "rows": rows[:limit],
        }

    def rematch_imported_affaires(self, dry_run: bool = False, limit: int = 500) -> dict[str, Any]:
        self.ensure_import_schema()
        moved_demandes = 0
        merged_affaires = 0
        updated_from_reference_db = 0
        unresolved_rows: list[dict[str, Any]] = []
        actions: list[dict[str, Any]] = []

        with sqlite3.connect(self.paths.target_db_path) as conn:
            conn.row_factory = sqlite3.Row
            imported_affaires = conn.execute(
                """
                SELECT id, reference, affaire_nge, chantier, titulaire, statut, responsable
                FROM affaires_rst
                WHERE COALESCE(statut, '') = 'Importée'
                  AND COALESCE(responsable, '') = 'Import historique'
                ORDER BY id ASC
                """
            ).fetchall()
            external_rows = self._load_reference_affaires()

            for imported in imported_affaires:
                imported_id = int(imported["id"])
                imported_key = self._normalize_affaire_key(self._stringify(imported["affaire_nge"]))
                if not imported_key:
                    unresolved_rows.append(
                        {
                            "imported_affaire_id": imported_id,
                            "imported_reference": self._stringify(imported["reference"]),
                            "affaire_nge_raw": self._stringify(imported["affaire_nge"]),
                            "reason": "Numéro d’affaire vide",
                        }
                    )
                    continue

                current_target_rows = conn.execute(
                    """
                    SELECT id, reference, affaire_nge, chantier, titulaire, statut, responsable
                    FROM affaires_rst
                    ORDER BY id ASC
                    """
                ).fetchall()

                target_match = self._find_best_target_match(current_target_rows, imported_key, exclude_id=imported_id)
                if target_match is not None:
                    target_id = int(target_match["id"])
                    demande_count_row = conn.execute(
                        "SELECT COUNT(*) FROM demandes WHERE affaire_rst_id = ?",
                        (imported_id,),
                    ).fetchone()
                    demande_count = int(demande_count_row[0] or 0)

                    actions.append(
                        {
                            "action": "merge_into_existing_affaire",
                            "source_affaire_reference": self._stringify(imported["reference"]),
                            "source_affaire_id": imported_id,
                            "target_affaire_reference": self._stringify(target_match["reference"]),
                            "target_affaire_id": target_id,
                            "moved_demandes": demande_count,
                            "normalized_affaire_key": imported_key,
                        }
                    )

                    if not dry_run:
                        conn.execute(
                            "UPDATE demandes SET affaire_rst_id = ? WHERE affaire_rst_id = ?",
                            (target_id, imported_id),
                        )
                        remaining_row = conn.execute(
                            "SELECT COUNT(*) FROM demandes WHERE affaire_rst_id = ?",
                            (imported_id,),
                        ).fetchone()
                        remaining = int(remaining_row[0] or 0)
                        if remaining == 0:
                            conn.execute("DELETE FROM affaires_rst WHERE id = ?", (imported_id,))
                        conn.commit()

                    moved_demandes += demande_count
                    merged_affaires += 1
                    continue

                external_match = self._find_reference_match(external_rows, imported_key)
                if external_match:
                    actions.append(
                        {
                            "action": "enrich_imported_affaire_from_reference_db",
                            "source_affaire_reference": self._stringify(imported["reference"]),
                            "source_affaire_id": imported_id,
                            "target_affaire_reference": "",
                            "target_affaire_id": None,
                            "moved_demandes": 0,
                            "normalized_affaire_key": imported_key,
                            "external_match_code": external_match["matched_code_raw"],
                            "external_match_column": external_match["matched_column"],
                        }
                    )
                    if not dry_run:
                        new_affaire_nge = external_match["matched_code_raw"] or self._stringify(imported["affaire_nge"])
                        new_titulaire = external_match["titulaire"] or self._stringify(imported["titulaire"])
                        new_chantier = external_match["libelle"] or self._stringify(imported["chantier"])
                        conn.execute(
                            """
                            UPDATE affaires_rst
                            SET affaire_nge = ?,
                                titulaire = ?,
                                chantier = ?,
                                updated_at = datetime('now')
                            WHERE id = ?
                            """,
                            (
                                new_affaire_nge,
                                new_titulaire,
                                new_chantier,
                                imported_id,
                            ),
                        )
                        conn.commit()
                    updated_from_reference_db += 1
                    continue

                unresolved_rows.append(
                    {
                        "imported_affaire_id": imported_id,
                        "imported_reference": self._stringify(imported["reference"]),
                        "affaire_nge_raw": self._stringify(imported["affaire_nge"]),
                        "normalized_affaire_key": imported_key,
                        "reason": "Aucun match trouvé après normalisation",
                    }
                )

        return {
            "dry_run": dry_run,
            "merged_affaires": merged_affaires,
            "moved_demandes": moved_demandes,
            "updated_from_reference_db": updated_from_reference_db,
            "unresolved_count": len(unresolved_rows),
            "actions": actions[:limit],
            "unresolved_rows": unresolved_rows[:limit],
        }

    def ensure_import_schema(self) -> None:
        self.paths.target_db_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(self.paths.target_db_path) as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS historical_import_batches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    folder_path TEXT NOT NULL,
                    dry_run INTEGER NOT NULL DEFAULT 0,
                    total_candidates INTEGER NOT NULL DEFAULT 0,
                    created_affaires INTEGER NOT NULL DEFAULT 0,
                    created_demandes INTEGER NOT NULL DEFAULT 0,
                    created_echantillons INTEGER NOT NULL DEFAULT 0,
                    created_essais INTEGER NOT NULL DEFAULT 0,
                    status TEXT NOT NULL DEFAULT 'started',
                    notes TEXT NOT NULL DEFAULT '',
                    started_at TEXT NOT NULL DEFAULT (datetime('now')),
                    finished_at TEXT
                );

                CREATE TABLE IF NOT EXISTS historical_import_files (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    batch_id INTEGER NOT NULL REFERENCES historical_import_batches(id) ON DELETE CASCADE,
                    file_name TEXT NOT NULL,
                    file_path TEXT NOT NULL,
                    file_hash TEXT NOT NULL DEFAULT '',
                    sheet_name TEXT NOT NULL DEFAULT '',
                    essai_code TEXT NOT NULL DEFAULT '',
                    sample_local_ref TEXT NOT NULL DEFAULT '',
                    status TEXT NOT NULL DEFAULT '',
                    message TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL DEFAULT (datetime('now'))
                );

                CREATE INDEX IF NOT EXISTS idx_historical_import_files_batch
                    ON historical_import_files(batch_id);
                """
            )
            conn.commit()

    def _scan_supported_files(self, folder_path: Path) -> list[Path]:
        if not folder_path.exists():
            raise FileNotFoundError(f"Folder not found: {folder_path}")
        if not folder_path.is_dir():
            raise NotADirectoryError(f"Not a folder: {folder_path}")
        return sorted(
            [
                file_path
                for file_path in folder_path.rglob("*.xlsx")
                if not file_path.name.startswith("~$")
            ]
        )


    def _parse_workbook(self, file_path: Path) -> dict[str, Any]:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        file_hash = hashlib.sha1(file_path.read_bytes()).hexdigest()
        supported_candidates: list[dict[str, Any]] = []
        skipped_sheets: list[dict[str, Any]] = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            raw_code = self._detect_raw_code(worksheet)
            if not raw_code:
                raw_code = self._detect_code_from_filename(file_path.name)
            essai_code = self._normalize_excel_code(raw_code)
            if essai_code not in SUPPORTED_CODES:
                if essai_code in KNOWN_CODES:
                    reason = f"Known code not yet supported in this V1: {raw_code or essai_code} ({KNOWN_CODES[essai_code]})"
                else:
                    reason = f"Unsupported code: {raw_code or 'empty'}"
                skipped_sheets.append(
                    {
                        "file_name": file_path.name,
                        "sheet_name": sheet_name,
                        "reason": reason,
                    }
                )
                continue

            base = self._build_base_candidate(file_path, file_hash, worksheet, sheet_name, essai_code)
            sheet_candidates = self._parse_supported_sheet(worksheet, base)
            if not sheet_candidates:
                skipped_sheets.append(
                    {
                        "file_name": file_path.name,
                        "sheet_name": sheet_name,
                        "reason": f"Supported code detected but no exploitable line parsed ({essai_code})",
                    }
                )
                continue
            supported_candidates.extend(sheet_candidates)

        return {
            "sheet_count": len(workbook.sheetnames),
            "supported_candidates": supported_candidates,
            "skipped_sheets": skipped_sheets,
        }



    def _build_base_candidate(
        self,
        file_path: Path,
        file_hash: str,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        sheet_name: str,
        essai_code: str,
    ) -> dict[str, Any]:
        target_entity = "intervention" if essai_code in {"DE", "CFE", "PLD", "DF", "PMT", "SC", "SO", "SOL", "FTP"} else "echantillon"
        if essai_code in {"MVA", "DE", "CFE"}:
            import_family = "enrobes"
        elif essai_code in {"PLD", "DF", "PMT", "SC", "SO"}:
            import_family = "terrain"
        elif essai_code == "SOL":
            import_family = "external_lab"
        elif essai_code == "FTP":
            import_family = "documents"
        else:
            import_family = "sols"
        import_mode = "composite" if essai_code in {"ID", "CFE", "SC", "SO", "SOL", "FTP"} else "simple"
        return {
            "file_name": file_path.name,
            "file_path": str(file_path),
            "file_hash": file_hash,
            "sheet_name": sheet_name,
            "essai_code": essai_code,
            "essai_label": SUPPORTED_CODES.get(essai_code, essai_code),
            "import_mode": import_mode,
            "target_entity": target_entity,
            "import_family": import_family,
            "title": self._stringify(worksheet["G2"].value),
            "chrono": self._stringify(worksheet["J5"].value),
            "affaire_nge": self._stringify(worksheet["L5"].value),
            "date_redaction": self._to_iso_date(worksheet["P5"].value),
            "operator": self._label_value(worksheet, "Opérateur :"),
            "date_prelevement": self._to_iso_date(
                self._label_value(
                    worksheet,
                    "Date de(s) prélèvement(s) :",
                    "Date du prélèvement :",
                )
            ),
            "date_essai": self._to_iso_date(
                self._label_value(
                    worksheet,
                    "Date de(s) essai(s) :",
                    "Date des essais : ",
                    "Date de l'essai :",
                )
            ),
            "date_mise_en_oeuvre": self._to_iso_date(
                self._label_value(
                    worksheet,
                    "Date de mise en œuvre :",
                    "Date de mise en œuvre:",
                )
            ),
            "provenance": self._label_value(
                worksheet,
                "Provenance :",
                "Origine du matériau :",
            ),
            "destination": self._label_value(
                worksheet,
                "Destination :",
                "Destination du matériau :",
                "Destination du produit",
            ),
            "nature_materiau": self._label_value(
                worksheet,
                "Nature du matériau :",
                "Nature du produit :",
                "Produit contrôlé :",
            ),
            "couche": self._label_value(worksheet, "Couche :", "Couche de :"),
            "section_controlee": self._label_value(worksheet, "Section contrôlée :"),
            "lieu_fabrication": self._label_value(worksheet, "Lieu de fabrication :"),
            "formula_code": self._label_value(worksheet, "N° formule :", "Code formule : "),
        }


    def _parse_supported_sheet(
        self,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        base: dict[str, Any],
    ) -> list[dict[str, Any]]:
        essai_code = base["essai_code"]
        if essai_code == "WE":
            return self._parse_we_sheet(worksheet, base)
        if essai_code == "ID":
            return self._parse_id_sheet(worksheet, base)
        if essai_code == "LCP":
            return self._parse_lcp_sheet(worksheet, base)
        if essai_code == "IPI - PR":
            return self._parse_proctor_sheet(worksheet, base)
        if essai_code == "IM":
            return self._parse_cbri_sheet(worksheet, base)
        if essai_code == "MVA":
            return self._parse_mva_sheet(worksheet, base)
        if essai_code == "DE":
            return self._parse_de_sheet(worksheet, base)
        if essai_code == "CFE":
            return self._parse_cfe_sheet(worksheet, base)
        if essai_code == "PLD":
            return self._parse_pld_sheet(worksheet, base)
        if essai_code == "DF":
            return self._parse_df_sheet(worksheet, base)
        if essai_code == "PMT":
            return self._parse_pmt_sheet(worksheet, base)
        if essai_code == "SC":
            return self._parse_sc_so_sheet(worksheet, base, "SC")
        if essai_code == "SO":
            return self._parse_sc_so_sheet(worksheet, base, "SO")
        if essai_code == "SOL":
            return self._parse_sol_sheet(worksheet, base)
        if essai_code == "FTP":
            return self._parse_ftp_sheet(worksheet, base)
        return []

    def _parse_we_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        row_index = 36
        while row_index <= worksheet.max_row:
            local_ref = worksheet[f"C{row_index}"].value
            water_value = worksheet[f"N{row_index}"].value
            ipi_value = worksheet[f"U{row_index}"].value

            if local_ref in (None, ""):
                break

            item = dict(base)
            item["sample_local_ref"] = self._stringify(local_ref)
            item["result_payload"] = {
                "teneur_eau_ponderale_percent": self._safe_float(water_value),
                "ipi": self._safe_float(ipi_value),
            }
            rows.append(item)
            row_index += 1
        return rows


    def _parse_id_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        granulometry_sizes = []
        granulometry_values = []

        for column in range(4, 18):
            size_value = worksheet.cell(47, column).value
            pass_value = worksheet.cell(48, column).value
            if size_value not in (None, ""):
                granulometry_sizes.append(size_value)
                granulometry_values.append(pass_value)

        granulometry = {}
        for size_value, pass_value in zip(granulometry_sizes, granulometry_values):
            safe_value = self._safe_float(pass_value)
            if safe_value is not None:
                granulometry[str(size_value)] = safe_value

        local_ref = self._build_id_sample_ref(worksheet, base)

        identification_payload = {
            "passants_percent": granulometry,
            "wn_percent": self._safe_float(worksheet["D57"].value),
            "dmax_mm": self._safe_float(worksheet["F57"].value),
            "ip": self._safe_float(worksheet["H57"].value),
            "ic": self._safe_float(worksheet["J57"].value),
            "vbs": self._safe_float(worksheet["L57"].value),
            "es": self._safe_float(worksheet["N57"].value),
            "ipi": self._safe_float(worksheet["P57"].value),
            "gtr_class": self._stringify(worksheet["W59"].value),
            "gtr_state": self._stringify(worksheet["W60"].value),
        }

        item = dict(base)
        item["sample_local_ref"] = local_ref
        item["result_payload"] = identification_payload
        item["composite_subtests"] = self._build_id_subtests(identification_payload)
        return [item]


    def _parse_lcp_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        local_ref = self._stringify(base["provenance"] or base["sheet_name"])
        item = dict(base)
        item["sample_local_ref"] = local_ref
        item["result_payload"] = {
            "wl": self._safe_float(worksheet["F55"].value),
            "wp": self._safe_float(worksheet["G64"].value),
            "ip": self._safe_float(worksheet["E69"].value),
            "wnat": self._safe_float(worksheet["H73"].value),
        }
        return [item]

    def _parse_proctor_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        local_ref = self._stringify(base["sheet_name"])
        series = []
        for row_index in range(25, 30):
            water_value = worksheet[f"B{row_index}"].value
            density_value = worksheet[f"H{row_index}"].value
            ipi_value = worksheet[f"J{row_index}"].value
            if water_value in (None, "") and density_value in (None, "") and ipi_value in (None, ""):
                continue
            series.append(
                {
                    "w_essai": self._safe_float(water_value),
                    "rho_d": self._safe_float(density_value),
                    "ipi": self._safe_float(ipi_value),
                }
            )

        item = dict(base)
        item["sample_local_ref"] = local_ref
        item["result_payload"] = {
            "natural_water_percent": self._safe_float(worksheet["X24"].value),
            "opn_rho_d": self._safe_float(worksheet["H31"].value),
            "opn_water_percent": self._safe_float(worksheet["H32"].value or worksheet["T32"].value),
            "series": series,
        }
        return [item]


    def _parse_cbri_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        item = dict(base)
        item["sample_local_ref"] = self._stringify(base["sheet_name"]) or self._stringify(base.get("chrono", ""))
        series = []
        for row_index in range(20, min(worksheet.max_row, 80) + 1):
            water_value = self._safe_float(worksheet[f"B{row_index}"].value)
            density_value = self._safe_float(worksheet[f"H{row_index}"].value)
            cbri_value = self._safe_float(worksheet[f"J{row_index}"].value)
            if water_value is None and density_value is None and cbri_value is None:
                continue
            series.append({
                "w_essai": water_value,
                "rho_d": density_value,
                "cbri": cbri_value,
            })

        item["result_payload"] = {
            "cbri_percent": self._first_numeric_in_cells(worksheet, ["J31", "J32", "T31", "T32", "X31", "X32"]),
            "opn_rho_d": self._first_numeric_in_cells(worksheet, ["H31", "H32", "T31", "T32"]),
            "natural_water_percent": self._first_numeric_in_cells(worksheet, ["X24", "X25", "W24", "W25"]),
            "series": series,
        }
        return [item]


    def _parse_mva_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        row_index = 35
        while row_index <= worksheet.max_row:
            local_ref = self._stringify(worksheet[f"B{row_index}"].value)
            dry_mass = worksheet[f"E{row_index}"].value
            if local_ref in (None, ""):
                break
            if dry_mass in (None, ""):
                row_index += 1
                continue

            item = dict(base)
            item["sample_local_ref"] = local_ref
            item["result_payload"] = {
                "water_temperature_c": self._safe_float(worksheet["F24"].value),
                "paraffin_density_kg_m3": self._safe_float(worksheet["S24"].value),
                "water_density_kg_m3": self._safe_float(worksheet["F27"].value),
                "mvr_kg_m3": self._safe_float(worksheet["S27"].value),
                "masse_seche_g": self._safe_float(dry_mass),
                "masse_seche_paraffinee_g": self._safe_float(worksheet[f"H{row_index}"].value),
                "masse_dans_eau_g": self._safe_float(worksheet[f"K{row_index}"].value),
                "masse_volumique_eprouvette_kg_m3": self._safe_float(worksheet[f"N{row_index}"].value),
                "compacite_percent": self._safe_float(worksheet[f"Q{row_index}"].value),
                "vides_percent": self._safe_float(worksheet[f"T{row_index}"].value),
                "hauteur_eprouvette_cm": self._safe_float(worksheet[f"W{row_index}"].value),
                "nature_produit": base.get("nature_materiau", ""),
                "couche": base.get("couche", ""),
                "provenance": base.get("provenance", ""),
            }
            rows.append(item)
            row_index += 1
        return rows


    def _parse_de_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        points = []
        row_index = 29
        while row_index <= worksheet.max_row:
            essai_no = worksheet[f"B{row_index}"].value
            density_value = worksheet[f"I{row_index}"].value
            compacity_value = worksheet[f"L{row_index}"].value
            voids_value = worksheet[f"O{row_index}"].value
            observation = self._stringify(worksheet[f"R{row_index}"].value)
            if essai_no in (None, ""):
                break
            points.append(
                {
                    "essai_no": self._stringify(essai_no),
                    "density_g_cm3": self._safe_float(density_value),
                    "compacite_percent": self._safe_float(compacity_value),
                    "vides_percent": self._safe_float(voids_value),
                    "observation": observation,
                }
            )
            row_index += 1

        if not points:
            return []

        item = dict(base)
        item["sample_local_ref"] = self._stringify(base.get("sheet_name", "")) or self._stringify(base.get("chrono", ""))
        item["intervention_subject"] = " / ".join(
            [value for value in [base.get("section_controlee", ""), base.get("couche", ""), base.get("nature_materiau", "")] if value]
        ) or self._stringify(base.get("sheet_name", ""))
        item["intervention_type"] = "Contrôle densité enrobés"
        item["result_payload"] = {
            "mvre_g_cm3": self._safe_float(worksheet["Q25"].value),
            "points": points,
            "moyenne_density_g_cm3": self._safe_float(worksheet["I51"].value),
            "moyenne_compacite_percent": self._safe_float(worksheet["L51"].value),
            "moyenne_vides_percent": self._safe_float(worksheet["O51"].value),
            "taux_conformes_percent": self._safe_float(worksheet["I52"].value),
            "nature_produit": base.get("nature_materiau", ""),
            "couche": base.get("couche", ""),
            "section_controlee": base.get("section_controlee", ""),
            "lieu_fabrication": base.get("lieu_fabrication", ""),
            "formula_code": base.get("formula_code", ""),
        }
        return [item]


    def _parse_cfe_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        granulometry_sizes = []
        for column in range(5, 19):
            size_value = worksheet.cell(52, column).value
            if size_value not in (None, ""):
                granulometry_sizes.append((openpyxl.utils.get_column_letter(column), size_value))

        rows = []
        row_index = 53
        while row_index <= worksheet.max_row:
            essai_no = worksheet[f"B{row_index}"].value
            if essai_no in (None, "", "Mini", "Moyenne", "Théorique"):
                break
            hour_value = self._stringify(worksheet[f"C{row_index}"].value)
            granulo = {}
            for column_letter, sieve_size in granulometry_sizes:
                pass_value = worksheet[f"{column_letter}{row_index}"].value
                safe_pass = self._safe_float(pass_value)
                if safe_pass is not None:
                    granulo[str(sieve_size)] = safe_pass
            rows.append(
                {
                    "essai_no": self._stringify(essai_no),
                    "hour": hour_value,
                    "granulometrie_passants_percent": granulo,
                    "temperature_c": self._safe_float(worksheet[f"U{row_index}"].value),
                    "teneur_liant_percent": self._safe_float(worksheet[f"W{row_index}"].value),
                    "module_richesse": self._safe_float(worksheet[f"Y{row_index}"].value),
                    "teneur_liant_ext_percent": self._safe_float(worksheet[f"AB{row_index}"].value),
                    "surface_specifique": self._safe_float(worksheet[f"AD{row_index}"].value),
                    "module_richesse_ext": self._safe_float(worksheet[f"AE{row_index}"].value),
                }
            )
            row_index += 1

        if not rows:
            return []

        payload = {
            "rows": rows,
            "moyenne": {
                "temperature_c": self._safe_float(worksheet["U59"].value),
                "teneur_liant_percent": self._safe_float(worksheet["W59"].value),
                "module_richesse": self._safe_float(worksheet["Y59"].value),
                "teneur_liant_ext_percent": self._safe_float(worksheet["AB59"].value),
                "surface_specifique": self._safe_float(worksheet["AD59"].value),
                "module_richesse_ext": self._safe_float(worksheet["AE59"].value),
            },
            "theorique": {
                "teneur_liant_percent": self._safe_float(worksheet["W60"].value),
                "module_richesse": self._safe_float(worksheet["Y60"].value),
                "teneur_liant_ext_percent": self._safe_float(worksheet["AB60"].value),
                "surface_specifique": self._safe_float(worksheet["AD60"].value),
                "module_richesse_ext": self._safe_float(worksheet["AE60"].value),
            },
            "thresholds": {
                "teneur_liant_min_percent": self._safe_float(worksheet["W62"].value),
                "teneur_liant_max_percent": self._safe_float(worksheet["W61"].value),
                "module_richesse_min": self._safe_float(worksheet["Y58"].value),
                "module_richesse_rule": self._stringify(worksheet["AE61"].value),
            },
            "appellation_europeenne": self._stringify(worksheet["G13"].value),
            "appellation_francaise": self._stringify(worksheet["G14"].value),
            "destination": base.get("destination", ""),
            "formula_code": base.get("formula_code", ""),
            "couche": base.get("couche", ""),
            "lieu_fabrication": base.get("lieu_fabrication", ""),
        }

        item = dict(base)
        item["sample_local_ref"] = self._stringify(base.get("sheet_name", "")) or self._stringify(base.get("chrono", ""))
        item["intervention_subject"] = " / ".join(
            [value for value in [base.get("destination", ""), base.get("formula_code", ""), base.get("couche", "")] if value]
        ) or self._stringify(base.get("sheet_name", ""))
        item["intervention_type"] = "Contrôle fabrication enrobés"
        item["result_payload"] = payload
        item["composite_subtests"] = self._build_cfe_subtests(payload)
        return [item]

    def _parse_pld_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        points = []
        row_index = 25
        while row_index <= worksheet.max_row:
            essai_no = worksheet[f"B{row_index}"].value
            if essai_no in (None, "", "Moyenne", "Valeur mini", "Valeur maxi"):
                break
            points.append(
                {
                    "point_no": self._stringify(essai_no),
                    "localisation": self._stringify(worksheet[f"C{row_index}"].value),
                    "ev2_mpa": self._safe_float(worksheet[f"H{row_index}"].value),
                    "observation": self._stringify(worksheet[f"K{row_index}"].value),
                }
            )
            row_index += 1

        if not points:
            return []

        item = dict(base)
        item["sample_local_ref"] = self._stringify(base.get("sheet_name", "")) or self._stringify(base.get("chrono", ""))
        item["intervention_subject"] = " / ".join(
            [value for value in [self._label_value(worksheet, "Partie de l'ouvrage :"), base.get("section_controlee", ""), base.get("nature_materiau", "")] if value]
        ) or self._stringify(base.get("sheet_name", ""))
        item["intervention_type"] = "Portances dynaplaque"
        item["result_payload"] = {
            "points": points,
            "diametre_plaque_mm": self._safe_float(worksheet["R21"].value),
            "moyenne_ev2_mpa": self._safe_float(worksheet["H47"].value),
            "valeur_min_mpa": self._safe_float(worksheet["H48"].value),
            "valeur_max_mpa": self._safe_float(worksheet["H49"].value),
            "taux_conformes_percent": self._safe_float(worksheet["K47"].value),
            "conclusion": self._stringify(worksheet["A54"].value or worksheet["A53"].value),
            "partie_ouvrage": self._label_value(worksheet, "Partie de l'ouvrage :"),
            "nature_materiau": base.get("nature_materiau", ""),
        }
        return [item]

    def _parse_df_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        rows = self._extract_generic_result_rows(
            worksheet,
            start_row=20,
            stop_markers={"Moyenne", "Valeur mini", "Valeur maxi", "Conclusion", "COMMENTAIRES"},
            max_rows=80,
        )
        if not rows:
            return []
        item = dict(base)
        item["sample_local_ref"] = self._stringify(base.get("sheet_name", "")) or self._stringify(base.get("chrono", ""))
        item["intervention_subject"] = " / ".join([value for value in [base.get("section_controlee", ""), base.get("couche", ""), base.get("nature_materiau", "")] if value]) or self._stringify(base.get("sheet_name", ""))
        item["intervention_type"] = "Déflexion"
        item["result_payload"] = {
            "rows": rows,
            "header_snapshot": self._sheet_snapshot(worksheet, max_row=18),
            "nature_materiau": base.get("nature_materiau", ""),
            "section_controlee": base.get("section_controlee", ""),
        }
        return [item]

    def _parse_pmt_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        rows = self._extract_generic_result_rows(
            worksheet,
            start_row=20,
            stop_markers={"Moyenne", "Valeur mini", "Valeur maxi", "Conclusion", "COMMENTAIRES"},
            max_rows=80,
        )
        if not rows:
            return []
        item = dict(base)
        item["sample_local_ref"] = self._stringify(base.get("sheet_name", "")) or self._stringify(base.get("chrono", ""))
        item["intervention_subject"] = " / ".join([value for value in [base.get("section_controlee", ""), base.get("couche", ""), base.get("nature_materiau", "")] if value]) or self._stringify(base.get("sheet_name", ""))
        item["intervention_type"] = "Macrotexture PMT"
        item["result_payload"] = {
            "rows": rows,
            "header_snapshot": self._sheet_snapshot(worksheet, max_row=18),
            "nature_materiau": base.get("nature_materiau", ""),
            "section_controlee": base.get("section_controlee", ""),
        }
        return [item]

    def _parse_sol_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        item = dict(base)
        item["sample_local_ref"] = self._stringify(base.get("sheet_name", "")) or self._stringify(base.get("chrono", ""))
        item["intervention_subject"] = base.get("title", "") or self._guess_chantier_from_filename(base["file_name"])
        item["intervention_type"] = "Analyses pollution - laboratoire externe"
        item["result_payload"] = {
            "title": base.get("title", ""),
            "operator": base.get("operator", ""),
            "provenance": base.get("provenance", ""),
            "destination": base.get("destination", ""),
            "nature_materiau": base.get("nature_materiau", ""),
            "header_snapshot": self._extract_header_snapshot(worksheet),
        }
        return [item]


    def _parse_ftp_sheet(self, worksheet, base: dict[str, Any]) -> list[dict[str, Any]]:
        item = dict(base)
        item["sample_local_ref"] = self._stringify(base.get("sheet_name", "")) or self._stringify(base.get("chrono", "")) or "FTP"
        item["intervention_subject"] = base.get("title", "") or self._guess_chantier_from_filename(base["file_name"])
        item["intervention_type"] = "Fiche technique produit"
        item["result_payload"] = {
            "title": base.get("title", ""),
            "nature_materiau": base.get("nature_materiau", ""),
            "formula_code": base.get("formula_code", ""),
            "header_snapshot": self._extract_header_snapshot(worksheet),
        }
        return [item]


    def _parse_sc_so_sheet(self, worksheet, base: dict[str, Any], code: str) -> list[dict[str, Any]]:
        rows = self._extract_generic_result_rows(
            worksheet,
            start_row=20,
            stop_markers={"Conclusion", "COMMENTAIRES", "Commentaires"},
            max_rows=120,
        )
        item = dict(base)
        item["sample_local_ref"] = self._stringify(base.get("sheet_name", "")) or self._stringify(base.get("chrono", ""))
        item["intervention_subject"] = " / ".join([value for value in [base.get("section_controlee", ""), base.get("couche", ""), base.get("nature_materiau", "")] if value]) or self._stringify(base.get("sheet_name", ""))
        item["intervention_type"] = "Coupe de sondage carotté" if code == "SC" else "Coupes de sondages"
        payload = {
            "rows": rows,
            "header_snapshot": self._sheet_snapshot(worksheet, max_row=20),
            "nature_materiau": base.get("nature_materiau", ""),
            "section_controlee": base.get("section_controlee", ""),
            "destination": base.get("destination", ""),
        }
        item["result_payload"] = payload
        item["composite_subtests"] = []
        return [item]

    def _extract_generic_result_rows(
        self,
        worksheet,
        start_row: int,
        stop_markers: set[str],
        max_rows: int = 80,
    ) -> list[dict[str, Any]]:
        headers = None
        data_rows: list[dict[str, Any]] = []
        max_column = min(worksheet.max_column, 26)
        scanned = 0
        for row_index in range(start_row, min(worksheet.max_row, start_row + max_rows) + 1):
            scanned += 1
            values = [worksheet.cell(row_index, col).value for col in range(1, max_column + 1)]
            texts = [self._stringify(v) for v in values]
            non_empty = [t for t in texts if t]
            if not non_empty:
                continue
            first_token = non_empty[0]
            if first_token in stop_markers:
                break

            if headers is None and self._looks_like_header_row(texts):
                headers = [t if t else f"col_{idx+1}" for idx, t in enumerate(texts)]
                continue

            if headers is None:
                continue

            row_data = {}
            has_value = False
            for idx, header in enumerate(headers):
                value = values[idx] if idx < len(values) else None
                text = self._stringify(value)
                if text:
                    has_value = True
                row_data[header] = value if value not in (None, "") else text
            if has_value:
                data_rows.append(row_data)
        return data_rows

    def _looks_like_header_row(self, texts: list[str]) -> bool:
        joined = " ".join(texts).lower()
        header_keywords = [
            "localisation",
            "observation",
            "module",
            "macrotexture",
            "déflex",
            "deflex",
            "point",
            "n°",
            "numero",
            "position",
            "chaine",
        ]
        score = sum(1 for keyword in header_keywords if keyword in joined)
        return score >= 1 and sum(1 for t in texts if t) >= 2

    def _sheet_snapshot(self, worksheet, max_row: int = 20, max_col: int = 12) -> list[list[str]]:
        snapshot: list[list[str]] = []
        for row_index in range(1, min(worksheet.max_row, max_row) + 1):
            row = [self._stringify(worksheet.cell(row_index, col).value) for col in range(1, min(worksheet.max_column, max_col) + 1)]
            if any(cell for cell in row):
                snapshot.append(row)
        return snapshot


    def _find_or_create_affaire(self, conn: sqlite3.Connection, candidate: dict[str, Any]) -> tuple[int, bool]:
        affaire_key = self._normalize_affaire_key(candidate["affaire_nge"])
        if not affaire_key:
            raise ValueError(f"Missing N° d'affaire in {candidate['file_name']} / {candidate['sheet_name']}")

        target_rows = conn.execute(
            """
            SELECT id, reference, affaire_nge, chantier, titulaire, statut, responsable
            FROM affaires_rst
            ORDER BY id ASC
            """
        ).fetchall()
        target_match = self._find_best_target_match(target_rows, affaire_key)
        if target_match is not None:
            return int(target_match["id"]), False

        source_row = self._find_reference_match(self._load_reference_affaires(), affaire_key)

        year_value = self._extract_year(candidate["date_essai"] or candidate["date_redaction"]) or 2026
        reference, numero = self._next_affaire_reference(conn, year_value)

        chantier = ""
        titulaire = ""
        affaire_nge = candidate["affaire_nge"]
        if source_row:
            chantier = self._stringify(source_row["libelle"])
            titulaire = self._stringify(source_row["titulaire"])
            affaire_nge = source_row["matched_code_raw"] or affaire_nge
        if not chantier:
            chantier = self._guess_chantier_from_filename(candidate["file_name"])

        cursor = conn.execute(
            """
            INSERT INTO affaires_rst (
                reference,
                annee,
                region,
                numero,
                client,
                titulaire,
                chantier,
                affaire_nge,
                date_ouverture,
                statut,
                responsable,
                created_at,
                updated_at
            ) VALUES (?, ?, 'RA', ?, 'Non communiqué', ?, ?, ?, ?, 'Importée', 'Import historique', datetime('now'), datetime('now'))
            """,
            (
                reference,
                year_value,
                numero,
                titulaire,
                chantier,
                affaire_nge,
                candidate["date_redaction"] or candidate["date_essai"] or datetime.now().date().isoformat(),
            ),
        )
        conn.commit()
        return int(cursor.lastrowid), True

    def _find_or_create_demande(self, conn: sqlite3.Connection, affaire_id: int, candidate: dict[str, Any]) -> tuple[int, bool]:
        year_value = self._extract_year(candidate["date_essai"] or candidate["date_redaction"]) or 2026
        nature, description, context_label, enabled_modules = self._demande_profile_for_candidate(candidate)
        existing = conn.execute(
            """
            SELECT id
            FROM demandes
            WHERE affaire_rst_id = ?
              AND labo_code = 'SP'
              AND nature = ?
              AND annee = ?
            ORDER BY id ASC
            LIMIT 1
            """,
            (affaire_id, nature, year_value),
        ).fetchone()
        if existing:
            return int(existing["id"]), False

        reference, numero = self._next_demande_reference(conn, year_value, "SP")
        cursor = conn.execute(
            """
            INSERT INTO demandes (
                reference,
                annee,
                labo_code,
                numero,
                affaire_rst_id,
                type_mission,
                nature,
                description,
                observations,
                demandeur,
                date_reception,
                statut,
                priorite,
                a_revoir,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, 'Laboratoire', ?, ?, ?, 'Import historique', ?, 'Importée', 'Normale', 0, datetime('now'), datetime('now'))
            """,
            (
                reference,
                year_value,
                "SP",
                numero,
                affaire_id,
                nature,
                description,
                f"Créée automatiquement depuis {candidate['file_name']}",
                candidate["date_essai"] or candidate["date_redaction"] or datetime.now().date().isoformat(),
            ),
        )
        demande_id = int(cursor.lastrowid)

        conn.execute(
            """
            INSERT OR IGNORE INTO demande_preparations (
                demande_id,
                phase_operation,
                contexte_operationnel,
                objectifs,
                commentaires,
                created_at,
                updated_at
            ) VALUES (?, 'Historique', ?, 'Rétro-intégration des essais réalisés', ?, datetime('now'), datetime('now'))
            """,
            (
                demande_id,
                context_label,
                f"Demande créée automatiquement depuis {candidate['file_name']}",
            ),
        )

        for module_code in enabled_modules:
            conn.execute(
                """
                INSERT OR IGNORE INTO demande_enabled_modules (
                    demande_id,
                    module_code,
                    is_enabled,
                    created_at,
                    updated_at
                ) VALUES (?, ?, 1, datetime('now'), datetime('now'))
                """,
                (demande_id, module_code),
            )
        conn.commit()
        return demande_id, True

    def _demande_profile_for_candidate(self, candidate: dict[str, Any]) -> tuple[str, str, str, tuple[str, ...]]:
        family = candidate.get("import_family", "sols")
        if family == "terrain":
            return (
                "Import historique essais terrain",
                f"Import historique essais terrain {candidate.get('date_essai', '')[:4] or '2026'}",
                "Import historique terrain",
                ("interventions", "documents"),
            )
        if family == "external_lab":
            return (
                "Import historique laboratoire externe",
                f"Import historique laboratoire externe {candidate.get('date_essai', '')[:4] or '2026'}",
                "Import historique laboratoire externe",
                ("interventions", "documents"),
            )
        if family == "documents":
            return (
                "Import historique documents techniques",
                f"Import historique documents techniques {candidate.get('date_essai', '')[:4] or '2026'}",
                "Import historique documents techniques",
                ("interventions", "documents"),
            )
        if family == "enrobes":
            if candidate.get("target_entity") == "intervention":
                return (
                    "Import historique essais enrobés",
                    f"Import historique essais enrobés {candidate.get('date_essai', '')[:4] or '2026'}",
                    "Import historique enrobés",
                    ("interventions", "documents"),
                )
            return (
                "Import historique essais enrobés",
                f"Import historique essais enrobés {candidate.get('date_essai', '')[:4] or '2026'}",
                "Import historique enrobés",
                ("echantillons", "essais_laboratoire", "documents"),
            )
        return (
            "Import historique essais labo",
            f"Import historique essais labo sols {candidate.get('date_essai', '')[:4] or '2026'}",
            "Import historique labo sols",
            ("echantillons", "essais_laboratoire", "documents"),
        )

    def _find_or_create_echantillon(self, conn: sqlite3.Connection, demande_id: int, candidate: dict[str, Any]) -> tuple[int, bool]:
        local_ref = candidate["sample_local_ref"].strip() or candidate["sheet_name"].strip()
        existing = conn.execute(
            """
            SELECT id
            FROM echantillons
            WHERE demande_id = ?
              AND COALESCE(designation, '') = ?
            ORDER BY id ASC
            LIMIT 1
            """,
            (demande_id, local_ref),
        ).fetchone()
        if existing:
            return int(existing["id"]), False

        year_value = self._extract_year(candidate["date_essai"] or candidate["date_prelevement"] or candidate["date_redaction"]) or 2026
        reference, numero = self._next_echantillon_reference(conn, year_value, "SP")
        localisation = " / ".join(
            [value for value in [candidate.get("provenance", ""), candidate.get("destination", "")] if value]
        )

        observations_payload = {
            "source_file": candidate["file_name"],
            "sheet_name": candidate["sheet_name"],
            "sample_local_ref": local_ref,
            "nature_materiau": candidate.get("nature_materiau", ""),
            "import_mode": candidate.get("import_mode", "simple"),
            "source_essai_code": candidate.get("essai_code", ""),
        }

        cursor = conn.execute(
            """
            INSERT INTO echantillons (
                reference,
                annee,
                labo_code,
                numero,
                demande_id,
                designation,
                date_prelevement,
                localisation,
                statut,
                date_reception_labo,
                observations,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Importé', ?, ?, datetime('now'), datetime('now'))
            """,
            (
                reference,
                year_value,
                "SP",
                numero,
                demande_id,
                local_ref,
                candidate["date_prelevement"] or candidate["date_essai"],
                localisation,
                candidate["date_essai"] or candidate["date_redaction"],
                json.dumps(observations_payload, ensure_ascii=False),
            ),
        )
        conn.commit()
        return int(cursor.lastrowid), True



    def _find_or_create_intervention(self, conn: sqlite3.Connection, demande_id: int, candidate: dict[str, Any]) -> tuple[int, bool]:
        signature = f"SRC_HASH={candidate['file_hash']}|SHEET={candidate['sheet_name']}|CODE={candidate['essai_code']}"
        existing = conn.execute(
            """
            SELECT id
            FROM interventions
            WHERE demande_id = ?
              AND COALESCE(observations, '') LIKE ?
            LIMIT 1
            """,
            (demande_id, f"%{signature}%"),
        ).fetchone()
        if existing:
            return int(existing["id"]), False

        year_value = self._extract_year(candidate["date_essai"] or candidate.get("date_mise_en_oeuvre", "") or candidate["date_redaction"]) or 2026
        reference, numero = self._next_intervention_reference(conn, year_value, "SP")
        observations = {
            "source_file": candidate["file_name"],
            "sheet_name": candidate["sheet_name"],
            "signature": signature,
            "essai_code": candidate.get("essai_code", ""),
            "essai_label": candidate.get("essai_label", ""),
            "payload": candidate.get("result_payload", {}),
            "import_mode": candidate.get("import_mode", "simple"),
            "composite_subtests": candidate.get("composite_subtests", []),
        }
        cursor = conn.execute(
            """
            INSERT INTO interventions (
                reference,
                annee,
                labo_code,
                numero,
                demande_id,
                type_intervention,
                sujet,
                date_intervention,
                geotechnicien,
                technicien,
                observations,
                statut,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', ?, ?, 'Importée', datetime('now'), datetime('now'))
            """,
            (
                reference,
                year_value,
                "SP",
                numero,
                demande_id,
                candidate.get("intervention_type", candidate.get("essai_label", "Intervention historique")),
                candidate.get("intervention_subject", candidate.get("sheet_name", "")),
                candidate.get("date_mise_en_oeuvre", "") or candidate.get("date_essai", "") or candidate.get("date_redaction", "") or datetime.now().date().isoformat(),
                candidate.get("operator", ""),
                json.dumps(observations, ensure_ascii=False),
            ),
        )
        conn.commit()
        return int(cursor.lastrowid), True


    def _store_intervention_payload_if_missing(self, conn: sqlite3.Connection, intervention_id: int, candidate: dict[str, Any]) -> int:
        return 0


    def _create_essai_if_missing(self, conn: sqlite3.Connection, echantillon_id: int, candidate: dict[str, Any]) -> int:
        if candidate.get("import_mode") == "composite" and candidate.get("essai_code") == "ID":
            return self._create_id_composite_essais(conn, echantillon_id, candidate)
        return self._create_simple_essai_if_missing(conn, echantillon_id, candidate)

    def _create_simple_essai_if_missing(self, conn: sqlite3.Connection, echantillon_id: int, candidate: dict[str, Any]) -> int:
        signature = f"SRC_HASH={candidate['file_hash']}|SHEET={candidate['sheet_name']}|REF={candidate['sample_local_ref']}|CODE={candidate['essai_code']}"
        existing = conn.execute(
            """
            SELECT id
            FROM essais
            WHERE echantillon_id = ?
              AND type_essai = ?
              AND COALESCE(observations, '') LIKE ?
            LIMIT 1
            """,
            (echantillon_id, candidate["essai_label"], f"%{signature}%"),
        ).fetchone()
        if existing:
            return 0

        observations = {
            "source_file": candidate["file_name"],
            "sheet_name": candidate["sheet_name"],
            "signature": signature,
            "destination": candidate.get("destination", ""),
            "provenance": candidate.get("provenance", ""),
            "import_mode": candidate.get("import_mode", "simple"),
        }

        conn.execute(
            """
            INSERT INTO essais (
                echantillon_id,
                type_essai,
                norme,
                statut,
                date_debut,
                date_fin,
                resultats,
                operateur,
                observations,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, 'Importé', ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
            """,
            (
                echantillon_id,
                candidate["essai_label"],
                candidate["title"],
                candidate["date_essai"] or candidate["date_redaction"],
                candidate["date_essai"] or candidate["date_redaction"],
                json.dumps(candidate["result_payload"], ensure_ascii=False),
                candidate.get("operator", ""),
                json.dumps(observations, ensure_ascii=False),
            ),
        )
        conn.commit()
        return 1

    def _create_id_composite_essais(self, conn: sqlite3.Connection, echantillon_id: int, candidate: dict[str, Any]) -> int:
        created_count = 0
        subtests = candidate.get("composite_subtests", [])
        if not subtests:
            return self._create_simple_essai_if_missing(conn, echantillon_id, candidate)

        for subtest in subtests:
            signature = (
                f"SRC_HASH={candidate['file_hash']}|SHEET={candidate['sheet_name']}|REF={candidate['sample_local_ref']}"
                f"|CODE={candidate['essai_code']}|SUB={subtest['subcode']}"
            )
            existing = conn.execute(
                """
                SELECT id
                FROM essais
                WHERE echantillon_id = ?
                  AND type_essai = ?
                  AND COALESCE(observations, '') LIKE ?
                LIMIT 1
                """,
                (echantillon_id, subtest["type_essai"], f"%{signature}%"),
            ).fetchone()
            if existing:
                continue

            observations = {
                "source_file": candidate["file_name"],
                "sheet_name": candidate["sheet_name"],
                "signature": signature,
                "destination": candidate.get("destination", ""),
                "provenance": candidate.get("provenance", ""),
                "import_mode": "composite",
                "parent_essai_code": candidate.get("essai_code", ""),
                "parent_essai_label": candidate.get("essai_label", ""),
                "subcode": subtest["subcode"],
            }

            conn.execute(
                """
                INSERT INTO essais (
                    echantillon_id,
                    type_essai,
                    norme,
                    statut,
                    date_debut,
                    date_fin,
                    resultats,
                    operateur,
                    observations,
                    created_at,
                    updated_at
                ) VALUES (?, ?, ?, 'Importé', ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
                """,
                (
                    echantillon_id,
                    subtest["type_essai"],
                    candidate["title"],
                    candidate["date_essai"] or candidate["date_redaction"],
                    candidate["date_essai"] or candidate["date_redaction"],
                    json.dumps(subtest["result_payload"], ensure_ascii=False),
                    candidate.get("operator", ""),
                    json.dumps(observations, ensure_ascii=False),
                ),
            )
            created_count += 1

        conn.commit()
        return created_count


    def _next_affaire_reference(self, conn: sqlite3.Connection, year_value: int) -> tuple[str, int]:
        row = conn.execute(
            "SELECT COALESCE(MAX(numero), 0) FROM affaires_rst WHERE annee = ?",
            (year_value,),
        ).fetchone()
        numero = int(row[0]) + 1
        return f"{year_value}-RA-{numero:03d}", numero

    def _next_demande_reference(self, conn: sqlite3.Connection, year_value: int, labo_code: str) -> tuple[str, int]:
        row = conn.execute(
            "SELECT COALESCE(MAX(numero), 0) FROM demandes WHERE annee = ? AND labo_code = ?",
            (year_value, labo_code),
        ).fetchone()
        numero = int(row[0]) + 1
        return f"{year_value}-{labo_code}-D{numero:04d}", numero

    def _next_echantillon_reference(self, conn: sqlite3.Connection, year_value: int, labo_code: str) -> tuple[str, int]:
        row = conn.execute(
            "SELECT COALESCE(MAX(numero), 0) FROM echantillons WHERE annee = ? AND labo_code = ?",
            (year_value, labo_code),
        ).fetchone()
        numero = int(row[0]) + 1
        return f"{year_value}-{labo_code}-E{numero:04d}", numero

    def _next_intervention_reference(self, conn: sqlite3.Connection, year_value: int, labo_code: str) -> tuple[str, int]:
        row = conn.execute(
            "SELECT COALESCE(MAX(numero), 0) FROM interventions WHERE annee = ? AND labo_code = ?",
            (year_value, labo_code),
        ).fetchone()
        numero = int(row[0]) + 1
        return f"{year_value}-{labo_code}-I{numero:04d}", numero

    def _create_batch(self, conn: sqlite3.Connection, folder_path: Path, total_candidates: int, dry_run: bool) -> int:
        cursor = conn.execute(
            """
            INSERT INTO historical_import_batches (
                folder_path,
                dry_run,
                total_candidates,
                status
            ) VALUES (?, ?, ?, 'started')
            """,
            (str(folder_path), 1 if dry_run else 0, total_candidates),
        )
        conn.commit()
        return int(cursor.lastrowid)

    def _finalize_batch(
        self,
        conn: sqlite3.Connection,
        batch_id: int,
        created_affaires: int,
        created_demandes: int,
        created_echantillons: int,
        created_essais: int,
        status: str,
        notes: str,
    ) -> None:
        conn.execute(
            """
            UPDATE historical_import_batches
            SET created_affaires = ?,
                created_demandes = ?,
                created_echantillons = ?,
                created_essais = ?,
                status = ?,
                notes = ?,
                finished_at = datetime('now')
            WHERE id = ?
            """,
            (
                created_affaires,
                created_demandes,
                created_echantillons,
                created_essais,
                status,
                notes,
                batch_id,
            ),
        )
        conn.commit()

    def _create_file_log(
        self,
        conn: sqlite3.Connection,
        batch_id: int,
        candidate: dict[str, Any],
        status: str,
        message: str,
    ) -> None:
        conn.execute(
            """
            INSERT INTO historical_import_files (
                batch_id,
                file_name,
                file_path,
                file_hash,
                sheet_name,
                essai_code,
                sample_local_ref,
                status,
                message
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                candidate["file_name"],
                candidate["file_path"],
                candidate["file_hash"],
                candidate["sheet_name"],
                candidate["essai_code"],
                candidate["sample_local_ref"],
                status,
                message,
            ),
        )
        conn.commit()

    def _load_reference_affaires(self) -> list[dict[str, Any]]:
        if not self.paths.affaires_db_path.exists():
            return []
        with sqlite3.connect(self.paths.affaires_db_path) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT *
                FROM affaires
                """
            ).fetchall()
        return [dict(row) for row in rows]

    def _find_reference_match(self, rows: list[dict[str, Any]], affaire_key: str) -> dict[str, Any] | None:
        if not affaire_key:
            return None
        for row in rows:
            for column_name in REFERENCE_AFFAIRE_COLUMNS:
                raw_value = self._stringify(row.get(column_name, ""))
                if not raw_value:
                    continue
                if self._normalize_affaire_key(raw_value) == affaire_key:
                    return {
                        "row_id": self._stringify(row.get("id", "")),
                        "libelle": self._stringify(row.get("libellé", "")),
                        "titulaire": self._stringify(row.get("titulaire", "")),
                        "matched_column": column_name,
                        "matched_code_raw": raw_value,
                    }
        return None

    def _find_best_target_match(
        self,
        rows: list[sqlite3.Row | dict[str, Any]],
        affaire_key: str,
        exclude_id: int | None = None,
    ) -> sqlite3.Row | dict[str, Any] | None:
        matches: list[sqlite3.Row | dict[str, Any]] = []
        for row in rows:
            row_id = int(row["id"])
            if exclude_id is not None and row_id == exclude_id:
                continue
            normalized = self._normalize_affaire_key(self._stringify(row["affaire_nge"]))
            if normalized == affaire_key:
                matches.append(row)
        if not matches:
            return None

        def rank(item: sqlite3.Row | dict[str, Any]) -> tuple[int, int]:
            is_imported = 1 if (
                self._stringify(item["statut"]) == "Importée"
                and self._stringify(item["responsable"]) == "Import historique"
            ) else 0
            return (is_imported, int(item["id"]))

        matches.sort(key=rank)
        return matches[0]


    def _build_id_sample_ref(self, worksheet, base: dict[str, Any]) -> str:
        explicit_local_ref = self._stringify(worksheet["B48"].value or worksheet["B57"].value)
        if explicit_local_ref and explicit_local_ref.lower() != "n° ech.":
            return explicit_local_ref

        chrono = self._stringify(base.get("chrono", ""))
        if chrono:
            return f"ID-{chrono}"

        sheet_name = self._stringify(base.get("sheet_name", ""))
        if sheet_name:
            return f"ID-{sheet_name}"

        return "ID-SANS-REF"

    def _build_id_subtests(self, payload: dict[str, Any]) -> list[dict[str, Any]]:
        subtests: list[dict[str, Any]] = []

        granulometry = payload.get("passants_percent") or {}
        if granulometry:
            subtests.append(
                {
                    "subcode": "ID-GRANULO",
                    "type_essai": "Granulométrie d'identification",
                    "result_payload": {
                        "passants_percent": granulometry,
                    },
                }
            )

        parameters = {
            "wn_percent": payload.get("wn_percent"),
            "dmax_mm": payload.get("dmax_mm"),
            "ip": payload.get("ip"),
            "ic": payload.get("ic"),
            "vbs": payload.get("vbs"),
            "es": payload.get("es"),
            "ipi": payload.get("ipi"),
        }
        if any(value not in (None, "", {}) for value in parameters.values()):
            subtests.append(
                {
                    "subcode": "ID-PARAM",
                    "type_essai": "Paramètres d'identification",
                    "result_payload": parameters,
                }
            )

        gtr_payload = {
            "gtr_class": payload.get("gtr_class", ""),
            "gtr_state": payload.get("gtr_state", ""),
        }
        if any(value not in (None, "") for value in gtr_payload.values()):
            subtests.append(
                {
                    "subcode": "ID-GTR",
                    "type_essai": "Classification GTR",
                    "result_payload": gtr_payload,
                }
            )

        return subtests

    def _build_cfe_subtests(self, payload: dict[str, Any]) -> list[dict[str, Any]]:
        subtests: list[dict[str, Any]] = []

        first_row = (payload.get("rows") or [{}])[0] if payload.get("rows") else {}
        granulometry = first_row.get("granulometrie_passants_percent") or {}
        if granulometry:
            subtests.append(
                {
                    "subcode": "CFE-GRANULO",
                    "type_essai": "Granulométrie enrobés",
                    "result_payload": {"passants_percent": granulometry},
                }
            )

        binder_payload = {
            "teneur_liant_percent": first_row.get("teneur_liant_percent"),
            "teneur_liant_ext_percent": first_row.get("teneur_liant_ext_percent"),
            "module_richesse": first_row.get("module_richesse"),
            "module_richesse_ext": first_row.get("module_richesse_ext"),
            "surface_specifique": first_row.get("surface_specifique"),
        }
        if any(value not in (None, "", {}) for value in binder_payload.values()):
            subtests.append(
                {
                    "subcode": "CFE-LIANT",
                    "type_essai": "Teneur en liant enrobés",
                    "result_payload": binder_payload,
                }
            )

        temp_value = first_row.get("temperature_c")
        if temp_value is not None:
            subtests.append(
                {
                    "subcode": "CFE-TEMP",
                    "type_essai": "Température enrobés",
                    "result_payload": {"temperature_c": temp_value},
                }
            )

        return subtests


    def _detect_code_from_filename(self, file_name: str) -> str:
        extracted = self._extract_code_from_text(file_name)
        if extracted:
            return extracted
        inferred = self._infer_code_from_keywords(file_name)
        if inferred:
            return inferred
        return ""


    def _detect_raw_code(self, worksheet) -> str:
        candidate_cells = ["G5", "F5", "H5", "E5", "G4", "F4", "H4"]
        for cell_ref in candidate_cells:
            value = self._stringify(worksheet[cell_ref].value)
            extracted = self._extract_code_from_text(value)
            if extracted:
                return extracted

        header_candidates: list[str] = []
        for row in range(1, min(12, worksheet.max_row) + 1):
            for column in range(1, min(26, worksheet.max_column) + 1):
                value = self._stringify(worksheet.cell(row, column).value)
                if not value:
                    continue
                header_candidates.append(value)
                extracted = self._extract_code_from_text(value)
                if extracted:
                    return extracted

        title_zone_candidates: list[str] = []
        for row in range(1, min(6, worksheet.max_row) + 1):
            for column in range(21, min(26, worksheet.max_column) + 1):
                value = self._stringify(worksheet.cell(row, column).value)
                if value:
                    title_zone_candidates.append(value)

        for value in header_candidates + title_zone_candidates:
            extracted = self._infer_code_from_keywords(value)
            if extracted:
                return extracted
        return ""

    def _label_value(self, worksheet, *labels: str) -> str:
        normalized_labels = {self._normalize_text(label) for label in labels}
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                if self._normalize_text(value) in normalized_labels:
                    for offset in range(1, 16):
                        candidate = worksheet.cell(cell.row, cell.column + offset).value
                        if candidate not in (None, ""):
                            return self._stringify(candidate)
        return ""

    def _normalize_excel_code(self, value: str) -> str:
        text = self._stringify(value).strip()
        if not text:
            return ""
        extracted = self._extract_code_from_text(text)
        if extracted:
            return extracted
        inferred = self._infer_code_from_keywords(text)
        if inferred:
            return inferred
        return text

    def _extract_code_from_text(self, value: str) -> str:
        text = self._stringify(value).strip()
        if not text:
            return ""

        compact = re.sub(r"\s+", " ", text).strip()
        upper = compact.upper()
        if upper in KNOWN_CODES:
            return upper
        if upper == "MVE":
            return "MVA"
        if upper == "CBRI":
            return "IM"

        for code, pattern in CODE_TOKEN_PATTERNS:
            if pattern.search(compact):
                return code
        return ""

    def _infer_code_from_keywords(self, value: str) -> str:
        text = self._stringify(value).strip()
        if not text:
            return ""
        for code, pattern in TITLE_KEYWORD_PATTERNS:
            if pattern.search(text):
                return code
        return ""

    def _first_numeric_in_cells(self, worksheet, cell_refs: list[str]) -> float | None:
        for cell_ref in cell_refs:
            if re.fullmatch(r"[A-Z]+\d+", cell_ref):
                value = self._safe_float(worksheet[cell_ref].value)
                if value is not None:
                    return value
        return None


    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        return str(value).strip()

    def _to_iso_date(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if hasattr(value, "isoformat") and not isinstance(value, str):
            try:
                return value.isoformat()
            except Exception:
                pass
        text = self._stringify(value)
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        return text

    def _safe_float(self, value: Any) -> float | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = self._stringify(value).replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    def _extract_year(self, date_text: str) -> int | None:
        if not date_text:
            return None
        match = re.search(r"(20\d{2})", date_text)
        if match:
            return int(match.group(1))
        return None

    def _normalize_text(self, value: str) -> str:
        return re.sub(r"\s+", " ", value or "").strip().lower()

    def _normalize_affaire_key(self, value: str) -> str:
        return re.sub(r"[^A-Z0-9]", "", (value or "").upper())

    def _guess_chantier_from_filename(self, file_name: str) -> str:
        stem = Path(file_name).stem
        if " - " in stem:
            return stem.split(" - ", 1)[1].strip()
        return stem.strip()
