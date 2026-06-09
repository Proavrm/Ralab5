"""
app/services/qsse_import_service.py
Snapshot importer for QSSE indicator and incident workbooks.

The import model is intentionally generic:
- each workbook is treated as a snapshot source
- rows are stored in qsse_records with a normalized envelope plus raw JSON
- re-importing the same source file replaces the previous snapshot for that file

This keeps the 2026 workbook live and makes the 2025 workbooks one-shot archives.
"""
from __future__ import annotations

import hashlib
import json
import re
import shutil
import sqlite3
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl

from app.core.database import get_qsse_db_path


IGNORE_SHEETS = {
    "sommaire",
    "graphiques",
    "feuil1",
    "liste deroulante",
    "fournisseurs",
}

MONTH_HEADERS = {
    "janvier",
    "fevrier",
    "mars",
    "avril",
    "mai",
    "juin",
    "juillet",
    "aout",
    "septembre",
    "octobre",
    "novembre",
    "decembre",
}


@dataclass(slots=True)
class WorkbookSource:
    path: Path
    source_year: int
    source_mode: str = "live"


@dataclass(slots=True)
class SheetSpec:
    pattern: re.Pattern[str]
    header_row: int
    sheet_kind: str
    register_code: str
    record_kind: str


SHEET_SPECS: tuple[SheetSpec, ...] = (
    SheetSpec(re.compile(r"\bregistre\s+fnc\b"), 2, "fnc_event", "FNC", "event"),
    SheetSpec(re.compile(r"\bfnc\s*2025\b"), 3, "fnc_2025_event", "FNC", "event"),
    # Keep the combined PASD objective sheet ahead of the generic FNC indicator rule.
    SheetSpec(re.compile(r"\bsuivi\s+obj\s+pasd\b"), 4, "pasd_indicator", "PASD", "indicator"),
    SheetSpec(re.compile(r"\bfnc\b"), 2, "fnc_indicator", "FNC", "indicator"),
    SheetSpec(re.compile(r"\bregistre\s+fae\b"), 2, "fae_event", "FAE", "event"),
    SheetSpec(re.compile(r"\bbonnes\s+pratiques\b"), 1, "bp_event", "BP", "event"),
    SheetSpec(re.compile(r"\bregistre\s+bp\b"), 2, "bp_event", "BP", "event"),
    SheetSpec(re.compile(r"\bregis?(?:tre|te)(?:\s+des)?\s+at\b"), 2, "at_event", "AT", "event"),
    SheetSpec(re.compile(r"\bregistre\s+pasd\b"), 2, "pasd_event", "PASD", "event"),
    SheetSpec(re.compile(r"\bpasd\s+d[eé]tails?\b"), 2, "pasd_detail", "PASD", "event"),
    SheetSpec(re.compile(r"\bplan\s+d['’]?actions?\b"), 2, "action_plan", "ACTION", "event"),
    SheetSpec(re.compile(r"\btests?\s+alcool\b.*\bstup"), 3, "tests_alcool_stup", "QSSE", "event"),
    SheetSpec(re.compile(r"\bnon\s*[- ]?respect\s+rv\b|\br[èe]gles\s+vitales\b"), 2, "non_respect_rv", "RV", "event"),
    SheetSpec(re.compile(r"\bnon\s*[- ]?respect\s+epi\b"), 2, "non_respect_epi", "EPI", "event"),
    SheetSpec(re.compile(r"\bquart\s+d['’]?heure\s+cc\s*[- ]\s*ce\b|\bquart\s+d['’]?heure\s+cc\s+ce\b"), 2, "quarter_hour_cc_ce", "QUART_HEURE", "indicator"),
    SheetSpec(re.compile(r"\bquart\s+d['’]?heure\s+encadrant\b"), 2, "quarter_hour_encadrant", "QUART_HEURE", "indicator"),
    SheetSpec(re.compile(r"\bsuggestions?\s*[- ]\s*mesures?\b"), 1, "suggestion_register", "INFO", "event"),
    SheetSpec(re.compile(r"\b5\s*pm\b"), 2, "five_pm_indicator", "QSE", "indicator"),
    SheetSpec(re.compile(r"\beveil\s+musculaire\b"), 2, "eveil_musculaire_indicator", "QSE", "indicator"),
    SheetSpec(re.compile(r"\bvisites\s+qsse\b.*\beveils?\b"), 2, "visite_qsse", "VISITE_QSSE", "indicator"),
    SheetSpec(re.compile(r"\bdetail\s+reporting\s+environnement\b"), 2, "environment_report", "FAE", "event"),
    SheetSpec(re.compile(r"\bautres\s+[- ]?reporting\s+qualite\b"), 2, "quality_other_report", "QSE", "event"),
    SheetSpec(re.compile(r"\bcasses?\s+reseaux\b"), 2, "network_breakage", "FNC", "event"),
    SheetSpec(re.compile(r"\barr[eê]t\s+de\s+chantier\b|\barret\s+de\s+chantier\b"), 2, "work_stop", "QSE", "event"),
    SheetSpec(re.compile(r"\bremontees?\s+d['’]?infos?\b"), 2, "information_indicator", "INFO", "indicator"),
)


class QsseImportService:
    def __init__(self, db_path: Path | None = None) -> None:
        self.db_path = db_path or get_qsse_db_path()

    def status(self) -> dict[str, Any]:
        self._ensure_schema()
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            run_count = conn.execute("SELECT COUNT(*) FROM qsse_import_runs").fetchone()[0]
            record_count = conn.execute("SELECT COUNT(*) FROM qsse_records").fetchone()[0]
            live_count = conn.execute(
                "SELECT COUNT(*) FROM qsse_records WHERE source_mode = 'live'"
            ).fetchone()[0]
            archive_count = conn.execute(
                "SELECT COUNT(*) FROM qsse_records WHERE source_mode = 'closed'"
            ).fetchone()[0]
        return {
            "db_path": str(self.db_path),
            "run_count": int(run_count or 0),
            "record_count": int(record_count or 0),
            "live_record_count": int(live_count or 0),
            "archive_record_count": int(archive_count or 0),
        }

    def preview_sources(self, sources: Iterable[WorkbookSource]) -> dict[str, Any]:
        preview: list[dict[str, Any]] = []
        for source in sources:
            workbook_result = self._preview_workbook(source)
            preview.append(workbook_result)
        return {
            "sources": preview,
            "source_count": len(preview),
            "sheet_count": sum(item["sheet_count"] for item in preview),
            "row_count": sum(item["row_count"] for item in preview),
        }

    def import_sources(self, sources: Iterable[WorkbookSource], replace_existing: bool = True) -> dict[str, Any]:
        self._ensure_schema()
        summary = {
            "source_count": 0,
            "sheet_count": 0,
            "row_count": 0,
            "inserted_count": 0,
            "skipped_count": 0,
            "sources": [],
        }

        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            for source in sources:
                result = self._import_single_source(conn, source, replace_existing=replace_existing)
                summary["source_count"] += 1
                summary["sheet_count"] += result["sheet_count"]
                summary["row_count"] += result["row_count"]
                summary["inserted_count"] += result["inserted_count"]
                summary["skipped_count"] += result["skipped_count"]
                summary["sources"].append(result)

        return summary

    def _import_single_source(
        self,
        conn: sqlite3.Connection,
        source: WorkbookSource,
        *,
        replace_existing: bool,
    ) -> dict[str, Any]:
        source_path = Path(source.path)
        if not source_path.exists():
            raise FileNotFoundError(f"Workbook not found: {source_path}")

        file_hash = self._hash_file(source_path)
        source_name = source_path.name
        workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=False)

        inserted_count = 0
        skipped_count = 0
        row_count = 0
        sheet_count = 0
        restored_documents = 0
        unmatched_documents = 0
        missing_document_files = 0
        restored_rex_drafts = 0
        unmatched_rex_drafts = 0

        try:
            with conn:
                captured_attachments: list[dict[str, Any]] = []
                captured_rex_drafts: list[dict[str, Any]] = []
                if replace_existing:
                    captured_attachments, captured_rex_drafts = self._capture_existing_links(
                        conn,
                        source_name=source_name,
                        source_year=int(source.source_year),
                    )

                if replace_existing:
                    conn.execute(
                        "DELETE FROM qsse_records WHERE source_file = ? AND source_year = ?",
                        (source_name, int(source.source_year)),
                    )

                run_id = self._create_run(
                    conn,
                    source_name=source_name,
                    source_year=int(source.source_year),
                    source_mode=source.source_mode,
                    file_hash=file_hash,
                    workbook_title=workbook.properties.title or source_name,
                )

                for ws in workbook.worksheets:
                    sheet_key = self._normalize_key(ws.title)
                    if sheet_key in IGNORE_SHEETS:
                        continue

                    spec = self._match_sheet_spec(sheet_key)
                    header_row = spec.header_row if spec else self._detect_header_row(ws)
                    if header_row is None:
                        continue

                    parsed_rows = list(self._iter_sheet_rows(ws, header_row))
                    if not parsed_rows:
                        continue

                    sheet_count += 1
                    for row_index, row_map, raw_row in parsed_rows:
                        record = self._build_record(
                            source=source,
                            source_name=source_name,
                            sheet_name=ws.title,
                            sheet_key=sheet_key,
                            row_index=row_index,
                            row_map=row_map,
                            raw_row=raw_row,
                            spec=spec,
                            file_hash=file_hash,
                        )
                        if record is None:
                            skipped_count += 1
                            continue
                        record["run_id"] = run_id
                        conn.execute(
                            """
                            INSERT INTO qsse_records (
                                run_id,
                                source_file,
                                source_year,
                                source_mode,
                                sheet_name,
                                sheet_kind,
                                row_index,
                                register_code,
                                record_kind,
                                agency,
                                entity,
                                person,
                                site,
                                theme,
                                title,
                                description,
                                cause,
                                treatment,
                                corrective_action,
                                action_label,
                                pilot,
                                status,
                                severity,
                                date_event,
                                date_closed,
                                date_saisie,
                                amount_text,
                                amount_value,
                                document_reference,
                                metrics_json,
                                raw_json,
                                row_hash
                            ) VALUES (
                                :run_id,
                                :source_file,
                                :source_year,
                                :source_mode,
                                :sheet_name,
                                :sheet_kind,
                                :row_index,
                                :register_code,
                                :record_kind,
                                :agency,
                                :entity,
                                :person,
                                :site,
                                :theme,
                                :title,
                                :description,
                                :cause,
                                :treatment,
                                :corrective_action,
                                :action_label,
                                :pilot,
                                :status,
                                :severity,
                                :date_event,
                                :date_closed,
                                :date_saisie,
                                :amount_text,
                                :amount_value,
                                :document_reference,
                                :metrics_json,
                                :raw_json,
                                :row_hash
                            )
                            """,
                            record,
                        )
                        inserted_count += 1
                        row_count += 1

                conn.execute(
                    """
                    UPDATE qsse_import_runs
                    SET sheet_count = ?, row_count = ?, inserted_count = ?, skipped_count = ?, status = 'done', updated_at = datetime('now')
                    WHERE id = ?
                    """,
                    (sheet_count, row_count, inserted_count, skipped_count, run_id),
                )

                if replace_existing and (captured_attachments or captured_rex_drafts):
                    restore_result = self._restore_existing_links(
                        conn,
                        source_name=source_name,
                        source_year=int(source.source_year),
                        attachments=captured_attachments,
                        rex_drafts=captured_rex_drafts,
                    )
                    restored_documents = int(restore_result["restored_documents"])
                    unmatched_documents = int(restore_result["unmatched_documents"])
                    missing_document_files = int(restore_result["missing_document_files"])
                    restored_rex_drafts = int(restore_result["restored_rex_drafts"])
                    unmatched_rex_drafts = int(restore_result["unmatched_rex_drafts"])
        finally:
            workbook.close()

        return {
            "source_file": source_name,
            "source_year": int(source.source_year),
            "source_mode": source.source_mode,
            "file_hash": file_hash,
            "sheet_count": sheet_count,
            "row_count": row_count,
            "inserted_count": inserted_count,
            "skipped_count": skipped_count,
            "restored_documents": restored_documents,
            "unmatched_documents": unmatched_documents,
            "missing_document_files": missing_document_files,
            "restored_rex_drafts": restored_rex_drafts,
            "unmatched_rex_drafts": unmatched_rex_drafts,
        }

    def _attachments_root(self) -> Path:
        project_root = Path(__file__).resolve().parents[4]
        return project_root / "storage" / "qsse" / "fnc"

    def _capture_existing_links(
        self,
        conn: sqlite3.Connection,
        *,
        source_name: str,
        source_year: int,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        attachment_rows = conn.execute(
            """
            SELECT
                rec.id AS old_record_id,
                rec.sheet_name,
                rec.row_index,
                doc.stored_name,
                doc.original_name,
                doc.content_type,
                doc.file_size
            FROM qsse_records rec
            JOIN qsse_documents doc ON doc.qsse_record_id = rec.id
            WHERE rec.source_file = ? AND rec.source_year = ?
            ORDER BY rec.id ASC, doc.id ASC
            """,
            (source_name, int(source_year)),
        ).fetchall()

        rex_rows = conn.execute(
            """
            SELECT
                rec.sheet_name,
                rec.row_index,
                draft.provider,
                draft.prompt_version,
                draft.status,
                draft.confidence_score,
                draft.source_payload_json,
                draft.draft_json,
                draft.generated_at,
                draft.reviewed_at,
                draft.approved_at
            FROM qsse_records rec
            JOIN qsse_rex_drafts draft ON draft.qsse_record_id = rec.id
            WHERE rec.source_file = ? AND rec.source_year = ?
            ORDER BY rec.id ASC
            """,
            (source_name, int(source_year)),
        ).fetchall()

        return [dict(row) for row in attachment_rows], [dict(row) for row in rex_rows]

    def _restore_existing_links(
        self,
        conn: sqlite3.Connection,
        *,
        source_name: str,
        source_year: int,
        attachments: list[dict[str, Any]],
        rex_drafts: list[dict[str, Any]],
    ) -> dict[str, int]:
        mapping_rows = conn.execute(
            """
            SELECT id, sheet_name, row_index
            FROM qsse_records
            WHERE source_file = ? AND source_year = ?
            """,
            (source_name, int(source_year)),
        ).fetchall()
        record_map: dict[tuple[str, int], int] = {
            (str(row["sheet_name"] or ""), int(row["row_index"] or 0)): int(row["id"])
            for row in mapping_rows
        }

        attachments_root = self._attachments_root()
        restored_documents = 0
        unmatched_documents = 0
        missing_document_files = 0

        for doc in attachments:
            key = (str(doc.get("sheet_name") or ""), int(doc.get("row_index") or 0))
            new_record_id = record_map.get(key)
            if not new_record_id:
                unmatched_documents += 1
                continue

            old_record_id = int(doc.get("old_record_id") or 0)
            stored_name = str(doc.get("stored_name") or "")
            if not stored_name:
                missing_document_files += 1
                continue

            old_path = attachments_root / str(old_record_id) / stored_name
            new_dir = attachments_root / str(new_record_id)
            new_dir.mkdir(parents=True, exist_ok=True)
            new_path = new_dir / stored_name

            if old_path.exists() and old_path.resolve() != new_path.resolve():
                shutil.copy2(old_path, new_path)
            elif not old_path.exists() and not new_path.exists():
                missing_document_files += 1
                continue

            conn.execute(
                """
                INSERT INTO qsse_documents (
                    qsse_record_id,
                    stored_name,
                    original_name,
                    content_type,
                    file_size
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (
                    int(new_record_id),
                    stored_name,
                    str(doc.get("original_name") or stored_name),
                    str(doc.get("content_type") or "application/octet-stream"),
                    int(doc.get("file_size") or 0),
                ),
            )
            restored_documents += 1

        restored_rex_drafts = 0
        unmatched_rex_drafts = 0
        for draft in rex_drafts:
            key = (str(draft.get("sheet_name") or ""), int(draft.get("row_index") or 0))
            new_record_id = record_map.get(key)
            if not new_record_id:
                unmatched_rex_drafts += 1
                continue

            conn.execute(
                """
                INSERT INTO qsse_rex_drafts (
                    qsse_record_id,
                    provider,
                    prompt_version,
                    status,
                    confidence_score,
                    source_payload_json,
                    draft_json,
                    generated_at,
                    reviewed_at,
                    approved_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(qsse_record_id) DO UPDATE SET
                    provider = excluded.provider,
                    prompt_version = excluded.prompt_version,
                    status = excluded.status,
                    confidence_score = excluded.confidence_score,
                    source_payload_json = excluded.source_payload_json,
                    draft_json = excluded.draft_json,
                    generated_at = excluded.generated_at,
                    reviewed_at = excluded.reviewed_at,
                    approved_at = excluded.approved_at,
                    updated_at = datetime('now')
                """,
                (
                    int(new_record_id),
                    str(draft.get("provider") or ""),
                    str(draft.get("prompt_version") or ""),
                    str(draft.get("status") or "draft"),
                    int(draft.get("confidence_score") or 0),
                    str(draft.get("source_payload_json") or "{}"),
                    str(draft.get("draft_json") or "{}"),
                    str(draft.get("generated_at") or ""),
                    str(draft.get("reviewed_at") or ""),
                    str(draft.get("approved_at") or ""),
                ),
            )
            restored_rex_drafts += 1

        return {
            "restored_documents": restored_documents,
            "unmatched_documents": unmatched_documents,
            "missing_document_files": missing_document_files,
            "restored_rex_drafts": restored_rex_drafts,
            "unmatched_rex_drafts": unmatched_rex_drafts,
        }

    def _preview_workbook(self, source: WorkbookSource) -> dict[str, Any]:
        source_path = Path(source.path)
        if not source_path.exists():
            raise FileNotFoundError(f"Workbook not found: {source_path}")

        workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=False)
        try:
            sheets: list[dict[str, Any]] = []
            row_count = 0
            sheet_count = 0
            for ws in workbook.worksheets:
                sheet_key = self._normalize_key(ws.title)
                if sheet_key in IGNORE_SHEETS:
                    continue
                spec = self._match_sheet_spec(sheet_key)
                header_row = spec.header_row if spec else self._detect_header_row(ws)
                if header_row is None:
                    continue
                parsed_rows = list(self._iter_sheet_rows(ws, header_row))
                if not parsed_rows:
                    continue
                sheet_count += 1
                row_count += len(parsed_rows)
                register_code = spec.register_code if spec else self._guess_register_code(ws.title)
                sheets.append(
                    {
                        "sheet_name": ws.title,
                        "sheet_kind": spec.sheet_kind if spec else "generic",
                        "register_code": register_code,
                        "header_row": header_row,
                        "row_count": len(parsed_rows),
                    }
                )
            return {
                "source_file": source_path.name,
                "source_year": int(source.source_year),
                "source_mode": source.source_mode,
                "sheet_count": sheet_count,
                "row_count": row_count,
                "sheets": sheets,
            }
        finally:
            workbook.close()

    def _create_run(
        self,
        conn: sqlite3.Connection,
        *,
        source_name: str,
        source_year: int,
        source_mode: str,
        file_hash: str,
        workbook_title: str,
    ) -> int:
        cur = conn.execute(
            """
            INSERT INTO qsse_import_runs (
                source_file,
                source_year,
                source_mode,
                file_hash,
                workbook_title,
                status
            ) VALUES (?, ?, ?, ?, ?, 'running')
            """,
            (source_name, source_year, source_mode, file_hash, workbook_title),
        )
        return int(cur.lastrowid)

    def _build_record(
        self,
        *,
        source: WorkbookSource,
        source_name: str,
        sheet_name: str,
        sheet_key: str,
        row_index: int,
        row_map: dict[str, Any],
        raw_row: dict[str, Any],
        spec: SheetSpec | None,
        file_hash: str,
    ) -> dict[str, Any] | None:
        normalized_row = {self._normalize_key(key): value for key, value in row_map.items()}
        record_kind = spec.record_kind if spec else self._record_kind_for_sheet(sheet_key)
        register_code = spec.register_code if spec else self._guess_register_code(sheet_name)
        generic_date = self._pick(normalized_row, "date")
        event_date = self._pick(
            normalized_row,
            "date de l'évènement",
            "date de l evenement",
            "date de l action",
        )
        closed_date = self._pick(
            normalized_row,
            "date de cloture",
            "date de clôture",
            "echeance de l'action",
            "échéance de l'action",
            "dates de dernier jour d'arrêt",
            "dates de dernier jour d arret",
        )
        saisie_date = self._pick(
            normalized_row,
            "date de saisie",
            "date saisie",
            "date d'enregistrement",
            "date d enregistrement",
        )

        # Many QSSE event sheets only expose a single generic DATE column.
        # Persist it as both entry date and event date when no dedicated field exists,
        # so reimports keep quarterly analyses and filters usable instead of blank.
        if record_kind == "event" and generic_date:
            if not saisie_date:
                saisie_date = generic_date
            if not event_date:
                event_date = generic_date

        data = {
            "source_file": source_name,
            "source_year": int(source.source_year),
            "source_mode": source.source_mode,
            "sheet_name": sheet_name,
            "sheet_kind": spec.sheet_kind if spec else "generic",
            "row_index": row_index,
            "register_code": register_code,
            "record_kind": record_kind,
            "agency": self._pick(normalized_row, "agence", "region / entite", "région / entité", "entite", "agence / entite"),
            "entity": self._pick(normalized_row, "entite", "région / entité", "region / entite", "entreprise", "société", "societe"),
            "person": self._pick(
                normalized_row,
                "nom / prenom responsable",
                "nom / prénom responsable",
                "nom / prenom",
                "nom / prénom",
                "nom et prenom",
                "nom et prénom",
                "nom prenom du personnel",
                "nom du preventeur",
                "nom",
                "redacteur",
                "rédacteur",
            ),
            "site": self._pick(
                normalized_row,
                "chantier ou site concerne",
                "chantier ou site concerné",
                "chantier",
                "nom de la plateforme",
                "plateforme fixe ou provisoire (chantier) ou associee a une carriere ou a une industrie",
            ),
            "theme": self._pick(
                normalized_row,
                "theme (pollution / poussiere / surconsommation / dechets…)",
                "thematique (q / s / e)",
                "theme q/s/sa/e/ systeme / financier",
                "theme",
                "rubrique(s) icpe",
            ),
            "title": self._pick(
                normalized_row,
                "recit de l'evenement",
                "recit de l evenement",
                "description de la non-conformite ouvrage",
                "description",
                "constat",
                "sujet",
                "presqu'accident/ situation dangereuse",
                "presqu accident/ situation dangereuse",
            ),
            "description": self._pick(
                normalized_row,
                "recit de l'evenement",
                "recit de l evenement",
                "contexte",
                "constat",
                "circonstances de l'accident",
                "circonstances de l accident",
                "recit de l'action",
                "recit de l action",
            ),
            "cause": self._pick(
                normalized_row,
                "cause",
                "causes identifiees",
                "causes identifiées",
                "cause de la nc",
                "origine de l'action",
            ),
            "treatment": self._pick(
                normalized_row,
                "traitement",
                "traitement de la nc",
                "traitement",
                "action définie",
                "action definie",
                "mesure immediate",
            ),
            "corrective_action": self._pick(
                normalized_row,
                "action corrective",
                "actions correctives",
                "action",
                "action definie",
                "action définie",
            ),
            "action_label": self._pick(
                normalized_row,
                "action",
                "action définie",
                "action definie",
                "détails",
                "details",
                "commentaire",
            ),
            "pilot": self._pick(
                normalized_row,
                "pilote",
                "responsable de l'action",
                "responsable de l action",
                "responsable",
            ),
            "status": self._pick(
                normalized_row,
                "statut",
                "statut de la realisation de l'action",
                "statut de la réalisation de l’action",
                "realisee (oui / non)",
                "réalisée (oui / non)",
                "enregistrement",
            ),
            "severity": self._pick(
                normalized_row,
                "severite",
                "sévérité",
                "types de risques afin de statuer sur l'efficacité des actions lors du bilan",
            ),
            "date_event": event_date,
            "date_closed": closed_date,
            "date_saisie": saisie_date,
            "amount_text": self._pick(normalized_row, "coût", "cout", "chiffrage du traitement", "coût du traitement"),
            "amount_value": self._parse_amount(self._pick(normalized_row, "coût", "cout", "chiffrage du traitement", "coût du traitement")),
            "document_reference": self._pick(
                normalized_row,
                "document de reference",
                "document de référence",
                "enregistrement",
                "support (kizeo bien vu / mail / fiche reference / fiche bonne pratique)",
                "support (kizéo bien vu / mail / fiche référence / fiche bonne pratique)",
            ),
            "metrics_json": json.dumps(self._extract_metrics(normalized_row), ensure_ascii=False, sort_keys=True, default=str),
            "raw_json": json.dumps(
                {
                    "source_year": int(source.source_year),
                    "source_mode": source.source_mode,
                    "source_file": source_name,
                    "sheet_name": sheet_name,
                    "sheet_kind": spec.sheet_kind if spec else "generic",
                    "row_index": row_index,
                    "headers": list(row_map.keys()),
                    "row": raw_row,
                },
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            ),
            "row_hash": self._hash_payload(raw_row),
        }

        # Post-normalisation tweaks for a few known formats.
        if record_kind == "indicator":
            if not data["title"]:
                data["title"] = self._pick(normalized_row, "nom / prenom responsable", "nom du preventeur", "nom du personnel", "nom")
            if not data["description"]:
                data["description"] = self._pick(normalized_row, "commentaire", "remarques", "détail noms et actions", "detail noms et actions")
            if not data["status"]:
                data["status"] = self._pick(normalized_row, "objectif année", "objectif annee", "nbr de 1/4 heure", "nbr de 1/4 h")

        if spec and spec.record_kind == "indicator":
            # Preserve the table row as an indicator snapshot rather than a narrative event.
            data["title"] = data["title"] or self._pick(normalized_row, "nom / prenom responsable", "nom / prénom responsable", "nom / prenom", "nom")
            data["description"] = data["description"] or self._pick(normalized_row, "poste", "fonction", "agence")

        if spec and spec.sheet_kind in {"fnc_event", "fnc_2025_event", "fae_event", "bp_event", "pasd_event", "pasd_detail"}:
            explicit_subject = self._pick(
                normalized_row,
                "sujet",
                "objet",
                "intitule",
                "intitulé",
                "designation",
                "désignation",
            )
            document_subject = self._derive_subject_from_document_reference(data["document_reference"])
            preferred_subject = self._first_non_empty((explicit_subject, document_subject))
            title_is_narrative = (
                not data["title"]
                or data["title"] == data["description"]
                or "\n" in data["title"]
                or len(data["title"]) > 140
            )
            if preferred_subject and title_is_narrative:
                data["title"] = preferred_subject

        if spec and spec.sheet_kind == "at_event":
            data["title"] = data["title"] or self._pick(
                normalized_row,
                "circonstances de l'accident",
                "circonstances de l accident",
                "nom et prenom",
                "nom et prénom",
            )
            data["description"] = data["description"] or self._pick(
                normalized_row,
                "circonstances de l'accident",
                "circonstances de l accident",
                "lésion",
                "lesion",
            )

        if spec and spec.sheet_kind == "suggestion_register":
            data["title"] = data["title"] or self._pick(normalized_row, "propositions", "suggestion remontée")
            data["description"] = data["description"] or self._pick(normalized_row, "observations")
            data["action_label"] = data["action_label"] or self._pick(normalized_row, "actions")
            data["corrective_action"] = data["corrective_action"] or data["action_label"]
            data["date_closed"] = data["date_closed"] or self._pick(normalized_row, "delais", "délais")

        if not data["title"]:
            data["title"] = self._first_non_empty(normalized_row.values())

        if not data["description"]:
            data["description"] = data["title"]

        if record_kind == "indicator" and not any((data["agency"], data["entity"], data["person"])) and data["title"] in {"", "#DIV/0!", "0"}:
            return None

        # AT sheets contain summary rows with only counters like 0/NON/4.
        # Keep only rows that carry at least one incident anchor.
        if spec and spec.sheet_kind == "at_event" and not any((data["date_event"], data["agency"], data["site"])):
            return None

        if spec and spec.sheet_kind == "pasd_event" and not any((data["date_event"], data["agency"], data["site"], data["person"])):
            return None

        return data

    def _extract_metrics(self, normalized_row: dict[str, Any]) -> dict[str, Any]:
        metrics: dict[str, Any] = {}
        for key, value in normalized_row.items():
            if key in {
                "agence",
                "region / entite",
                "région / entité",
                "entite",
                "entreprise",
                "societe",
                "société",
                "nom / prenom responsable",
                "nom et prenom",
                "nom prenom du personnel",
                "nom du preventeur",
                "nom",
                "redacteur",
                "rédacteur",
                "chantier ou site concerne",
                "chantier ou site concerné",
                "chantier",
                "nom de la plateforme",
                "date",
                "date de clôture",
                "date de cloture",
                "coût",
                "cout",
                "document de reference",
                "document de référence",
                "description de la non-conformite ouvrage",
                "description",
                "recit de l'evenement",
                "recit de l evenement",
                "causes identifiees",
                "causes identifiées",
                "traitement",
                "action corrective",
                "action",
                "pilote",
                "statut",
                "severite",
                "sévérité",
                "theme (pollution / poussiere / surconsommation / dechets…)",
                "thematique (q / s / e)",
                "theme q/s/sa/e/ systeme / financier",
                "theme",
            }:
                continue
            if key in MONTH_HEADERS or key.startswith("nombre ") or key.startswith("objectifs"):
                metrics[key] = value
                continue
            if value not in (None, ""):
                metrics[key] = value
        return metrics

    def _iter_sheet_rows(self, ws, header_row: int) -> Iterable[tuple[int, dict[str, Any], dict[str, Any]]]:
        headers = self._read_headers(ws, header_row)
        if not headers:
            return []

        last_col = max(headers)
        blank_streak = 0
        for row_index in range(header_row + 1, ws.max_row + 1):
            raw_cells = [ws.cell(row=row_index, column=col_index).value for col_index in range(1, last_col + 1)]
            if self._row_is_blank(raw_cells):
                blank_streak += 1
                if blank_streak >= 5:
                    break
                continue
            blank_streak = 0
            row_map = {
                self._clean(headers[col_index]): raw_cells[col_index - 1]
                for col_index in headers
                if self._clean(headers[col_index])
            }
            raw_row = {str(col_index): raw_cells[col_index - 1] for col_index in range(1, last_col + 1)}
            yield row_index, row_map, raw_row

    def _read_headers(self, ws, header_row: int) -> dict[int, str]:
        headers: dict[int, str] = {}
        last_col = 0
        for col_index in range(1, ws.max_column + 1):
            value = ws.cell(row=header_row, column=col_index).value
            if value not in (None, ""):
                headers[col_index] = self._clean(value)
                last_col = col_index
        return {index: headers[index] for index in range(1, last_col + 1) if index in headers}

    def _detect_header_row(self, ws) -> int | None:
        best_row = None
        best_score = 0
        for row_index in range(1, min(ws.max_row, 15) + 1):
            values = [self._clean(ws.cell(row=row_index, column=col_index).value) for col_index in range(1, min(ws.max_column, 40) + 1)]
            score = sum(1 for value in values if value)
            if score > best_score:
                best_score = score
                best_row = row_index
        return best_row if best_score >= 3 else None

    def _match_sheet_spec(self, sheet_key: str) -> SheetSpec | None:
        for spec in SHEET_SPECS:
            if spec.pattern.search(sheet_key):
                return spec
        return None

    def _record_kind_for_sheet(self, sheet_key: str) -> str:
        if any(token in sheet_key for token in ("quart heure", "visite qsse", "fnc ", "suivi obj pasd", "5 pm", "eveil musculaire", "remontees d infos")):
            return "indicator"
        return "event"

    def _guess_register_code(self, sheet_title: str) -> str:
        key = self._normalize_key(sheet_title)
        if "fae" in key:
            return "FAE"
        if "bp" in key:
            return "BP"
        if "fnc" in key:
            return "FNC"
        if "pasd" in key:
            return "PASD"
        if "alcool" in key or "stup" in key:
            return "QSSE"
        if "epi" in key:
            return "EPI"
        if "rv" in key or "vitales" in key:
            return "RV"
        if "at" in key:
            return "AT"
        if "action" in key:
            return "ACTION"
        if "visite" in key or "quart" in key:
            return "QUART_HEURE"
        return "QSE"

    def _pick(self, row_map: dict[str, Any], *aliases: str) -> str:
        alias_keys = [self._normalize_key(alias) for alias in aliases]
        for alias in alias_keys:
            if alias in row_map:
                value = self._clean(row_map.get(alias))
                if value:
                    return value
        for alias in alias_keys:
            for key, value in row_map.items():
                if alias and alias in key:
                    cleaned = self._clean(value)
                    if cleaned:
                        return cleaned
        return ""

    def _first_non_empty(self, values: Iterable[Any]) -> str:
        for value in values:
            cleaned = self._clean(value)
            if cleaned:
                return cleaned
        return ""

    def _derive_subject_from_document_reference(self, value: Any) -> str:
        text = self._clean(value)
        if not text:
            return ""

        if " - " in text:
            candidate = self._clean(text.rsplit(" - ", 1)[-1])
            if candidate and candidate != text:
                return candidate

        if "_" in text:
            candidate = self._clean(text.rsplit("_", 1)[-1])
            if candidate and candidate != text:
                return candidate

        return ""

    def _row_is_blank(self, raw_cells: list[Any]) -> bool:
        return all(self._clean(value) == "" for value in raw_cells)

    def _hash_file(self, path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _hash_payload(self, payload: dict[str, Any]) -> str:
        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    def _parse_amount(self, value: Any) -> float | None:
        text = self._clean(value).replace("€", "").replace("ht", "").replace("HT", "")
        if not text:
            return None
        match = re.search(r"(-?\d+(?:[.,]\d+)?)", text)
        if not match:
            return None
        number = match.group(1).replace(",", ".")
        try:
            return float(number)
        except ValueError:
            return None

    def _normalize_key(self, value: Any) -> str:
        text = self._clean(value).lower()
        normalized = unicodedata.normalize("NFKD", text)
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = normalized.replace("’", "'")
        normalized = normalized.replace("/", " ")
        normalized = normalized.replace("-", " ")
        normalized = normalized.replace("_", " ")
        normalized = normalized.replace(".", " ")
        normalized = normalized.replace(",", " ")
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized.strip()

    def _clean(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return str(value).strip()

    def _ensure_schema(self) -> None:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        from app.core.database import ensure_ralab4_schema

        ensure_ralab4_schema(self.db_path)


__all__ = ["WorkbookSource", "QsseImportService"]
