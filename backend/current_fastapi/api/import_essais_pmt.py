from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from datetime import date, datetime
from io import BytesIO
from typing import Any

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from app.core.database import get_db_path
from api.import_essais_base import (
    _resolve_affaire_context,
    ensure_hierarchy,
    ensure_modules_enabled,
    group_rows_by_temporal_gap,
    _predict_references,
)

router = APIRouter(prefix="/api/import-essais-pmt", tags=["Import Essais PMT"])
DB_PATH = get_db_path()


def _conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    text = _clean(value).replace(",", ".").replace(" ", "")
    try:
        number = float(text)
    except Exception:
        return None
    return number if number == number else None


def _extract_date_iso(text: str) -> str:
    if isinstance(text, (datetime, date)):
        return text.strftime("%Y-%m-%d")
    raw = _clean(text)
    if not raw:
        return ""
    match_iso = re.search(r"(\d{4})-(\d{2})-(\d{2})", raw)
    if match_iso:
        year, month, day = match_iso.groups()
        return f"{year}-{month}-{day}"
    match = re.search(r"(\d{2})[/-](\d{2})[/-](\d{4})", raw)
    if not match:
        return ""
    day, month, year = match.groups()
    return f"{year}-{month}-{day}"


def _extract_night_range(raw: str) -> tuple[str, str]:
    text = _clean(raw)
    if not text:
        return "", ""
    # Example: "Nuit 09-10/10/2025"
    match = re.search(r"(\d{2})-(\d{2})/(\d{2})/(\d{4})", text)
    if match:
        day_start, day_end, month, year = match.groups()
        return f"{year}-{month}-{day_start}", f"{year}-{month}-{day_end}"
    iso = _extract_date_iso(text)
    return iso, iso


def _extract_threshold(criteria_definition: str) -> float | None:
    # Handles forms like "PMT ≥ 0.4", "PMT >= 0,4"
    text = _clean(criteria_definition).replace(",", ".")
    match = re.search(r"PMT\s*[≥>=]+\s*([0-9]+(?:\.[0-9]+)?)", text, flags=re.IGNORECASE)
    if not match:
        return None
    return _as_float(match.group(1))


def _extract_year_from_sheet(sheet: dict[str, Any], fallback_year: int | None = None) -> int:
    date_iso = _clean(sheet.get("date_essai_debut"))
    if re.match(r"^\d{4}-\d{2}-\d{2}$", date_iso):
        return int(date_iso[:4])
    if fallback_year is not None:
        return int(fallback_year)
    return datetime.utcnow().year


def _next_pmt_reference(conn: sqlite3.Connection, year: int) -> str:
    prefix = f"{int(year)}-SP-PMT"
    rows = conn.execute(
        "SELECT reference FROM pmt_essais WHERE reference LIKE ?",
        (f"{prefix}%",),
    ).fetchall()
    max_num = 0
    for row in rows:
        ref = _clean(row["reference"]).upper()
        match = re.match(rf"^{re.escape(prefix)}(\d+)$", ref)
        if not match:
            continue
        try:
            max_num = max(max_num, int(match.group(1)))
        except ValueError:
            continue
    return f"{prefix}{max_num + 1:04d}"


def _load_next_pmt_sequence_by_year(conn: sqlite3.Connection) -> dict[int, int]:
    rows = conn.execute("SELECT reference FROM pmt_essais WHERE reference LIKE '%-SP-PMT%'").fetchall()
    by_year: dict[int, int] = {}
    for row in rows:
        ref = _clean(row["reference"]).upper()
        match = re.match(r"^(\d{4})-SP-PMT(\d+)$", ref)
        if not match:
            continue
        year = int(match.group(1))
        seq = int(match.group(2))
        current = by_year.get(year, 1)
        if seq + 1 > current:
            by_year[year] = seq + 1
    return by_year


def _is_pmt_reference_format(value: str) -> bool:
    return bool(re.match(r"^\d{4}-SP-PMT\d{4,}$", _clean(value).upper()))


def _find_label(ws, label: str, max_row: int = 180, max_col: int = 32) -> tuple[int, int] | None:
    wanted = _clean(label).upper()
    for row_idx in range(1, min(max_row, ws.max_row) + 1):
        for col_idx in range(1, min(max_col, ws.max_column) + 1):
            if _clean(ws.cell(row=row_idx, column=col_idx).value).upper() == wanted:
                return row_idx, col_idx
    return None


def _first_value_right(ws, row_idx: int, col_idx: int, span: int = 10) -> Any:
    for probe_col in range(col_idx + 1, min(ws.max_column, col_idx + span) + 1):
        value = ws.cell(row=row_idx, column=probe_col).value
        if value not in (None, ""):
            return value
    return None


def _collect_comment_block(ws) -> str:
    marker = _find_label(ws, "COMMENTAIRES", max_row=220)
    if not marker:
        return ""
    row_idx, _ = marker
    lines: list[str] = []
    for r in range(row_idx + 1, min(row_idx + 6, ws.max_row) + 1):
        for c in range(1, min(20, ws.max_column) + 1):
            text = _clean(ws.cell(row=r, column=c).value)
            if text and text.upper() not in {"COMMENTAIRES", "VISA"}:
                lines.append(text)
    return "\n".join(lines).strip()


def _extract_points(ws) -> list[dict[str, Any]]:
    header_row = None
    for row_idx in range(1, min(ws.max_row, 220) + 1):
        row_values = [_clean(ws.cell(row=row_idx, column=c).value) for c in range(1, min(ws.max_column, 32) + 1)]
        joined = " | ".join(v for v in row_values if v)
        if "N° essai" in joined and "Diamètre moyen de la tache" in joined and "Profondeurs de macrotexture" in joined:
            header_row = row_idx
            break
    if not header_row:
        return []

    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, min(header_row + 120, ws.max_row) + 1):
        n_essai_raw = ws.cell(row=row_idx, column=3).value
        diam_raw = ws.cell(row=row_idx, column=9).value
        pmt_raw = ws.cell(row=row_idx, column=15).value
        profil = _clean(ws.cell(row=row_idx, column=4).value)
        position = _clean(ws.cell(row=row_idx, column=6).value)

        if _clean(n_essai_raw).lower().startswith("nb d'essais"):
            break
        if _clean(position).lower().startswith("pourcentage de valeurs conformes"):
            break

        if all(v in (None, "") for v in (n_essai_raw, diam_raw, pmt_raw, profil, position)):
            continue

        numero = _clean(n_essai_raw)
        if not numero:
            continue

        rows.append(
            {
                "ordre": len(rows) + 1,
                "numero_essai": numero,
                "profil": profil,
                "position": position,
                "diametre_moyen_tache_mm": _as_float(diam_raw),
                "profondeur_macrotexture_mm": _as_float(pmt_raw),
                "observation": "",
                "donnees_ligne_json": {
                    "row_index": row_idx,
                    "n_essai_raw": n_essai_raw,
                    "diametre_raw": diam_raw,
                    "profondeur_raw": pmt_raw,
                },
            }
        )

    return rows


def _summarize(points: list[dict[str, Any]], seuil_pmt_min_mm: float | None) -> dict[str, Any]:
    values = [p.get("profondeur_macrotexture_mm") for p in points if p.get("profondeur_macrotexture_mm") is not None]
    if not values:
        return {
            "nombre_essais": 0,
            "pmt_moyenne_mm": None,
            "pmt_min_mm": None,
            "pmt_max_mm": None,
            "pmt_ecart_type_mm": None,
            "pourcentage_valeurs_conformes": None,
            "nombre_points_conformes": 0,
            "nombre_points_non_conformes": 0,
        }
    n = len(values)
    avg = sum(values) / n
    var = sum((v - avg) ** 2 for v in values) / n
    std = var ** 0.5
    conformes = 0
    if seuil_pmt_min_mm is not None:
        conformes = sum(1 for v in values if v >= seuil_pmt_min_mm)
    non_conformes = n - conformes if seuil_pmt_min_mm is not None else 0
    pct = (conformes * 100.0 / n) if (seuil_pmt_min_mm is not None and n > 0) else None
    return {
        "nombre_essais": n,
        "pmt_moyenne_mm": round(avg, 4),
        "pmt_min_mm": min(values),
        "pmt_max_mm": max(values),
        "pmt_ecart_type_mm": round(std, 4),
        "pourcentage_valeurs_conformes": round(pct, 2) if pct is not None else None,
        "nombre_points_conformes": conformes,
        "nombre_points_non_conformes": non_conformes,
    }


def _sheet_payload(ws, file_name: str, file_hash: str) -> dict[str, Any]:
    def read(label: str) -> Any:
        found = _find_label(ws, label)
        if not found:
            return None
        return _first_value_right(ws, found[0], found[1])

    criteria_def = _clean(read("Définition des critères / objectifs :"))
    seuil = _extract_threshold(criteria_def)
    date_essai_texte = _clean(read("Date de l'essai :"))
    date_moe_texte = _clean(read("Date de mise en œuvre :"))
    date_essai_debut, date_essai_fin = _extract_night_range(date_essai_texte)
    date_moe_debut, date_moe_fin = _extract_night_range(date_moe_texte)
    points = _extract_points(ws)
    summary = _summarize(points, seuil)

    sheet_name = _clean(ws.title)
    import_uid = hashlib.sha1(f"{file_hash}|{sheet_name}".encode("utf-8")).hexdigest()
    commentaire = _collect_comment_block(ws)

    # Header row on PMT templates usually carries:
    # PMT | n° | <chrono> | <affaire> | <date_redaction>
    chrono_header = ""
    affaire_header = ""
    date_redaction_header = ""
    for row_idx in (5, 4, 3, 6, 7):
        if not chrono_header:
            chrono_header = _clean(ws.cell(row=row_idx, column=10).value) or _clean(ws.cell(row=row_idx, column=11).value)
        if not affaire_header:
            affaire_header = _clean(ws.cell(row=row_idx, column=12).value) or _clean(ws.cell(row=row_idx, column=13).value)
        if not date_redaction_header:
            date_redaction_header = _extract_date_iso(ws.cell(row=row_idx, column=16).value) or _extract_date_iso(
                ws.cell(row=row_idx, column=17).value
            )
        if chrono_header and affaire_header and date_redaction_header:
            break

    payload = {
        "import_source_file": file_name,
        "import_source_sheet": sheet_name,
        "import_uid": import_uid,
        "code_essai": "PMT",
        "norme": "NF EN 13036-1",
        "chrono": chrono_header or _clean(read("Chrono")),
        "reference_affaire": _sanitize_affaire_reference(affaire_header or _clean(read("N° d'affaire"))),
        "date_redaction": date_redaction_header or _extract_date_iso(_clean(read("Date de rédaction"))),
        "operateur": _clean(read("Opérateur :")),
        "conditions_meteorologiques": _clean(read("Conditions météorologiques :")),
        "laboratoire": _clean(read("Laboratoire :")),
        "produit_controle": _clean(read("Produit contrôlé :")),
        "numero_formule": _clean(read("Numéro de formule")),
        "couche": _clean(read("Couche :")),
        "lieu_fabrication": _clean(read("Lieu de fabrication :")),
        "date_essai_texte": date_essai_texte,
        "date_essai_debut": date_essai_debut,
        "date_essai_fin": date_essai_fin,
        "date_mise_en_oeuvre_texte": date_moe_texte,
        "date_mise_en_oeuvre_debut": date_moe_debut,
        "date_mise_en_oeuvre_fin": date_moe_fin,
        "epaisseur_couche_texte": _clean(read("Epaisseur de la couche :")),
        "section_controlee": _clean(read("Section contrôlée :")),
        "atelier_mise_en_oeuvre": _clean(read("Atelier de mise en œuvre :")),
        "volume_materiau_texte": _clean(read("Volume de matériau utilisé:")),
        "source_criteres": _clean(read("Source des critères :")),
        "definition_criteres": criteria_def,
        "seuil_pmt_min_mm": seuil,
        "conclusion_excel_texte": _clean(read("CONCLUSIONS")),
        "commentaire": commentaire,
        "signataire_nom": _clean(read("Nom")),
        "signataire_fonction": _clean(read("Fonction")),
        "visa_texte": _clean(read("Visa")),
        "donnees_entete_json": {},
        "points_rows": points,
        "summary": summary,
    }
    payload.update(summary)
    return payload


def _normalize_affaire_nge(value: Any) -> str:
    return re.sub(r"\W+", "", _clean(value)).upper()


def _sanitize_affaire_reference(value: Any) -> str:
    raw = _clean(value)
    if not raw:
        return ""
    normalized = _normalize_affaire_nge(raw)
    blocked_tokens = {
        "DATEDEDACTION",
        "DATEDERDACTION",
        "DATEDEREDACTION",
        "NDAFFAIRE",
        "NUMERODAFFAIRE",
        "CHRONO",
    }
    if normalized in blocked_tokens:
        return ""
    if normalized.startswith("DATEDE") and "REDACTION" in normalized:
        return ""
    return raw


def _next_affaire_reference(conn: sqlite3.Connection, year: int, region: str = "RA") -> tuple[str, int]:
    prefix = f"{year}-{region}-A"
    rows = conn.execute(
        "SELECT reference FROM affaires_rst WHERE reference LIKE ?",
        (f"{prefix}%",),
    ).fetchall()
    nums: list[int] = []
    for row in rows:
        ref = _clean(row["reference"] if isinstance(row, sqlite3.Row) else row[0])
        match = re.match(rf"^{re.escape(prefix)}(\d+)$", ref)
        if match:
            nums.append(int(match.group(1)))
    number = max(nums, default=0) + 1
    return f"{prefix}{number:04d}", number


def _ensure_affaire_context(conn: sqlite3.Connection, row: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    affaire_ref = _sanitize_affaire_reference(row.get("reference_affaire"))
    affaire_nge = _normalize_affaire_nge(affaire_ref)
    context = _resolve_affaire_context(conn, affaire_reference=affaire_ref, affaire_nge=affaire_nge)
    if context.get("selected"):
        return context, False

    year = datetime.now().year
    date_essai = _clean(row.get("date_essai_debut"))
    if re.match(r"^\d{4}-\d{2}-\d{2}$", date_essai):
        year = int(date_essai[:4])
    ref_auto, numero = _next_affaire_reference(conn, year=year, region="RA")
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    chantier = _clean(row.get("section_controlee"))
    reference_to_use = affaire_ref or ref_auto
    conn.execute(
        """
        INSERT INTO affaires_rst (
            reference, annee, region, numero, client, titulaire, chantier, affaire_nge,
            date_ouverture, statut, responsable, created_at, updated_at
        ) VALUES (?, ?, ?, ?, '', '', ?, ?, ?, 'À qualifier', 'Import PMT', ?, ?)
        """,
        (
            reference_to_use,
            year,
            "RA",
            numero,
            chantier,
            affaire_nge,
            date_essai or datetime.utcnow().strftime("%Y-%m-%d"),
            now,
            now,
        ),
    )
    context = _resolve_affaire_context(conn, affaire_reference=reference_to_use, affaire_nge=affaire_nge)
    if not context.get("selected"):
        raise HTTPException(status_code=500, detail="Impossible de résoudre/créer le contexte affaire PMT.")
    return context, True


def _preview_workbook(file_name: str, content: bytes) -> dict[str, Any]:
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Impossible de lire le fichier Excel PMT: {exc}") from exc
    file_hash = hashlib.sha256(content).hexdigest()
    sheets = [_sheet_payload(ws, file_name, file_hash) for ws in wb.worksheets]
    return {"file_name": file_name, "file_hash": file_hash, "sheet_count": len(sheets), "sheets": sheets}


def _build_already_imported_map(conn: sqlite3.Connection, file_name: str) -> dict[str, dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT id, import_source_sheet, demande_id, campaign_id, intervention_id
        FROM pmt_essais
        WHERE import_source_file = ?
        ORDER BY id DESC
        """,
        (file_name,),
    ).fetchall()
    mapping: dict[str, dict[str, Any]] = {}
    for row in rows:
        sheet_name = _clean(row["import_source_sheet"])
        if not sheet_name:
            continue
        if sheet_name in mapping:
            continue
        mapping[sheet_name] = {
            "pmt_id": int(row["id"]),
            "demande_id": row["demande_id"],
            "campagne_id": row["campaign_id"],
            "intervention_id": row["intervention_id"],
        }
    return mapping


def _lookup_reference_map(conn: sqlite3.Connection, table: str, ids: set[int]) -> dict[int, str]:
    if not ids:
        return {}
    placeholders = ", ".join("?" for _ in ids)
    rows = conn.execute(
        f"SELECT id, reference FROM {table} WHERE id IN ({placeholders})",
        tuple(sorted(ids)),
    ).fetchall()
    return {int(row["id"]): _clean(row["reference"]) for row in rows}


def _build_preview_proposals(
    conn: sqlite3.Connection,
    sheets: list[dict[str, Any]],
    demande_gap_days: int,
    campagne_gap_days: int,
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for sheet in sheets:
        anchor = _clean(sheet.get("date_essai_debut"))
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", anchor):
            continue
        rows.append(
            {
                "sheet": _clean(sheet.get("import_source_sheet")),
                "anchor_date": anchor,
                "already_imported": bool(sheet.get("already_imported")),
            }
        )
    rows.sort(key=lambda item: (item["anchor_date"], item["sheet"]))
    demande_groups = group_rows_by_temporal_gap(rows, gap_days=max(1, int(demande_gap_days)), date_field="anchor_date")
    base_groups: list[list[dict[str, Any]]] = []
    for demande_group in demande_groups:
        base_groups.append(
            [
                {
                    "sheet_name": row.get("sheet"),
                    "date_sondage": row.get("anchor_date"),
                }
                for row in demande_group
            ]
        )
    predictions = _predict_references(
        conn,
        base_groups,
        max(1, int(campagne_gap_days)),
        labo_code="SP",
    )
    demandes = []
    for d_idx, d_group in enumerate(demande_groups, start=1):
        d_pred = predictions[d_idx - 1] if d_idx - 1 < len(predictions) else {}
        campagne_groups = group_rows_by_temporal_gap(
            d_group,
            gap_days=max(1, int(campagne_gap_days)),
            date_field="anchor_date",
        )
        campagnes = []
        interventions_count = 0
        for c_idx, c_group in enumerate(campagne_groups, start=1):
            c_pred = (d_pred.get("campagnes") or [])[c_idx - 1] if c_idx - 1 < len(d_pred.get("campagnes") or []) else {}
            sheets_names = [item["sheet"] for item in c_group]
            dates = sorted({_clean(item["anchor_date"]) for item in c_group if _clean(item.get("anchor_date"))})
            raw_i_preds = c_pred.get("interventions") or []
            predicted_i_map = {
                _clean(pred.get("sheet_name")): _clean(pred.get("predicted_intervention_reference"))
                for pred in raw_i_preds
                if _clean(pred.get("sheet_name"))
            }
            interventions_count += len(dates)
            campagnes.append(
                {
                    "proposal_index": c_idx,
                    "start_date": dates[0] if dates else "",
                    "end_date": dates[-1] if dates else "",
                    "interventions_count": len(dates),
                    "sheets": sheets_names,
                    "predicted_campagne_reference": _clean(c_pred.get("predicted_campagne_reference")),
                    "predicted_intervention_references": predicted_i_map,
                }
            )
        demandes.append(
            {
                "proposal_index": d_idx,
                "campagnes_count": len(campagnes),
                "interventions_count": interventions_count,
                "imported_count": sum(1 for row in d_group if row.get("already_imported")),
                "campagnes": campagnes,
                "sheets": [row["sheet"] for row in d_group],
                "predicted_demande_reference": _clean(d_pred.get("predicted_demande_reference")),
            }
        )
    return {"demandes_count": len(demandes), "demandes": demandes}


def _track_hierarchy_stat(stats: dict[str, dict[str, set[int]]], entity: str, entity_id: int, created: bool) -> None:
    if created:
        stats[entity]["created"].add(entity_id)
        stats[entity]["existing"].discard(entity_id)
        return
    if entity_id not in stats[entity]["created"]:
        stats[entity]["existing"].add(entity_id)


@router.post("/preview-upload")
async def preview_upload(
    file: UploadFile = File(...),
    affaire_reference: str = Form(default=""),
    affaire_nge: str = Form(default=""),
    demande_gap_days: int = Form(default=120),
    campagne_gap_days: int = Form(default=7),
):
    name = _clean(file.filename) or "pmt.xlsx"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Fichier vide.")
    preview = _preview_workbook(name, content)

    with _conn() as conn:
        imported_map = _build_already_imported_map(conn, preview["file_name"])
        for sheet in preview["sheets"]:
            sheet_name = _clean(sheet.get("import_source_sheet"))
            imported = imported_map.get(sheet_name)
            sheet["already_imported"] = bool(imported)
            if imported:
                sheet["existing_pmt_id"] = imported["pmt_id"]
                sheet["existing_demande_id"] = imported["demande_id"]
                sheet["existing_campagne_id"] = imported["campagne_id"]
                sheet["existing_intervention_id"] = imported["intervention_id"]

        existing_demande_ids = {
            int(sheet["existing_demande_id"])
            for sheet in preview["sheets"]
            if sheet.get("existing_demande_id") not in (None, "")
        }
        existing_campagne_ids = {
            int(sheet["existing_campagne_id"])
            for sheet in preview["sheets"]
            if sheet.get("existing_campagne_id") not in (None, "")
        }
        existing_intervention_ids = {
            int(sheet["existing_intervention_id"])
            for sheet in preview["sheets"]
            if sheet.get("existing_intervention_id") not in (None, "")
        }
        existing_pmt_ids = {
            int(sheet["existing_pmt_id"])
            for sheet in preview["sheets"]
            if sheet.get("existing_pmt_id") not in (None, "")
        }
        demande_ref_map = _lookup_reference_map(conn, "demandes", existing_demande_ids)
        campagne_ref_map = _lookup_reference_map(conn, "campagnes", existing_campagne_ids)
        intervention_ref_map = _lookup_reference_map(conn, "interventions", existing_intervention_ids)
        pmt_ref_map = _lookup_reference_map(conn, "pmt_essais", existing_pmt_ids)
        for sheet in preview["sheets"]:
            existing_demande_id = sheet.get("existing_demande_id")
            existing_campagne_id = sheet.get("existing_campagne_id")
            existing_intervention_id = sheet.get("existing_intervention_id")
            existing_pmt_id = sheet.get("existing_pmt_id")
            if existing_demande_id not in (None, ""):
                sheet["existing_demande_reference"] = demande_ref_map.get(int(existing_demande_id), "")
            if existing_campagne_id not in (None, ""):
                sheet["existing_campagne_reference"] = campagne_ref_map.get(int(existing_campagne_id), "")
            if existing_intervention_id not in (None, ""):
                sheet["existing_intervention_reference"] = intervention_ref_map.get(int(existing_intervention_id), "")
            if existing_pmt_id not in (None, ""):
                sheet["existing_essai_reference"] = pmt_ref_map.get(int(existing_pmt_id), "")

        affaire_candidates = [
            _sanitize_affaire_reference(sheet.get("reference_affaire"))
            for sheet in preview["sheets"]
            if _sanitize_affaire_reference(sheet.get("reference_affaire"))
        ]
        detected_affaire_nge = sorted({_normalize_affaire_nge(candidate) for candidate in affaire_candidates if _normalize_affaire_nge(candidate)})

        selected_affaire_reference = _clean(affaire_reference) or (affaire_candidates[0] if affaire_candidates else "")
        selected_affaire_nge = _clean(affaire_nge) or (detected_affaire_nge[0] if detected_affaire_nge else "")
        affaire_context = _resolve_affaire_context(
            conn,
            affaire_reference=selected_affaire_reference,
            affaire_nge=selected_affaire_nge,
        )
        proposals = _build_preview_proposals(
            conn,
            preview["sheets"],
            demande_gap_days=demande_gap_days,
            campagne_gap_days=campagne_gap_days,
        )
        sheet_to_intervention_ref: dict[str, str] = {}
        for demande in proposals.get("demandes", []):
            for campagne in (demande.get("campagnes") or []):
                for sheet_name, ref in (campagne.get("predicted_intervention_references") or {}).items():
                    clean_sheet = _clean(sheet_name)
                    clean_ref = _clean(ref)
                    if clean_sheet and clean_ref:
                        sheet_to_intervention_ref[clean_sheet] = clean_ref
        next_seq_by_year = _load_next_pmt_sequence_by_year(conn)
        for sheet in preview["sheets"]:
            sheet_name = _clean(sheet.get("import_source_sheet"))
            predicted_intervention_reference = sheet_to_intervention_ref.get(sheet_name, "")
            sheet["predicted_intervention_reference"] = predicted_intervention_reference
            if sheet.get("already_imported"):
                sheet["predicted_essai_reference"] = (
                    _clean(sheet.get("existing_essai_reference"))
                    or _clean(sheet.get("import_source_sheet"))
                    or _clean(sheet.get("import_uid"))
                )
            else:
                year = _extract_year_from_sheet(sheet)
                next_seq = int(next_seq_by_year.get(year, 1))
                sheet["predicted_essai_reference"] = f"{year}-SP-PMT{next_seq:04d}"
                next_seq_by_year[year] = next_seq + 1

    return {
        **preview,
        "already_imported_count": sum(1 for sheet in preview["sheets"] if sheet.get("already_imported")),
        "affaire_nge_detected": detected_affaire_nge,
        "affaire_context": affaire_context,
        "proposals": proposals,
        "auto_defaults": {
            "affaire_reference_suggested": selected_affaire_reference,
            "affaire_nge_suggested": selected_affaire_nge,
            "demande_gap_days_suggested": int(demande_gap_days or 120),
            "campagne_gap_days_suggested": int(campagne_gap_days or 7),
        },
    }


@router.post("/import-upload")
async def import_upload(
    file: UploadFile = File(...),
    sheet_name: str = Form(default=""),
    demande_id: int | None = Form(default=None),
    campagne_id: int | None = Form(default=None),
    intervention_id: int | None = Form(default=None),
):
    name = _clean(file.filename) or "pmt.xlsx"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Fichier vide.")

    preview = _preview_workbook(name, content)
    imported: list[dict[str, Any]] = []
    selected_sheet_name = _clean(sheet_name)
    stats = {
        "affaires_rst": {"created": set(), "existing": set()},
        "demandes": {"created": set(), "existing": set()},
        "campagnes": {"created": set(), "existing": set()},
        "interventions": {"created": set(), "existing": set()},
    }

    with _conn() as conn:
        sheets_to_import = preview["sheets"]
        if selected_sheet_name:
            sheets_to_import = [
                s for s in sheets_to_import
                if _clean(s.get("import_source_sheet")) == selected_sheet_name
            ]
            if not sheets_to_import:
                raise HTTPException(status_code=400, detail=f"Feuille PMT introuvable: {selected_sheet_name}")

        for sheet in sheets_to_import:
            points = list(sheet.pop("points_rows", []))
            summary = dict(sheet.pop("summary", {}))
            now = datetime.utcnow().isoformat(timespec="seconds")
            affaire_context, affaire_created = _ensure_affaire_context(conn, sheet)
            affaire_id = int(affaire_context["selected"]["id"])
            existing_before_upsert = conn.execute(
                """
                SELECT id, demande_id, campaign_id, intervention_id, reference
                FROM pmt_essais
                WHERE import_source_file = ? AND import_source_sheet = ?
                """,
                (sheet["import_source_file"], sheet["import_source_sheet"]),
            ).fetchone()

            anchor_date = datetime.utcnow().date()
            date_text = _clean(sheet.get("date_essai_debut"))
            if re.match(r"^\d{4}-\d{2}-\d{2}$", date_text):
                anchor_date = date.fromisoformat(date_text)
            # Do not reuse hierarchy IDs from a previous pmt_essais row unless the caller
            # explicitly passes demande_id / campagne_id / intervention_id (bind mode).
            # Otherwise ensure_hierarchy re-applies 120/7 temporal matching on each import.
            hierarchy = ensure_hierarchy(
                conn,
                affaire_context=affaire_context,
                anchor_date=anchor_date,
                demande_id=demande_id,
                campagne_id=campagne_id,
                intervention_id=intervention_id,
                import_profile_label="PMT import",
            )
            ensure_modules_enabled(conn, int(hierarchy["demande_id"]), ["interventions", "essais_terrain"])

            _track_hierarchy_stat(stats, "affaires_rst", affaire_id, affaire_created)
            _track_hierarchy_stat(
                stats,
                "demandes",
                int(hierarchy["demande_id"]),
                bool(hierarchy["created"]["demande"]),
            )
            _track_hierarchy_stat(
                stats,
                "campagnes",
                int(hierarchy["campagne_id"]),
                bool(hierarchy["created"]["campagne"]),
            )
            _track_hierarchy_stat(
                stats,
                "interventions",
                int(hierarchy["intervention_id"]),
                bool(hierarchy["created"]["intervention"]),
            )

            existing_reference = _clean(existing_before_upsert["reference"]) if existing_before_upsert else ""
            if _is_pmt_reference_format(existing_reference):
                reference_to_use = existing_reference
            else:
                reference_to_use = _next_pmt_reference(conn, _extract_year_from_sheet(sheet, fallback_year=anchor_date.year))

            payload = {
                **sheet,
                "campaign_id": int(hierarchy["campagne_id"]),
                "demande_id": int(hierarchy["demande_id"]),
                "intervention_id": int(hierarchy["intervention_id"]),
                "reference": reference_to_use,
                "statut": "Importé",
                "date_essai": sheet.get("date_essai_debut") or "",
                "imported_at": now,
                "updated_at": now,
                "observations": _clean(sheet.get("commentaire")),
                "donnees_entete_json": json.dumps(sheet.get("donnees_entete_json", {}), ensure_ascii=False),
                "donnees_synthese_json": json.dumps(summary, ensure_ascii=False),
                "resultats_json": json.dumps({"points": points, "summary": summary}, ensure_ascii=False),
            }

            columns = [
                "campaign_id", "demande_id", "intervention_id", "reference", "statut", "date_essai",
                "operateur", "section_controlee", "observations", "resultats_json",
                "import_source_file", "import_source_sheet", "import_uid", "imported_at",
                "code_essai", "norme", "chrono", "reference_affaire", "date_redaction", "laboratoire",
                "produit_controle", "numero_formule", "couche", "lieu_fabrication",
                "date_essai_texte", "date_essai_debut", "date_essai_fin",
                "date_mise_en_oeuvre_texte", "date_mise_en_oeuvre_debut", "date_mise_en_oeuvre_fin",
                "epaisseur_couche_texte", "epaisseur_couche_cm", "conditions_meteorologiques",
                "atelier_mise_en_oeuvre", "volume_materiau_texte", "volume_materiau_mm3", "volume_materiau_cm3",
                "source_criteres", "definition_criteres", "seuil_pmt_min_mm", "pourcentage_conformite_min",
                "nombre_essais", "pmt_moyenne_mm", "pmt_min_mm", "pmt_max_mm", "pmt_ecart_type_mm",
                "pourcentage_valeurs_conformes", "nombre_points_conformes", "nombre_points_non_conformes",
                "conclusion_excel_texte", "conclusion_calculee", "conclusion_finale", "commentaire",
                "signataire_nom", "signataire_fonction", "visa_texte", "donnees_entete_json", "donnees_synthese_json",
                "updated_at",
            ]
            nullable_columns = {
                "campaign_id",
                "demande_id",
                "intervention_id",
                "essai_id",
                "imported_at",
                "epaisseur_couche_cm",
                "volume_materiau_mm3",
                "volume_materiau_cm3",
                "pourcentage_conformite_min",
                "seuil_pmt_min_mm",
                "nombre_essais",
                "pmt_moyenne_mm",
                "pmt_min_mm",
                "pmt_max_mm",
                "pmt_ecart_type_mm",
                "pourcentage_valeurs_conformes",
                "nombre_points_conformes",
                "nombre_points_non_conformes",
            }
            values: list[Any] = []
            for column in columns:
                value = payload.get(column)
                if value is None and column not in nullable_columns:
                    value = ""
                values.append(value)
            placeholders = ", ".join("?" for _ in columns)
            insert_columns = ", ".join(columns)
            update_assignments = ", ".join(f"{col}=excluded.{col}" for col in columns if col not in {"import_source_file", "import_source_sheet"})

            conn.execute(
                f"""
                INSERT INTO pmt_essais ({insert_columns})
                VALUES ({placeholders})
                ON CONFLICT(import_source_file, import_source_sheet) DO UPDATE SET
                    {update_assignments}
                """,
                values,
            )

            row = conn.execute(
                "SELECT id FROM pmt_essais WHERE import_source_file = ? AND import_source_sheet = ?",
                (sheet["import_source_file"], sheet["import_source_sheet"]),
            ).fetchone()
            if not row:
                raise HTTPException(status_code=500, detail="Echec de récupération de l'essai PMT upserté.")
            pmt_id = int(row["id"])

            conn.execute("DELETE FROM pmt_essais_points WHERE pmt_id = ?", (pmt_id,))
            for point in points:
                profondeur = point.get("profondeur_macrotexture_mm")
                seuil = sheet.get("seuil_pmt_min_mm")
                conforme = None
                ecart = None
                if isinstance(profondeur, (int, float)) and isinstance(seuil, (int, float)):
                    conforme = 1 if profondeur >= seuil else 0
                    ecart = float(profondeur) - float(seuil)
                conn.execute(
                    """
                    INSERT INTO pmt_essais_points (
                        pmt_id, ordre, numero_essai, profil, position, localisation,
                        diametre_moyen_tache_mm, profondeur_macrotexture_mm, observation,
                        volume_materiau_mm3, seuil_pmt_min_mm, conforme, ecart_au_seuil_mm,
                        donnees_ligne_json, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        pmt_id,
                        point.get("ordre") or 0,
                        _clean(point.get("numero_essai")),
                        _clean(point.get("profil")),
                        _clean(point.get("position")),
                        _clean(point.get("localisation")),
                        point.get("diametre_moyen_tache_mm"),
                        point.get("profondeur_macrotexture_mm"),
                        _clean(point.get("observation")),
                        sheet.get("volume_materiau_mm3"),
                        sheet.get("seuil_pmt_min_mm"),
                        conforme,
                        ecart,
                        json.dumps(point.get("donnees_ligne_json", {}), ensure_ascii=False),
                        now,
                    ),
                )

            imported.append(
                {
                    "pmt_id": pmt_id,
                    "action": "updated" if existing_before_upsert else "inserted",
                    "affaire_id": affaire_id,
                    "demande_id": int(hierarchy["demande_id"]),
                    "campagne_id": int(hierarchy["campagne_id"]),
                    "intervention_id": int(hierarchy["intervention_id"]),
                    "sheet": sheet["import_source_sheet"],
                    "reference": reference_to_use,
                    "date_essai": _clean(sheet.get("date_essai_debut")),
                    "import_uid": sheet["import_uid"],
                    "points_imported": len(points),
                    "hierarchy_created": {
                        "affaire": bool(affaire_created),
                        "demande": bool(hierarchy["created"]["demande"]),
                        "campagne": bool(hierarchy["created"]["campagne"]),
                        "intervention": bool(hierarchy["created"]["intervention"]),
                    },
                }
            )

        conn.commit()

    hierarchy_summary = {
        entity: {
            "created": len(values["created"]),
            "existing": len(values["existing"]),
        }
        for entity, values in stats.items()
    }
    return {
        "ok": True,
        "file": preview["file_name"],
        "imported_count": len(imported),
        "hierarchy_summary": hierarchy_summary,
        "imported": imported,
    }
