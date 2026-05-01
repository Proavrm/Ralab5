from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from io import BytesIO
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from pydantic import BaseModel, Field

from app.core.database import get_db_path
from api.point_code_logic import allocate_next_point_code_for_scope
from api.import_essais_base import (
    group_rows_by_temporal_gap,
    _predict_references as _base_predict_references,
    _resolve_affaire_context as _base_resolve_affaire_context,
    _next_demande_reference,
    _next_campaign_reference,
    _next_intervention_reference,
)

router = APIRouter()
DB_PATH = get_db_path()


class PreviewDERequest(BaseModel):
    file_path: str = Field(..., min_length=1)
    affaire_reference: str = Field(default="", max_length=64)
    affaire_nge: str = Field(default="", max_length=128)
    demande_gap_days: int = Field(default=120, ge=1, le=3650)
    campagne_gap_days: int = Field(default=7, ge=1, le=365)


class ImportDESheetRequest(BaseModel):
    file_path: str = Field(..., min_length=1)
    sheet_name: str = Field(..., min_length=1)
    affaire_reference: str = Field(default="", max_length=64)
    affaire_nge: str = Field(default="", max_length=128)
    demande_gap_days: int = Field(default=120, ge=1, le=3650)
    campagne_gap_days: int = Field(default=7, ge=1, le=365)
    demande_reference_override: str = Field(default="", max_length=64)
    campagne_reference_override: str = Field(default="", max_length=64)
    intervention_reference_override: str = Field(default="", max_length=64)


def _conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _display_value(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return _clean(value)


def _normalize_affaire_nge(value: str) -> str:
    return re.sub(r"\W+", "", _clean(value)).upper()


def _parse_date(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = _clean(value)
    if not text:
        return ""

    match = re.search(r"(\d{2})[/-](\d{2})[/-](20\d{2})", text)
    if match:
        day, month, year = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

    match = re.search(r"(20\d{2})[/-](\d{2})[/-](\d{2})", text)
    if match:
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

    return ""


def _join_cells(ws, cells: list[str]) -> str:
    values = [_clean(ws[cell].value) for cell in cells]
    return " ".join(value for value in values if value).strip()


def _find_label_row(ws, label: str, start_row: int = 1, end_row: int = 120, max_col: int = 24) -> int | None:
    wanted = _clean(label).upper()
    for row_index in range(start_row, end_row + 1):
        for col_index in range(1, max_col + 1):
            if _clean(ws.cell(row=row_index, column=col_index).value).upper() == wanted:
                return row_index
    return None


def _extract_commentaires(ws) -> str:
    comment_row = _find_label_row(ws, "COMMENTAIRES", start_row=50, end_row=80)
    if comment_row is None:
        return ""

    values: list[str] = []
    for row_index in range(comment_row + 1, min(comment_row + 4, ws.max_row) + 1):
        for col_index in range(2, 18):
            text = _clean(ws.cell(row=row_index, column=col_index).value)
            if not text or text.upper() == "COMMENTAIRES":
                continue
            values.append(text)
    return "\n".join(values).strip()


def _extract_conclusion(ws) -> str:
    conclusion_row = _find_label_row(ws, "CONCLUSIONS", start_row=50, end_row=80)
    if conclusion_row is None:
        return ""

    for row_index in range(conclusion_row + 1, min(conclusion_row + 4, ws.max_row) + 1):
        text = _clean(ws[f"A{row_index}"].value)
        if text and "CONFORM" in text.upper():
            return text
    return ""


def _extract_sheet_row(ws) -> dict[str, Any]:
    date_essai = _parse_date(ws["G11"].value)
    date_redaction = _parse_date(ws["P5"].value)
    anchor_date = date_essai or date_redaction
    date_essai_raw = _display_value(ws["G11"].value)
    date_mise_en_oeuvre = _display_value(ws["G13"].value)
    criteria_source = _join_cells(ws, [f"{col}21" for col in "CDEFGHIJKLMNOPQRSTUVWX"])
    criteria_definition = _join_cells(ws, [f"{col}22" for col in "CDEFGHIJKLMNOPQRSTUVWX"])
    conclusion = _extract_conclusion(ws)
    commentaires = _extract_commentaires(ws)
    moyenne_row = _find_label_row(ws, "Moyenne", start_row=45, end_row=60, max_col=4)
    conformite_row = _find_label_row(ws, "Pourcentage de valeurs conformes :", start_row=45, end_row=65, max_col=8)

    point_count = 0
    points_rows: list[dict[str, Any]] = []
    for row_index in range(29, 61):
        point_label = _clean(ws[f"B{row_index}"].value).upper()
        if point_label in {
            "MOYENNE",
            "POURCENTAGE DE VALEURS CONFORMES :",
            "POURCENTAGE DE VALEURS CONFORMES",
        }:
            break
        line = {
            "row_index": row_index,
            "point": _clean(ws[f"B{row_index}"].value),
            "profil": _clean(ws[f"D{row_index}"].value),
            "position_g": _clean(ws[f"F{row_index}"].value),
            "position_a": _clean(ws[f"G{row_index}"].value),
            "position_d": _clean(ws[f"H{row_index}"].value),
            "position": " / ".join(
                value
                for value in (
                    _clean(ws[f"F{row_index}"].value),
                    _clean(ws[f"G{row_index}"].value),
                    _clean(ws[f"H{row_index}"].value),
                )
                if value
            ),
            "masse_volumique": ws[f"I{row_index}"].value,
            "compacite_pct": ws[f"L{row_index}"].value,
            "vides_pct": ws[f"O{row_index}"].value,
            "observations": _clean(ws[f"R{row_index}"].value),
        }
        has_value = any(
            line[key] not in (None, "")
            for key in (
                "point",
                "profil",
                "position_g",
                "position_a",
                "position_d",
                "position",
                "masse_volumique",
                "compacite_pct",
                "vides_pct",
                "observations",
            )
        )
        if has_value:
            point_count += 1
            points_rows.append(line)

    return {
        "sheet": ws.title,
        "affaire_nge_raw": _clean(ws["L5"].value),
        "affaire_nge_normalized": _normalize_affaire_nge(_clean(ws["L5"].value)),
        "chrono": _clean(ws["I5"].value),
        "date_essai": date_essai,
        "date_essai_raw": date_essai_raw or date_essai,
        "date_redaction": date_redaction,
        "anchor_date": anchor_date,
        "operateur": _clean(ws["G10"].value),
        "produit_controle": _clean(ws["S10"].value),
        "numero_formule": _clean(ws["P11"].value),
        "couche": _clean(ws["G12"].value),
        "epaisseur_couche_cm": _clean(ws["X11"].value),
        "date_mise_en_oeuvre": date_mise_en_oeuvre,
        "lieu_fabrication": _clean(ws["S12"].value),
        "section_controlee": _clean(ws["S13"].value),
        "gammadensimetre": _clean(ws["G14"].value),
        "date_dernier_calibrage": _display_value(ws["G15"].value),
        "conditions_meteo": _clean(ws["S14"].value),
        "profondeur_mesure": _clean(ws["S15"].value),
        "atelier_mise_en_oeuvre": _clean(ws["G16"].value),
        "criteria_source": criteria_source,
        "criteria_definition": criteria_definition,
        "criteria_vides_min": _display_value(ws["N22"].value),
        "criteria_vides_max": _display_value(ws["R22"].value),
        "mvre": ws["Q25"].value,
        "points": point_count,
        "moyenne_mv": ws[f"I{moyenne_row}"].value if moyenne_row else ws["I50"].value,
        "moyenne_compacite_pct": ws[f"L{moyenne_row}"].value if moyenne_row else ws["L50"].value,
        "moyenne_vides_pct": ws[f"O{moyenne_row}"].value if moyenne_row else ws["O50"].value,
        "conformite_pct": ws[f"I{conformite_row}"].value if conformite_row else ws["I51"].value,
        "conclusion": conclusion,
        "commentaires": commentaires,
        "points_rows": points_rows,
    }


def _group_by_gap(rows: list[dict[str, Any]], gap_days: int) -> list[list[dict[str, Any]]]:
    valid_rows = [row for row in rows if row.get("anchor_date")]
    valid_rows.sort(key=lambda row: (row.get("anchor_date") or "9999-99-99", row.get("sheet", "")))
    return group_rows_by_temporal_gap(valid_rows, gap_days, date_field="anchor_date")


def _build_imported_sheet_map(
    conn: sqlite3.Connection,
    *,
    file_hash: str,
    sheet_names: list[str],
) -> dict[str, dict[str, Any]]:
    if not file_hash or not sheet_names:
        return {}

    signatures = [f"DE_IMPORT|ESSAI|{file_hash}|{_clean(name)}" for name in sheet_names if _clean(name)]
    if not signatures:
        return {}

    placeholders = ",".join("?" for _ in signatures)
    rows = conn.execute(
        f"""
        SELECT id, intervention_id, source_signature
        FROM essais
        WHERE source_signature IN ({placeholders})
        """,
        tuple(signatures),
    ).fetchall()

    by_sheet: dict[str, dict[str, Any]] = {}
    for row in rows:
        signature = _clean(row["source_signature"])
        parts = signature.split("|", 3)
        sheet_name = parts[3] if len(parts) == 4 else ""
        if not sheet_name:
            continue
        by_sheet[sheet_name] = {
            "imported": True,
            "essai_id": int(row["id"]),
            "intervention_id": int(row["intervention_id"]) if row["intervention_id"] is not None else None,
        }
    return by_sheet


def _build_preview_response(
    workbook,
    *,
    source_name: str,
    file_hash: str,
    affaire_reference: str,
    affaire_nge: str,
    demande_gap_days: int,
    campagne_gap_days: int,
) -> dict[str, Any]:
    sheet_rows = [_extract_sheet_row(ws) for ws in workbook.worksheets]
    sheet_rows.sort(key=lambda row: (row.get("anchor_date") or "9999-99-99", row.get("sheet", "")))

    normalized_from_sheets = sorted({row["affaire_nge_normalized"] for row in sheet_rows if row["affaire_nge_normalized"]})
    inferred_affaire_nge = affaire_nge or (normalized_from_sheets[0] if len(normalized_from_sheets) == 1 else "")

    max_essai_id = 0
    has_essai_reference_column = False
    existing_essai_reference_by_id: dict[int, str] = {}

    with _conn() as conn:
        affaire_context = _base_resolve_affaire_context(conn, affaire_reference, inferred_affaire_nge)
        essais_cols = {col[1] for col in conn.execute("PRAGMA table_info(essais)").fetchall()}
        has_essai_reference_column = "reference" in essais_cols
        max_essai_id = int(conn.execute("SELECT COALESCE(MAX(id), 0) FROM essais").fetchone()[0] or 0)
        imported_sheet_map = _build_imported_sheet_map(
            conn,
            file_hash=file_hash,
            sheet_names=[_clean(row.get("sheet")) for row in sheet_rows],
        )

        existing_essai_ids = sorted(
            {
                int(info.get("essai_id"))
                for info in imported_sheet_map.values()
                if info.get("essai_id") is not None
            }
        )
        if has_essai_reference_column and existing_essai_ids:
            placeholders = ",".join("?" for _ in existing_essai_ids)
            rows = conn.execute(
                f"SELECT id, reference FROM essais WHERE id IN ({placeholders})",
                tuple(existing_essai_ids),
            ).fetchall()
            existing_essai_reference_by_id = {
                int(row["id"]): _clean(row["reference"])
                for row in rows
            }

        for row in sheet_rows:
            imported_info = imported_sheet_map.get(_clean(row.get("sheet"))) or {}
            row["already_imported"] = bool(imported_info.get("imported"))
            row["existing_essai_id"] = imported_info.get("essai_id")
            row["existing_intervention_id"] = imported_info.get("intervention_id")

        preview_candidate_rows = [row for row in sheet_rows if not row.get("already_imported")]
        demande_groups = _group_by_gap(preview_candidate_rows, demande_gap_days)
        predictions = _predict_references_for_preview(conn, demande_groups, campagne_gap_days)

    demande_proposals: list[dict[str, Any]] = []
    sheet_to_i_ref: dict[str, str] = {}
    for index, (demande_group, d_pred) in enumerate(zip(demande_groups, predictions), start=1):
        campagne_groups = _group_by_gap(demande_group, campagne_gap_days)
        campagne_proposals: list[dict[str, Any]] = []
        for c_index, (campagne_group, c_pred) in enumerate(zip(campagne_groups, d_pred.get("campagnes", [])), start=1):
            raw_interventions = c_pred.get("interventions", [])
            sheet_rows_by_name = {
                _clean(row.get("sheet")): row
                for row in campagne_group
                if _clean(row.get("sheet"))
            }
            date_to_ref: dict[str, str] = {}
            i_pred_map: dict[str, str] = {}
            for ip in raw_interventions:
                sheet_name = _clean(ip.get("sheet"))
                predicted_ref = _clean(ip.get("predicted_intervention_reference"))
                row = sheet_rows_by_name.get(sheet_name) or {}
                anchor_date = _clean(row.get("anchor_date"))
                if anchor_date and anchor_date not in date_to_ref:
                    date_to_ref[anchor_date] = predicted_ref
                if anchor_date:
                    i_pred_map[sheet_name] = date_to_ref[anchor_date]
                else:
                    i_pred_map[sheet_name] = predicted_ref
            for row in campagne_group:
                sheet_name = _clean(row.get("sheet"))
                anchor_date = _clean(row.get("anchor_date"))
                predicted_ref = date_to_ref.get(anchor_date) if anchor_date else ""
                if not predicted_ref:
                    predicted_ref = i_pred_map.get(sheet_name, "")
                i_pred_map[sheet_name] = predicted_ref
                sheet_to_i_ref[sheet_name] = predicted_ref
            unique_dates = {
                _clean(row.get("anchor_date"))
                for row in campagne_group
                if _clean(row.get("anchor_date"))
            }
            campagne_proposals.append(
                {
                    "proposal_index": c_index,
                    "start_date": campagne_group[0]["anchor_date"],
                    "end_date": campagne_group[-1]["anchor_date"],
                    "interventions_count": len(unique_dates),
                    "sheets": [row["sheet"] for row in campagne_group],
                    "predicted_campagne_reference": c_pred.get("predicted_campagne_reference", ""),
                    "predicted_intervention_references": i_pred_map,
                }
            )

        imported_count = sum(1 for r in demande_group if r.get("already_imported"))
        demande_proposals.append(
            {
                "proposal_index": index,
                "start_date": demande_group[0]["anchor_date"],
                "end_date": demande_group[-1]["anchor_date"],
                "interventions_count": sum(int(c.get("interventions_count") or 0) for c in campagne_proposals),
                "campagnes_count": len(campagne_proposals),
                "campagnes": campagne_proposals,
                "sheets": [row["sheet"] for row in demande_group],
                "predicted_demande_reference": d_pred.get("predicted_demande_reference", ""),
                "imported_count": imported_count,
            }
        )

    sheet_to_e_ref: dict[str, str] = {}
    next_preview_essai_id = max_essai_id

    for row in sheet_rows:
        row["predicted_intervention_reference"] = sheet_to_i_ref.get(_clean(row.get("sheet")), "")
        row["existing_essai_reference"] = ""

        if row.get("already_imported"):
            existing_essai_id = row.get("existing_essai_id")
            if existing_essai_id is not None:
                row["existing_essai_reference"] = existing_essai_reference_by_id.get(int(existing_essai_id), "")
            row["predicted_essai_reference"] = row.get("existing_essai_reference") or ""
            continue

        next_preview_essai_id += 1
        sheet_name = _clean(row.get("sheet"))
        sheet_to_e_ref[sheet_name] = _build_de_essai_reference_from_intervention_ref(
            row.get("predicted_intervention_reference", ""),
            row,
            next_preview_essai_id,
        )
        row["predicted_essai_reference"] = sheet_to_e_ref.get(sheet_name, "")

    return {
        "mode": "preview_only",
        "file_name": source_name,
        "file_hash": file_hash,
        "sheet_count": len(sheet_rows),
        "already_imported_count": sum(1 for row in sheet_rows if row.get("already_imported")),
        "affaire_nge_detected": normalized_from_sheets,
        "auto_defaults": {
            "affaire_nge_suggested": inferred_affaire_nge,
            "affaire_reference_suggested": (affaire_context.get("selected") or {}).get("reference", ""),
            "demande_gap_days_suggested": 120,
            "campagne_gap_days_suggested": 7,
        },
        "affaire_context": affaire_context,
        "params": {
            "demande_gap_days": demande_gap_days,
            "campagne_gap_days": campagne_gap_days,
        },
        "proposals": {
            "demandes_count": len(demande_proposals),
            "demandes": demande_proposals,
        },
        "sheets": sheet_rows,
    }


def _predict_references_for_preview(
    conn: sqlite3.Connection,
    demande_groups: list[list[dict]],
    campagne_gap_days: int,
    labo_code: str = "SP",
) -> list[dict]:
    """Predict references using the shared global engine, preserving DE response keys."""
    base_groups: list[list[dict[str, Any]]] = []
    for group in demande_groups:
        base_group: list[dict[str, Any]] = []
        for row in group:
            base_group.append(
                {
                    "sheet_name": row.get("sheet"),
                    "date_sondage": row.get("anchor_date"),
                }
            )
        base_groups.append(base_group)

    base_predictions = _base_predict_references(
        conn,
        base_groups,
        campagne_gap_days,
        labo_code=labo_code,
    )

    result: list[dict[str, Any]] = []
    for d_pred in base_predictions:
        campagnes: list[dict[str, Any]] = []
        for c_pred in d_pred.get("campagnes", []):
            interventions = [
                {
                    "sheet": i_pred.get("sheet_name"),
                    "predicted_intervention_reference": i_pred.get("predicted_intervention_reference", ""),
                }
                for i_pred in c_pred.get("interventions", [])
            ]
            campagnes.append(
                {
                    "predicted_campagne_reference": c_pred.get("predicted_campagne_reference", ""),
                    "interventions": interventions,
                }
            )
        result.append(
            {
                "predicted_demande_reference": d_pred.get("predicted_demande_reference", ""),
                "campagnes": campagnes,
            }
        )

    return result


def _find_sheet_group_indices(
    sheet_rows: list[dict[str, Any]],
    sheet_name: str,
    demande_gap_days: int,
    campagne_gap_days: int,
) -> tuple[int, int]:
    demande_groups = _group_by_gap(sheet_rows, demande_gap_days)
    for demande_index, demande_group in enumerate(demande_groups, start=1):
        campagne_groups = _group_by_gap(demande_group, campagne_gap_days)
        for campagne_index, campagne_group in enumerate(campagne_groups, start=1):
            for row in campagne_group:
                if _clean(row.get("sheet")) == _clean(sheet_name):
                    return demande_index, campagne_index
    raise HTTPException(status_code=400, detail=f"Feuille introuvable dans les groupes: {sheet_name}")


def _find_or_create_demande(
    conn: sqlite3.Connection,
    affaire_rst_id: int,
    demande_index: int,
    demande_group: dict[str, Any],
    reference_override: str = "",
) -> tuple[int, bool]:
    signature = f"DE_IMPORT|DEMANDE_GROUP|{demande_index}|{demande_group['start_date']}|{demande_group['end_date']}"
    row = conn.execute(
        """
        SELECT id
        FROM demandes
        WHERE affaire_rst_id = ?
          AND nature = 'Import DE gammadensimetre'
          AND observations LIKE ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (affaire_rst_id, f"%{signature}%"),
    ).fetchone()
    if row:
        return int(row["id"]), False

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    reception_date = demande_group["start_date"] or datetime.now().strftime("%Y-%m-%d")
    annee = int(reception_date[0:4]) if reception_date else datetime.now().year
    observations_payload = {
        "type": "DE_IMPORT_GROUP",
        "signature": signature,
        "source": "import_essais_de",
        "start_date": demande_group["start_date"],
        "end_date": demande_group["end_date"],
        "sheets": demande_group.get("sheets") or [],
    }
    override_ref = _clean(reference_override)
    if override_ref:
        match = re.match(r"^(\d{4})-([A-Za-z0-9]+)-D(\d+)$", override_ref)
        if not match:
            raise HTTPException(status_code=400, detail=f"Référence Demande invalide: {override_ref}")
        override_year = int(match.group(1))
        override_labo = match.group(2).upper()
        override_num = int(match.group(3))
        if conn.execute("SELECT 1 FROM demandes WHERE reference = ? LIMIT 1", (override_ref,)).fetchone():
            raise HTTPException(status_code=409, detail=f"Référence Demande déjà utilisée: {override_ref}")
        conn.execute(
            """
            INSERT INTO demandes
            (reference, annee, labo_code, numero, affaire_rst_id,
             numero_dst, type_mission, nature, domaine_etude, type_prestation_attendue,
             documents_fournis, lien_pieces_jointes, service_interne, societe_interne, urgence_source,
             description, observations,
             demandeur, date_reception, date_echeance, statut, priorite,
             a_revoir, note_reconciliation, suivi_notes, dossier_nom, dossier_path,
             rapport_ref, devis_ref, facture_ref, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                override_ref,
                override_year,
                override_labo,
                override_num,
                affaire_rst_id,
                "",
                "routine",
                "Import DE gammadensimetre",
                "Enrobes",
                "Feuille d'essai DE",
                "",
                "",
                "",
                "",
                "",
                f"Import DE automatique - groupe {demande_index}",
                json.dumps(observations_payload, ensure_ascii=False),
                "Import Outils",
                reception_date,
                None,
                "en_cours",
                "normale",
                0,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                now,
                now,
            ),
        )
        uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return int(uid), True

    last_reference = ""
    for _ in range(50):
        reference, ref_year, ref_num = _next_demande_reference(conn, labo_code="SP", annee=annee)
        last_reference = reference
        try:
            conn.execute(
                """
                INSERT INTO demandes
                (reference, annee, labo_code, numero, affaire_rst_id,
                 numero_dst, type_mission, nature, domaine_etude, type_prestation_attendue,
                 documents_fournis, lien_pieces_jointes, service_interne, societe_interne, urgence_source,
                 description, observations,
                 demandeur, date_reception, date_echeance, statut, priorite,
                 a_revoir, note_reconciliation, suivi_notes, dossier_nom, dossier_path,
                 rapport_ref, devis_ref, facture_ref, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    reference,
                    ref_year,
                    "SP",
                    ref_num,
                    affaire_rst_id,
                    "",
                    "routine",
                    "Import DE gammadensimetre",
                    "Enrobes",
                    "Feuille d'essai DE",
                    "",
                    "",
                    "",
                    "",
                    "",
                    f"Import DE automatique - groupe {demande_index}",
                    json.dumps(observations_payload, ensure_ascii=False),
                    "Import Outils",
                    reception_date,
                    None,
                    "en_cours",
                    "normale",
                    0,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    now,
                    now,
                ),
            )
            uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            return int(uid), True
        except sqlite3.IntegrityError as exc:
            # Rare race/collision on unique reference, regenerate and retry.
            if "demandes.reference" not in str(exc):
                raise

    raise HTTPException(
        status_code=409,
        detail=f"Impossible de générer une référence Demande unique (dernière tentative: {last_reference or 'n/a'})",
    )


def _find_or_create_campagne(
    conn: sqlite3.Connection,
    demande_id: int,
    demande_index: int,
    campagne_index: int,
    campagne_group: dict[str, Any],
    reference_override: str = "",
) -> tuple[int, bool]:
    signature = (
        f"DE_IMPORT|CAMPAGNE_GROUP|D{demande_index}|C{campagne_index}|"
        f"{campagne_group['start_date']}|{campagne_group['end_date']}"
    )
    row = conn.execute(
        """
        SELECT id
        FROM campagnes
        WHERE demande_id = ?
          AND code = 'DE'
          AND notes LIKE ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (demande_id, f"%{signature}%"),
    ).fetchone()
    if row:
        return int(row["id"]), False

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    notes_payload = {
        "type": "DE_IMPORT_CAMPAGNE_GROUP",
        "signature": signature,
        "source": "import_essais_de",
        "start_date": campagne_group["start_date"],
        "end_date": campagne_group["end_date"],
        "sheets": campagne_group.get("sheets") or [],
    }
    override_ref = _clean(reference_override)
    if override_ref:
        if not re.match(r"^\d{4}-[A-Za-z0-9]+-C\d+$", override_ref):
            raise HTTPException(status_code=400, detail=f"Référence Campagne invalide: {override_ref}")
        if conn.execute("SELECT 1 FROM campagnes WHERE reference = ? LIMIT 1", (override_ref,)).fetchone():
            raise HTTPException(status_code=409, detail=f"Référence Campagne déjà utilisée: {override_ref}")
        conn.execute(
            """
            INSERT INTO campagnes (
                demande_id, reference, label, type_campagne, code, designation,
                zone_scope, temporalite, programme_specifique, nb_points_prevus,
                types_essais_prevus, date_debut_prevue, date_fin_prevue, priorite,
                responsable_technique, attribue_a, criteres_controle,
                livrables_attendus, workflow_label, statut, notes, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                demande_id,
                override_ref,
                f"Import DE D{demande_index} C{campagne_index}",
                "DE",
                "DE",
                "Densite enrobes au gammadensimetre",
                "",
                "ponctuelle",
                "",
                campagne_group.get("interventions_count") or 0,
                "DE",
                campagne_group.get("start_date") or None,
                campagne_group.get("end_date") or None,
                "normale",
                "",
                "",
                "",
                "Rapport DE",
                "Affaire -> Demande -> Campagne -> Intervention",
                "a_faire",
                json.dumps(notes_payload, ensure_ascii=False),
                now,
            ),
        )
        uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return int(uid), True

    for _ in range(6):
        reference = _next_campaign_reference(conn, demande_id)
        try:
            conn.execute(
                """
                INSERT INTO campagnes (
                    demande_id, reference, label, type_campagne, code, designation,
                    zone_scope, temporalite, programme_specifique, nb_points_prevus,
                    types_essais_prevus, date_debut_prevue, date_fin_prevue, priorite,
                    responsable_technique, attribue_a, criteres_controle,
                    livrables_attendus, workflow_label, statut, notes, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    demande_id,
                    reference,
                    f"Import DE D{demande_index} C{campagne_index}",
                    "DE",
                    "DE",
                    "Densite enrobes au gammadensimetre",
                    "",
                    "ponctuelle",
                    "",
                    campagne_group.get("interventions_count") or 0,
                    "DE",
                    campagne_group.get("start_date") or None,
                    campagne_group.get("end_date") or None,
                    "normale",
                    "",
                    "",
                    "",
                    "Rapport DE",
                    "Affaire -> Demande -> Campagne -> Intervention",
                    "a_faire",
                    json.dumps(notes_payload, ensure_ascii=False),
                    now,
                ),
            )
            uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            return int(uid), True
        except sqlite3.IntegrityError as exc:
            if "campagnes.reference" not in str(exc):
                raise

    raise HTTPException(status_code=409, detail="Impossible de générer une référence Campagne unique")


def _build_resultats_payload(sheet_row: dict[str, Any], file_name: str, file_hash: str) -> dict[str, Any]:
    return {
        "source": "de_excel_import",
        "version": 1,
        "file_name": file_name,
        "file_hash": file_hash,
        "sheet": sheet_row.get("sheet") or "",
        "meta": {
            "affaire_nge_raw": sheet_row.get("affaire_nge_raw") or "",
            "chrono": sheet_row.get("chrono") or "",
            "date_essai": sheet_row.get("date_essai_raw") or sheet_row.get("date_essai") or "",
            "date_redaction": sheet_row.get("date_redaction") or "",
            "operateur": sheet_row.get("operateur") or "",
            "produit_controle": sheet_row.get("produit_controle") or "",
            "numero_formule": sheet_row.get("numero_formule") or "",
            "couche": sheet_row.get("couche") or "",
            "epaisseur_couche_cm": sheet_row.get("epaisseur_couche_cm") or "",
            "date_mise_en_oeuvre": sheet_row.get("date_mise_en_oeuvre") or "",
            "lieu_fabrication": sheet_row.get("lieu_fabrication") or "",
            "section_controlee": sheet_row.get("section_controlee") or "",
            "gammadensimetre": sheet_row.get("gammadensimetre") or "",
            "date_dernier_calibrage": sheet_row.get("date_dernier_calibrage") or "",
            "conditions_meteo": sheet_row.get("conditions_meteo") or "",
            "profondeur_mesure": sheet_row.get("profondeur_mesure") or "",
            "atelier_mise_en_oeuvre": sheet_row.get("atelier_mise_en_oeuvre") or "",
            "criteria_source": sheet_row.get("criteria_source") or "",
            "criteria_definition": sheet_row.get("criteria_definition") or "",
            "criteria_vides_min": sheet_row.get("criteria_vides_min"),
            "criteria_vides_max": sheet_row.get("criteria_vides_max"),
            "mvre": sheet_row.get("mvre"),
            "conclusion": sheet_row.get("conclusion") or "",
            "commentaires": sheet_row.get("commentaires") or "",
        },
        "resume": {
            "points": sheet_row.get("points") or 0,
            "moyenne_mv": sheet_row.get("moyenne_mv"),
            "moyenne_compacite_pct": sheet_row.get("moyenne_compacite_pct"),
            "moyenne_vides_pct": sheet_row.get("moyenne_vides_pct"),
            "conformite_pct": sheet_row.get("conformite_pct"),
        },
        "points_rows": sheet_row.get("points_rows") or [],
    }


def _find_or_create_intervention(
    conn: sqlite3.Connection,
    demande_id: int,
    campagne_id: int,
    sheet_row: dict[str, Any],
    file_name: str,
    file_hash: str,
    reference_override: str = "",
) -> tuple[int, bool]:
    intervention_date = sheet_row.get("date_essai") or sheet_row.get("anchor_date") or datetime.now().strftime("%Y-%m-%d")
    signature = f"DE_IMPORT|INTERVENTION_DATE|{file_hash}|{intervention_date}"
    row = conn.execute(
        """
        SELECT id
        FROM interventions
        WHERE demande_id = ?
          AND campagne_id = ?
          AND date_intervention = ?
          AND observations LIKE ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (demande_id, campagne_id, intervention_date, f"%{file_hash}%"),
    ).fetchone()
    if row:
        return int(row["id"]), False

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet_signature = f"DE_IMPORT|INTERVENTION_SHEET|{file_hash}|{_clean(sheet_row.get('sheet'))}"
    observations_payload = {
        "type": "DE_IMPORT_INTERVENTION",
        "signature": signature,
        "sheet_signature": sheet_signature,
        "source": "import_essais_de",
        "file_name": file_name,
        "file_hash": file_hash,
        "sheet": sheet_row.get("sheet"),
        "chrono": sheet_row.get("chrono"),
        "couche": sheet_row.get("couche"),
        "section_controlee": sheet_row.get("section_controlee"),
        "anchor_date": intervention_date,
    }
    override_ref = _clean(reference_override)
    if override_ref:
        match = re.match(r"^(\d{4})-([A-Za-z0-9]+)-I(\d+)$", override_ref)
        if not match:
            raise HTTPException(status_code=400, detail=f"Référence Intervention invalide: {override_ref}")
        annee = int(match.group(1))
        labo = match.group(2).upper()
        numero = int(match.group(3))
        if conn.execute("SELECT 1 FROM interventions WHERE reference = ? LIMIT 1", (override_ref,)).fetchone():
            raise HTTPException(status_code=409, detail=f"Référence Intervention déjà utilisée: {override_ref}")
        conn.execute(
            """
            INSERT INTO interventions (
                reference, annee, labo_code, numero, demande_id, campagne_id,
                type_intervention, sujet, date_intervention, duree_heures,
                geotechnicien, technicien, observations, anomalie_detectee,
                niveau_alerte, pv_ref, rapport_ref, photos_dossier, statut,
                nature_reelle, finalite, zone, heure_debut, heure_fin,
                tri_updated_at, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                override_ref,
                annee,
                labo,
                numero,
                demande_id,
                campagne_id,
                "Controle DE gammadensimetre",
                f"{_clean(sheet_row.get('sheet'))} - {_clean(sheet_row.get('couche'))}".strip(" -"),
                intervention_date,
                None,
                _clean(sheet_row.get("operateur")),
                "",
                json.dumps(observations_payload, ensure_ascii=False),
                0,
                "info",
                "",
                "",
                "",
                "a_faire",
                "Intervention terrain",
                "Controle DE",
                _clean(sheet_row.get("section_controlee")),
                "",
                "",
                now,
                now,
                now,
            ),
        )
        uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return int(uid), True

    for _ in range(6):
        reference, annee, labo, numero = _next_intervention_reference(conn, demande_id)
        try:
            conn.execute(
                """
                INSERT INTO interventions (
                    reference, annee, labo_code, numero, demande_id, campagne_id,
                    type_intervention, sujet, date_intervention, duree_heures,
                    geotechnicien, technicien, observations, anomalie_detectee,
                    niveau_alerte, pv_ref, rapport_ref, photos_dossier, statut,
                    nature_reelle, finalite, zone, heure_debut, heure_fin,
                    tri_updated_at, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    reference,
                    annee,
                    labo,
                    numero,
                    demande_id,
                    campagne_id,
                    "Controle DE gammadensimetre",
                    f"{_clean(sheet_row.get('sheet'))} - {_clean(sheet_row.get('couche'))}".strip(" -"),
                    intervention_date,
                    None,
                    _clean(sheet_row.get("operateur")),
                    "",
                    json.dumps(observations_payload, ensure_ascii=False),
                    0,
                    "info",
                    "",
                    "",
                    "",
                    "a_faire",
                    "Intervention terrain",
                    "Controle DE",
                    _clean(sheet_row.get("section_controlee")),
                    "",
                    "",
                    now,
                    now,
                    now,
                ),
            )
            uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            return int(uid), True
        except sqlite3.IntegrityError as exc:
            if "interventions.reference" not in str(exc):
                raise

    raise HTTPException(status_code=409, detail="Impossible de générer une référence Intervention unique")


def _ensure_modules_enabled(conn: sqlite3.Connection, demande_id: int, module_codes: list[str]) -> None:
    """Enable given modules for a demande if not already present/enabled (upsert)."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for code in module_codes:
        conn.execute(
            """
            INSERT INTO demande_enabled_modules (demande_id, module_code, is_enabled, created_at, updated_at)
            VALUES (?, ?, 1, ?, ?)
            ON CONFLICT(demande_id, module_code) DO UPDATE SET is_enabled = 1, updated_at = excluded.updated_at
            """,
            (demande_id, code, now, now),
        )


def _build_wbs(
    affaire_ref: str,
    demande_ref: str,
    campagne_ref: str,
    intervention_ref: str,
    essai_id: int,
    essai_code: str = "DE",
) -> tuple[str, str]:
    """Return (wbs_full, wbs_short).

    wbs_full  : 2021-RA-002-2022-SP-D0007-2022-SP-C001-2022-SP-I0022-2022-SP-DE1027
    wbs_short : RA-SP-A2-D7-C1-I22-E1027
      - first lab  = affaire lab (from affaire_ref YYYY-LAB-NNN)
      - second lab = demande lab (originating lab, always, even if another lab executes)
      - numeric parts stripped of leading zeros for readability
    """
    # ── extract affaire lab + num ──────────────────────────────────────────
    m_aff = re.match(r"\d{4}-([A-Z]+)-(\d+)", affaire_ref)
    affaire_lab = m_aff.group(1) if m_aff else ""
    affaire_num = str(int(m_aff.group(2))) if m_aff else "0"

    # ── extract demande lab + num ──────────────────────────────────────────
    m_dem = re.match(r"\d{4}-([A-Z]+)-D(\d+)", demande_ref)
    demande_lab = m_dem.group(1) if m_dem else ""
    demande_num = str(int(m_dem.group(2))) if m_dem else "0"

    # ── extract campagne num ───────────────────────────────────────────────
    m_cam = re.match(r"\d{4}-[A-Z]+-C(\d+)", campagne_ref)
    camp_num = str(int(m_cam.group(1))) if m_cam else "0"

    # ── extract intervention num ───────────────────────────────────────────
    m_int = re.match(r"\d{4}-[A-Z]+-I(\d+)", intervention_ref)
    int_num = str(int(m_int.group(1))) if m_int else "0"

    wbs_full = f"{affaire_ref}-{demande_ref}-{campagne_ref}-{intervention_ref}-{demande_lab}{essai_code}{essai_id:04d}"
    wbs_short = f"{affaire_lab}-{demande_lab}-A{affaire_num}-D{demande_num}-C{camp_num}-I{int_num}-E{essai_id}"
    return wbs_full, wbs_short


def _build_de_essai_reference(
    conn: sqlite3.Connection,
    intervention_id: int,
    sheet_row: dict[str, Any],
    essai_id: int,
) -> str:
    """Build DE essai reference in format YYYY-LAB-DE####."""
    year = datetime.now().year
    labo = "SP"

    try:
        i_row = conn.execute("SELECT reference FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
        i_ref = _clean(i_row[0] if i_row and not isinstance(i_row, sqlite3.Row) else (i_row["reference"] if i_row else ""))
        match = re.match(r"^(\d{4})-([A-Z]+)-I\d+$", i_ref)
        if match:
            year = int(match.group(1))
            labo = match.group(2)
        else:
            parsed = _parse_date(sheet_row.get("date_essai") or sheet_row.get("anchor_date") or "")
            if parsed and len(parsed) >= 4 and parsed[:4].isdigit():
                year = int(parsed[:4])
    except Exception:
        parsed = _parse_date(sheet_row.get("date_essai") or sheet_row.get("anchor_date") or "")
        if parsed and len(parsed) >= 4 and parsed[:4].isdigit():
            year = int(parsed[:4])

    return f"{year}-{labo}-DE{int(essai_id):04d}"


def _build_de_essai_reference_from_intervention_ref(
    intervention_reference: str,
    sheet_row: dict[str, Any],
    essai_id: int,
) -> str:
    year = datetime.now().year
    labo = "SP"

    ref = _clean(intervention_reference)
    match = re.match(r"^(\d{4})-([A-Z]+)-I\d+$", ref)
    if match:
        year = int(match.group(1))
        labo = match.group(2)
    else:
        parsed = _parse_date(sheet_row.get("date_essai") or sheet_row.get("anchor_date") or "")
        if parsed and len(parsed) >= 4 and parsed[:4].isdigit():
            year = int(parsed[:4])

    return f"{year}-{labo}-DE{int(essai_id):04d}"


# ─── DE terrain helpers (terrain model) ──────────────────────────────────────

def _de_table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def _de_table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    if not _de_table_exists(conn, table_name):
        return set()
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}


def _next_de_feuille_reference(conn: sqlite3.Connection, year: int, labo_code: str) -> str:
    prefix = f"{year}-{labo_code}-DE"
    rows = conn.execute(
        "SELECT reference FROM feuilles_terrain WHERE reference LIKE ?",
        (f"{prefix}%",),
    ).fetchall()
    numbers: list[int] = []
    for row in rows:
        ref_val = str(row["reference"] if isinstance(row, sqlite3.Row) else row[0] or "").strip().upper()
        m = re.match(rf"^{re.escape(prefix)}(\d+)$", ref_val)
        if m:
            numbers.append(int(m.group(1)))
    return f"{prefix}{max(numbers, default=0) + 1:04d}"


def _next_de_point_reference(conn: sqlite3.Connection, year: int, labo_code: str) -> str:
    prefix = f"{year}-{labo_code}-DE"
    rows = conn.execute(
        "SELECT reference FROM points_terrain WHERE reference LIKE ?",
        (f"{prefix}%",),
    ).fetchall()
    numbers: list[int] = []
    for row in rows:
        ref_val = str(row["reference"] if isinstance(row, sqlite3.Row) else row[0] or "").strip().upper()
        m = re.match(rf"^{re.escape(prefix)}(\d+)$", ref_val)
        if m:
            numbers.append(int(m.group(1)))
    return f"{prefix}{max(numbers, default=0) + 1:04d}"


def _next_de_point_code_for_intervention(conn: sqlite3.Connection, intervention_id: int) -> str:
    """Generate next DE{n} point code within one intervention scope."""
    return allocate_next_point_code_for_scope(
        conn,
        'DE',
        intervention_id=int(intervention_id),
    )


def _find_or_create_de_terrain_point(
    conn: sqlite3.Connection,
    intervention_id: int,
    sheet_row: dict[str, Any],
    file_name: str,
    file_hash: str,
) -> tuple[int, int, bool]:
    """Create or find terrain feuille+point for one DE sheet.

    Returns (point_terrain_id, feuille_terrain_id, created).
    """
    sheet_name = _clean(sheet_row.get("sheet"))
    group_signature = f"DE_IMPORT|{file_hash}|{sheet_name}"

    # Idempotency: if a series for this exact sheet already exists, return existing point
    existing_serie = conn.execute(
        "SELECT id FROM series_essais_terrain WHERE group_signature = ? ORDER BY id DESC LIMIT 1",
        (group_signature,),
    ).fetchone()
    if existing_serie:
        existing_feuille = conn.execute(
            "SELECT id FROM feuilles_terrain WHERE serie_id = ? ORDER BY id DESC LIMIT 1",
            (int(existing_serie["id"]),),
        ).fetchone()
        existing_point = conn.execute(
            "SELECT id FROM points_terrain WHERE serie_id = ? ORDER BY id DESC LIMIT 1",
            (int(existing_serie["id"]),),
        ).fetchone()
        feuille_id_val = int(existing_feuille["id"]) if existing_feuille else 0
        point_id_val = int(existing_point["id"]) if existing_point else 0
        return point_id_val, feuille_id_val, False

    # Resolve year/labo from intervention reference
    intervention_row = conn.execute(
        """
        SELECT i.id, i.demande_id, i.campagne_id, i.reference, i.date_intervention,
               d.labo_code, d.reference AS demande_reference
        FROM interventions i
        JOIN demandes d ON d.id = i.demande_id
        WHERE i.id = ?
        """,
        (intervention_id,),
    ).fetchone()
    if not intervention_row:
        raise HTTPException(status_code=404, detail=f"Intervention #{intervention_id} introuvable")

    year_ref = datetime.now().year
    labo_ref = str(intervention_row["labo_code"] or "SP").strip().upper() or "SP"
    i_ref = _clean(intervention_row["reference"])
    m = re.match(r"^(\d{4})-([A-Z]+)-I\d+$", i_ref)
    if m:
        year_ref = int(m.group(1))
        labo_ref = m.group(2)
    else:
        date_str = _parse_date(sheet_row.get("date_essai") or sheet_row.get("anchor_date") or "")
        if date_str and len(date_str) >= 4 and date_str[:4].isdigit():
            year_ref = int(date_str[:4])

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    now_stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    code_feuille = "DE"
    resultats_payload = _build_resultats_payload(sheet_row, file_name, file_hash)
    compacite = sheet_row.get("moyenne_compacite_pct")
    try:
        principal = float(compacite) if compacite is not None else None
    except Exception:
        principal = None

    # Terrain tables now use interventions directly.
    series_int_id = intervention_id
    feuille_int_id = intervention_id
    point_int_id = intervention_id

    # 1. series_essais_terrain
    series_columns = _de_table_columns(conn, "series_essais_terrain")
    if not series_columns:
        raise HTTPException(status_code=400, detail="Table series_essais_terrain indisponible")

    series_values: dict[str, Any] = {
        "reference": f"SER-DE-{intervention_id}-{now_stamp}",
        "demande_id": int(intervention_row["demande_id"]),
        "campagne_id": intervention_row["campagne_id"],
        "intervention_id": series_int_id,
        "code_essai": code_feuille,
        "libelle_essai": f"Densité enrobés - {file_name}",
        "source_file": file_name,
        "sheet_name": sheet_name,
        "group_signature": group_signature,
        "import_mode": "de_excel_import",
        "statut": "Importée",
        "date_essai": sheet_row.get("date_essai") or sheet_row.get("anchor_date") or "",
        "operateur": _clean(sheet_row.get("operateur")),
        "section_controlee": _clean(sheet_row.get("section_controlee")),
        "couche": _clean(sheet_row.get("couche")),
        "observations": f"Import DE - {file_name}::{sheet_name}",
        "payload_json": json.dumps(resultats_payload, ensure_ascii=False),
        "created_at": now,
        "updated_at": now,
    }
    series_insert = {k: v for k, v in series_values.items() if k in series_columns}
    serie_id = conn.execute(
        f"INSERT INTO series_essais_terrain ({', '.join(series_insert.keys())}) "
        f"VALUES ({', '.join('?' for _ in series_insert)})",
        tuple(series_insert.values()),
    ).lastrowid

    # 2. feuilles_terrain
    feuille_reference = _next_de_feuille_reference(conn, year_ref, labo_ref)
    feuilles_columns = _de_table_columns(conn, "feuilles_terrain")
    feuille_values: dict[str, Any] = {
        "reference": feuille_reference,
        "demande_id": int(intervention_row["demande_id"]),
        "campagne_id": intervention_row["campagne_id"],
        "intervention_id": feuille_int_id,
        "serie_id": int(serie_id),
        "code_feuille": code_feuille,
        "label": f"DE - {_clean(sheet_row.get('section_controlee')) or sheet_name}",
        "norme": "NF P 98-241-1",
        "date_feuille": sheet_row.get("date_essai") or sheet_row.get("anchor_date") or "",
        "operateur": _clean(sheet_row.get("operateur")),
        "statut": "Importée",
        "observations": f"Import DE - {file_name}::{sheet_name}",
        "resultats_json": json.dumps(resultats_payload, ensure_ascii=False),
        "resultat_principal": principal,
        "resultat_unite": "%",
        "resultat_label": f"Compacite moyenne = {principal:.2f} %" if principal is not None else "",
        "created_at": now,
        "updated_at": now,
    }
    feuille_insert = {k: v for k, v in feuille_values.items() if k in feuilles_columns}
    feuille_uid = conn.execute(
        f"INSERT INTO feuilles_terrain ({', '.join(feuille_insert.keys())}) "
        f"VALUES ({', '.join('?' for _ in feuille_insert)})",
        tuple(feuille_insert.values()),
    ).lastrowid

    # 3. points_terrain (one point per sheet = one DE test location)
    point_reference = _next_de_point_reference(conn, year_ref, labo_ref)
    point_code = _next_de_point_code_for_intervention(conn, point_int_id)
    points_columns = _de_table_columns(conn, "points_terrain")
    point_values: dict[str, Any] = {
        "serie_id": int(serie_id),
        "intervention_id": point_int_id,
        "campagne_id": intervention_row["campagne_id"],
        "demande_id": int(intervention_row["demande_id"]),
        "reference": point_reference,
        "point_code": point_code,
        "point_type": "DENSITE_ENROBES",
        "ordre": 1,
        "localisation": _clean(sheet_row.get("section_controlee")),
        "position_label": _clean(sheet_row.get("section_controlee")),
        "valeur_principale": principal,
        "unite_principale": "%",
        "observation": f"Import DE - {sheet_name}",
        "payload_json": json.dumps(
            {
                "reference": point_reference,
                "point_code": point_code,
                "section_controlee": _clean(sheet_row.get("section_controlee")),
                "couche": _clean(sheet_row.get("couche")),
                "date_essai": sheet_row.get("date_essai") or sheet_row.get("anchor_date") or "",
                "operateur": _clean(sheet_row.get("operateur")),
                "moyenne_compacite_pct": principal,
                "resultats": resultats_payload,
            },
            ensure_ascii=False,
        ),
        "created_at": now,
    }
    point_insert = {k: v for k, v in point_values.items() if k in points_columns}
    point_uid = conn.execute(
        f"INSERT INTO points_terrain ({', '.join(point_insert.keys())}) "
        f"VALUES ({', '.join('?' for _ in point_insert)})",
        tuple(point_insert.values()),
    ).lastrowid

    return int(point_uid), int(feuille_uid), True


def _find_or_create_essai(
    conn: sqlite3.Connection,
    intervention_id: int,
    sheet_row: dict[str, Any],
    file_name: str,
    file_hash: str,
) -> tuple[int, bool]:
    source_signature = f"DE_IMPORT|ESSAI|{file_hash}|{_clean(sheet_row.get('sheet'))}"
    essais_cols = {col[1] for col in conn.execute("PRAGMA table_info(essais)").fetchall()}
    has_reference_column = "reference" in essais_cols
    select_sql = (
        "SELECT id, reference FROM essais WHERE intervention_id = ? AND source_signature = ? ORDER BY id DESC LIMIT 1"
        if has_reference_column
        else "SELECT id FROM essais WHERE intervention_id = ? AND source_signature = ? ORDER BY id DESC LIMIT 1"
    )
    row = conn.execute(select_sql, (intervention_id, source_signature)).fetchone()
    if row:
        essai_id = int(row["id"])
        if has_reference_column and not _clean(row["reference"]):
            conn.execute(
                "UPDATE essais SET reference = ?, updated_at = ? WHERE id = ?",
                (
                    _build_de_essai_reference(conn, intervention_id, sheet_row, essai_id),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    essai_id,
                ),
            )
        return essai_id, False

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    compacite = sheet_row.get("moyenne_compacite_pct")
    try:
        principal = float(compacite) if compacite is not None else None
    except Exception:
        principal = None
    resultat_label = ""
    if principal is not None:
        resultat_label = f"Compacite moyenne = {principal:.2f} %"
    resultats_payload = _build_resultats_payload(sheet_row, file_name, file_hash)

    conn.execute(
        """
        INSERT INTO essais
        (echantillon_id, intervention_id, essai_code, type_essai, norme, statut, date_debut, date_fin,
         resultats, operateur, observations, source_signature, source_label,
         resultat_principal, resultat_unite, resultat_label, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            None,
            intervention_id,
            "DE",
            "Densite enrobes",
            "NF P 98-241-1",
            "a_faire",
            sheet_row.get("date_essai") or sheet_row.get("anchor_date") or None,
            sheet_row.get("date_redaction") or None,
            json.dumps(resultats_payload, ensure_ascii=False),
            _clean(sheet_row.get("operateur")),
            f"Import DE - feuille {_clean(sheet_row.get('sheet'))}",
            source_signature,
            f"DE Excel {file_name}::{_clean(sheet_row.get('sheet'))}",
            principal,
            "%",
            resultat_label,
            now,
            now,
        ),
    )
    uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    if has_reference_column:
        conn.execute(
            "UPDATE essais SET reference = ?, updated_at = ? WHERE id = ?",
            (
                _build_de_essai_reference(conn, intervention_id, sheet_row, int(uid)),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                int(uid),
            ),
        )

    return int(uid), True


def _materialize_sheet_import(
    *,
    workbook,
    file_name: str,
    file_hash: str,
    sheet_name: str,
    affaire_reference: str,
    affaire_nge: str,
    demande_gap_days: int,
    campagne_gap_days: int,
    demande_reference_override: str = "",
    campagne_reference_override: str = "",
    intervention_reference_override: str = "",
) -> dict[str, Any]:
    sheet_rows = [_extract_sheet_row(ws) for ws in workbook.worksheets]
    if not sheet_rows:
        raise HTTPException(status_code=400, detail="Aucune feuille exploitable")

    target_sheet = next((row for row in sheet_rows if _clean(row.get("sheet")) == _clean(sheet_name)), None)
    if not target_sheet:
        raise HTTPException(status_code=404, detail=f"Feuille introuvable: {sheet_name}")
    if not target_sheet.get("anchor_date"):
        raise HTTPException(status_code=400, detail=f"Date manquante pour la feuille {sheet_name}")

    normalized_from_sheets = sorted({row["affaire_nge_normalized"] for row in sheet_rows if row["affaire_nge_normalized"]})
    inferred_affaire_nge = affaire_nge or (normalized_from_sheets[0] if len(normalized_from_sheets) == 1 else "")

    with _conn() as conn:
        affaire_context = _base_resolve_affaire_context(conn, affaire_reference, inferred_affaire_nge)
        selected_affaire = affaire_context.get("selected")
        if not selected_affaire:
            raise HTTPException(status_code=400, detail="Affaire RST non resolue. Fournir reference ou affaire NGE valide.")

        preview = _build_preview_response(
            workbook,
            source_name=file_name,
            file_hash=file_hash,
            affaire_reference=affaire_reference,
            affaire_nge=inferred_affaire_nge,
            demande_gap_days=demande_gap_days,
            campagne_gap_days=campagne_gap_days,
        )
        demande_idx, campagne_idx = _find_sheet_group_indices(
            preview["sheets"],
            sheet_name,
            demande_gap_days,
            campagne_gap_days,
        )

        demande_group = preview["proposals"]["demandes"][demande_idx - 1]
        campagne_group = preview["proposals"]["demandes"][demande_idx - 1]["campagnes"][campagne_idx - 1]
        campagne_sheet_names = {_clean(name) for name in (campagne_group.get("sheets") or []) if _clean(name)}
        campagne_rows = [row for row in preview["sheets"] if _clean(row.get("sheet")) in campagne_sheet_names]
        campagne_unique_dates = {
            _clean(row.get("anchor_date"))
            for row in campagne_rows
            if _clean(row.get("anchor_date"))
        }
        campagne_group_for_import = dict(campagne_group)
        campagne_group_for_import["interventions_count"] = len(campagne_unique_dates)

        demande_id, demande_created = _find_or_create_demande(
            conn,
            int(selected_affaire["id"]),
            demande_idx,
            demande_group,
            reference_override=demande_reference_override,
        )
        campagne_id, campagne_created = _find_or_create_campagne(
            conn,
            demande_id,
            demande_idx,
            campagne_idx,
            campagne_group_for_import,
            reference_override=campagne_reference_override,
        )
        intervention_id, intervention_created = _find_or_create_intervention(
            conn,
            demande_id,
            campagne_id,
            target_sheet,
            file_name,
            file_hash,
            reference_override=intervention_reference_override,
        )
        point_terrain_id, feuille_terrain_id, terrain_created = _find_or_create_de_terrain_point(
            conn,
            intervention_id,
            target_sheet,
            file_name,
            file_hash,
        )

        _ensure_modules_enabled(conn, demande_id, ["essais_terrain", "interventions"])

        demande_ref_row = conn.execute("SELECT reference FROM demandes WHERE id = ?", (demande_id,)).fetchone()
        campagne_ref_row = conn.execute("SELECT reference FROM campagnes WHERE id = ?", (campagne_id,)).fetchone()
        intervention_ref_row = conn.execute("SELECT reference FROM interventions WHERE id = ?", (intervention_id,)).fetchone()

        dem_ref = _clean(demande_ref_row["reference"]) if demande_ref_row else ""
        cam_ref = _clean(campagne_ref_row["reference"]) if campagne_ref_row else ""
        int_ref = _clean(intervention_ref_row["reference"]) if intervention_ref_row else ""

        conn.commit()

    return {
        "status": "ok",
        "mode": "single_sheet_import",
        "file_name": file_name,
        "sheet_name": sheet_name,
        "affaire": {
            "id": selected_affaire["id"],
            "reference": selected_affaire.get("reference") or "",
            "affaire_nge": selected_affaire.get("affaire_nge") or "",
        },
        "grouping": {
            "demande_index": demande_idx,
            "campagne_index": campagne_idx,
            "demande_gap_days": demande_gap_days,
            "campagne_gap_days": campagne_gap_days,
        },
        "created": {
            "demande": demande_created,
            "campagne": campagne_created,
            "intervention": intervention_created,
            "terrain": terrain_created,
        },
        "ids": {
            "demande_id": demande_id,
            "campagne_id": campagne_id,
            "intervention_id": intervention_id,
            "point_terrain_id": point_terrain_id,
            "feuille_terrain_id": feuille_terrain_id,
        },
        "references": {
            "demande_reference": dem_ref,
            "campagne_reference": cam_ref,
            "intervention_reference": int_ref,
        },
    }


@router.post("/preview")
def preview_de_import(payload: PreviewDERequest) -> dict[str, Any]:
    workbook_path = Path(payload.file_path)
    if not workbook_path.exists():
        raise HTTPException(status_code=404, detail=f"Fichier introuvable: {workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Le fichier doit être .xlsx ou .xlsm")

    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Lecture Excel impossible: {exc}") from exc

    return _build_preview_response(
        workbook,
        source_name=workbook_path.name,
        file_hash=hashlib.sha1(workbook_path.read_bytes()).hexdigest(),
        affaire_reference=payload.affaire_reference,
        affaire_nge=payload.affaire_nge,
        demande_gap_days=payload.demande_gap_days,
        campagne_gap_days=payload.campagne_gap_days,
    )


@router.post("/preview-upload")
async def preview_de_import_upload(
    file: UploadFile = File(...),
    affaire_reference: str = Form(default=""),
    affaire_nge: str = Form(default=""),
    demande_gap_days: int = Form(default=120),
    campagne_gap_days: int = Form(default=7),
) -> dict[str, Any]:
    filename = _clean(file.filename)
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Le fichier doit être .xlsx ou .xlsm")

    try:
        content = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Lecture Excel impossible: {exc}") from exc

    return _build_preview_response(
        workbook,
        source_name=filename or "uploaded_file.xlsx",
        file_hash=hashlib.sha1(content).hexdigest(),
        affaire_reference=affaire_reference,
        affaire_nge=affaire_nge,
        demande_gap_days=max(1, min(int(demande_gap_days or 120), 3650)),
        campagne_gap_days=max(1, min(int(campagne_gap_days or 7), 365)),
    )


@router.post("/import-sheet")
def import_de_sheet(payload: ImportDESheetRequest) -> dict[str, Any]:
    workbook_path = Path(payload.file_path)
    if not workbook_path.exists():
        raise HTTPException(status_code=404, detail=f"Fichier introuvable: {workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Le fichier doit etre .xlsx ou .xlsm")

    try:
        content = workbook_path.read_bytes()
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Lecture Excel impossible: {exc}") from exc

    file_hash = hashlib.sha1(content).hexdigest()
    return _materialize_sheet_import(
        workbook=workbook,
        file_name=workbook_path.name,
        file_hash=file_hash,
        sheet_name=payload.sheet_name,
        affaire_reference=payload.affaire_reference,
        affaire_nge=payload.affaire_nge,
        demande_gap_days=payload.demande_gap_days,
        campagne_gap_days=payload.campagne_gap_days,
        demande_reference_override=payload.demande_reference_override,
        campagne_reference_override=payload.campagne_reference_override,
        intervention_reference_override=payload.intervention_reference_override,
    )


@router.post("/import-sheet-upload")
async def import_de_sheet_upload(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    affaire_reference: str = Form(default=""),
    affaire_nge: str = Form(default=""),
    demande_gap_days: int = Form(default=120),
    campagne_gap_days: int = Form(default=7),
    demande_reference_override: str = Form(default=""),
    campagne_reference_override: str = Form(default=""),
    intervention_reference_override: str = Form(default=""),
) -> dict[str, Any]:
    filename = _clean(file.filename) or "uploaded_file.xlsx"
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Le fichier doit etre .xlsx ou .xlsm")

    try:
        content = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Lecture Excel impossible: {exc}") from exc

    file_hash = hashlib.sha1(content).hexdigest()
    return _materialize_sheet_import(
        workbook=workbook,
        file_name=filename,
        file_hash=file_hash,
        sheet_name=sheet_name,
        affaire_reference=affaire_reference,
        affaire_nge=affaire_nge,
        demande_gap_days=max(1, min(int(demande_gap_days or 120), 3650)),
        campagne_gap_days=max(1, min(int(campagne_gap_days or 7), 365)),
        demande_reference_override=demande_reference_override,
        campagne_reference_override=campagne_reference_override,
        intervention_reference_override=intervention_reference_override,
    )
