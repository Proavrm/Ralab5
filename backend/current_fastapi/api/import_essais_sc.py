"""
api/import_essais_sc.py

SC (Sondage Carotté / Core Sample Cut) Excel importer.
Follows the pattern of import_essais_de.py but tailored to SC structure:
- Graphical layout with visual positioning of layers
- Sparse lab data (optional)
- Photo reference system
- Affaire matching and demande hierarchy

Structure:
1. Preview: parse Excel, detect structure, estimate demandes
2. Extract: read sheet cells → normalized payload
3. Materialize: create demande/campagne/intervention/essai hierarchy + feuille_terrain
"""

import sqlite3
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from pathlib import PurePosixPath
from typing import Any, Optional
import json
import hashlib
import zipfile
from io import BytesIO
import tempfile
import xml.etree.ElementTree as ET

import openpyxl
from fastapi import APIRouter, UploadFile, File, Form, HTTPException

from app.core.database import get_db_path
from app.core.wbs import build_sc_wbs_short, join_wbs_display
from api.point_code_logic import allocate_next_point_code_for_scope
from api.import_essais_base import (
    group_rows_by_temporal_gap,
    ensure_hierarchy,
    ensure_modules_enabled,
    _predict_references as _base_predict_references,
    _resolve_affaire_context as _base_resolve_affaire_context,
    _find_demandes_by_affaire as _base_find_demandes_by_affaire,
    _find_campagnes_by_demande as _base_find_campagnes_by_demande,
    _find_interventions_by_campagne as _base_find_interventions_by_campagne,
    _next_demande_reference,
    _next_campaign_reference,
    _next_intervention_reference,
)
from api.sc_point_schema import build_sc_point_payload


router = APIRouter(prefix="/api/import-sc", tags=["import"])


# ──────────────────────────────────────────────────────────────────────────────
# UTILITY: Existing hierarchy lookup by affaire_nge
# ──────────────────────────────────────────────────────────────────────────────

def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalize_affaire_nge(value: str) -> str:
    return re.sub(r"\W+", "", _clean(value)).upper()


def _resolve_sc_affaire_context(
    conn: sqlite3.Connection,
    affaire_reference: str = "",
    affaire_nge_hint: str = "",
    affaire_rst_id: Optional[int] = None,
) -> dict[str, Any]:
    """
    Resolve affaire for SC import/preview.
    If affaire_rst_id is set, it wins (explicit context) so temporal matching always has an affaire row.
    """
    if affaire_rst_id is not None and int(affaire_rst_id) > 0:
        row = conn.execute(
            """
            SELECT id, reference, affaire_nge, chantier, statut
            FROM affaires_rst
            WHERE id = ?
            LIMIT 1
            """,
            (int(affaire_rst_id),),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=400, detail=f"Affaire introuvable: id={affaire_rst_id}")
        return {
            "by_reference": None,
            "by_affaire_nge": None,
            "selected": dict(row),
            "match_mode": "affaire_rst_id",
        }
    return _base_resolve_affaire_context(conn, affaire_reference, affaire_nge_hint)


def _enrich_hierarchy_references(conn: sqlite3.Connection, hierarchy: dict[str, Any]) -> None:
    """Attach demande/campagne/intervention references from DB (ensure_hierarchy does not return them)."""
    did = hierarchy.get("demande_id")
    if did:
        r = conn.execute("SELECT reference FROM demandes WHERE id = ? LIMIT 1", (int(did),)).fetchone()
        if r:
            hierarchy["demande_reference"] = _clean(r["reference"])
    cid = hierarchy.get("campagne_id")
    if cid:
        r = conn.execute("SELECT reference FROM campagnes WHERE id = ? LIMIT 1", (int(cid),)).fetchone()
        if r:
            hierarchy["campagne_reference"] = _clean(r["reference"])
    iid = hierarchy.get("intervention_id")
    if iid:
        r = conn.execute("SELECT reference FROM interventions WHERE id = ? LIMIT 1", (int(iid),)).fetchone()
        if r:
            hierarchy["intervention_reference"] = _clean(r["reference"])


def _sheet_key(name: Any) -> str:
    return str(name or "").strip().upper()


def _get_essais_result_column(conn: sqlite3.Connection) -> str:
    """Return the available JSON result column name for essais table."""
    cols = [row[1] for row in conn.execute("PRAGMA table_info(essais)").fetchall()]
    if "resultats_json" in cols:
        return "resultats_json"
    return "resultats"


def _normalize_filename_key(value: Any) -> str:
    name = Path(str(value or "")).name.strip().lower()
    if not name:
        return ""
    normalized = unicodedata.normalize("NFKD", name)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _build_imported_sc_sheet_map(file_hash: str, file_name: str = "") -> dict[str, int]:
    """
    Build map of already imported SC sheets for this file name (+ optional hash fallback).
    Returns: { sheet_name_normalized: source_uid }
    source_uid is best-effort: feuille_terrain.id or essai_id (legacy).
    """
    file_key = _normalize_filename_key(file_name)
    if not file_key and not file_hash:
        return {}

    imported: dict[str, int] = {}
    db_path = get_db_path()
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        # 1) Import traces in feuilles_terrain.resultats_json.
        feuille_rows = conn.execute(
            """
            SELECT id, resultats_json
            FROM feuilles_terrain
            WHERE code_feuille = 'SC'
            ORDER BY id DESC
            """
        ).fetchall()
        for row in feuille_rows:
            raw = row["resultats_json"]
            if not raw:
                continue
            try:
                payload = json.loads(raw)
            except Exception:
                continue
            payload_name_key = _normalize_filename_key(payload.get("file_name"))
            payload_hash = str(payload.get("file_hash") or "").strip()
            matches_file = bool(file_key and payload_name_key and payload_name_key == file_key)
            if not matches_file and file_hash:
                matches_file = payload_hash == file_hash
            if not matches_file:
                continue
            # Workbook payload shape
            sheets = payload.get("sheets")
            if isinstance(sheets, list):
                for sheet_payload in sheets:
                    if not isinstance(sheet_payload, dict):
                        continue
                    sheet_name = _sheet_key(sheet_payload.get("sheet"))
                    if sheet_name and sheet_name not in imported:
                        imported[sheet_name] = int(row["id"])
            # Single-sheet payload shape
            single_sheet = _sheet_key(payload.get("sheet"))
            if single_sheet and single_sheet not in imported:
                imported[single_sheet] = int(row["id"])

        # 2) Legacy path: SC entries in essais.
        result_col = _get_essais_result_column(conn)
        essai_rows = conn.execute(
            f"""
            SELECT id, {result_col} AS result_blob
            FROM essais
            WHERE essai_code = 'SC'
            ORDER BY id DESC
            """
        ).fetchall()
        for row in essai_rows:
            raw = row["result_blob"]
            if not raw:
                continue
            try:
                payload = json.loads(raw)
            except Exception:
                continue
            if payload.get("source") != "sc_excel_import":
                continue
            payload_name_key = _normalize_filename_key(payload.get("file_name"))
            payload_hash = str(payload.get("file_hash") or "").strip()
            matches_file = bool(file_key and payload_name_key and payload_name_key == file_key)
            if not matches_file and file_hash:
                matches_file = payload_hash == file_hash
            if not matches_file:
                continue
            sheet_name = _sheet_key(payload.get("sheet"))
            if sheet_name and sheet_name not in imported:
                imported[sheet_name] = int(row["id"])

    return imported


def _parse_iso_date(value: Any) -> Optional[datetime.date]:
    raw = _clean(value)
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw[:10]).date()
    except Exception:
        return None


def _extract_sc_number(ws) -> Optional[int]:
    """
    Extract SC carotte number from the workbook header.
    Expected pattern in templates: "SC ... n° ... <number>".
    """

    def _to_int(value: Any) -> Optional[int]:
        if value is None:
            return None
        text = _clean(value)
        if not text:
            return None
        match = re.search(r"(\d+)", text)
        return int(match.group(1)) if match else None

    # Pass 1: compact pattern in a single cell (e.g. "SC n° 1").
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        for col_idx in range(1, min(ws.max_column, 30) + 1):
            cell_text = _clean(ws.cell(row_idx, col_idx).value)
            if not cell_text:
                continue
            upper = cell_text.upper()
            match = re.search(r"\bSC\b.*?N[°ºO]?\s*[:\-]?\s*(\d+)\b", upper)
            if match:
                return int(match.group(1))

    # Pass 2: split cells pattern ("SC" | "n°" | "1") on same row.
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        for col_idx in range(1, min(ws.max_column, 25) + 1):
            head = _clean(ws.cell(row_idx, col_idx).value).upper()
            if head != "SC":
                continue
            window_end = min(ws.max_column, col_idx + 8)
            for scan_col in range(col_idx + 1, window_end + 1):
                token = _clean(ws.cell(row_idx, scan_col).value).upper()
                if "N°" in token or "Nº" in token or token in {"N", "NO", "N O"}:
                    # Prefer explicit number right after n°
                    for num_col in range(scan_col + 1, min(window_end, scan_col + 3) + 1):
                        number = _to_int(ws.cell(row_idx, num_col).value)
                        if number is not None:
                            return number
                    # Fallback: number embedded inside n° token
                    embedded = _to_int(token)
                    if embedded is not None:
                        return embedded

    return None


def _next_sc_feuille_reference(conn: sqlite3.Connection, year: int, labo_code: str) -> str:
    """Generate next feuille reference in format YYYY-LAB-SC####."""
    prefix = f"{year}-{labo_code}-SC"
    rows = conn.execute(
        "SELECT reference FROM feuilles_terrain WHERE reference LIKE ?",
        (f"{prefix}%",),
    ).fetchall()
    numbers: list[int] = []
    for row in rows:
        ref = _clean(row[0] if not isinstance(row, sqlite3.Row) else row["reference"])
        match = re.match(rf"^{re.escape(prefix)}(\d+)$", ref)
        if match:
            numbers.append(int(match.group(1)))
    return f"{prefix}{max(numbers, default=0) + 1:04d}"


def _next_sc_point_reference(conn: sqlite3.Connection, year: int, labo_code: str, prefix_code: str = "CE") -> str:
    prefix = f"{year}-{labo_code}-{prefix_code}"
    rows = conn.execute(
        "SELECT reference FROM points_terrain WHERE reference LIKE ?",
        (f"{prefix}%",),
    ).fetchall()
    numbers: list[int] = []
    for row in rows:
        ref = _clean(row[0] if not isinstance(row, sqlite3.Row) else row["reference"])
        match = re.match(rf"^{re.escape(prefix)}(\d+)$", ref)
        if match:
            numbers.append(int(match.group(1)))
    return f"{prefix}{max(numbers, default=0) + 1:04d}"


def _extract_year_labo_for_sc_reference(payload: dict[str, Any], hierarchy: dict[str, Any]) -> tuple[int, str]:
    """Resolve year/labo for SC feuille reference.

    Rule:
    - Year comes from Excel sheet date (meta.date_sondage) when available.
    - Labo comes from demande reference when available.
    """
    date_sondage = _parse_iso_date((payload.get("meta") or {}).get("date_sondage"))
    year = int(date_sondage.year) if date_sondage else datetime.now().year

    demande_ref = _clean(hierarchy.get("demande_reference"))
    match_demande = re.match(r"^(\d{4})-([A-Z]+)-D\d+", demande_ref)
    if match_demande:
        return year, match_demande.group(2)

    return year, "SP"


def _extract_labo_from_demande_reference(reference: Any) -> str:
    ref = _clean(reference)
    match = re.match(r"^\d{4}-([A-Z]+)-D\d+", ref)
    if match:
        return _clean(match.group(1)) or "SP"
    return "SP"




# ──────────────────────────────────────────────────────────────────────────────

def _preview_sc_workbook(
    file_path: Path,
    affaire_reference: str = "",
    affaire_nge_hint: str = "",
    affaire_rst_id: Optional[int] = None,
) -> dict[str, Any]:
    """
    Preview an SC Excel file: detect sheets, extract headers, list matching demandes/campagnes/interventions.
    Returns structure that can be shown to user before materialization.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as exc:
        raise ValueError(f"Cannot read Excel file: {exc}")

    file_hash = _file_hash(file_path)
    file_name = file_path.name
    imported_sc_map = _build_imported_sc_sheet_map(file_hash, file_name)

    sheets_preview = []
    affaires_detected = set()
    existing_hierarchy = {}  # Cache per affaire_nge
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            sheet_data = _extract_sc_header(ws)
            affaire = sheet_data.get("affaire")
            
            if affaire:
                affaires_detected.add(affaire)
            
            sheets_preview.append({
                "sheet_name": sheet_name,
                "affaire": affaire,
                "date_sondage": sheet_data.get("date_sondage"),
                "type_ouvrage": sheet_data.get("type_ouvrage"),
                "partie_ouvrage": sheet_data.get("partie_ouvrage"),
                "arret_sondage": sheet_data.get("arret_sondage"),
                "sc_number": sheet_data.get("sc_number"),
                "couches_count": _count_couches(ws),
                "already_imported": _sheet_key(sheet_name) in imported_sc_map,
                "existing_essai_id": imported_sc_map.get(_sheet_key(sheet_name)),
            })
        except Exception as exc:
            sheets_preview.append({
                "sheet_name": sheet_name,
                "error": str(exc),
            })

    normalized_from_sheets = sorted({_normalize_affaire_nge(v) for v in affaires_detected if _normalize_affaire_nge(v)})
    inferred_affaire_nge = _normalize_affaire_nge(_clean(affaire_nge_hint))
    if not inferred_affaire_nge and len(normalized_from_sheets) == 1:
        inferred_affaire_nge = normalized_from_sheets[0]

    db_path = get_db_path()
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        affaire_context = _resolve_sc_affaire_context(
            conn,
            affaire_reference,
            inferred_affaire_nge,
            affaire_rst_id=affaire_rst_id,
        )
        selected_affaire = affaire_context.get("selected") or {}

        # Rebuild hierarchy using resolved affaire context to mirror DE behavior.
        existing_hierarchy = {}
        hierarchy_key = inferred_affaire_nge or _normalize_affaire_nge(_clean(selected_affaire.get("affaire_nge")))
        affaire_pk = int(selected_affaire["id"]) if selected_affaire.get("id") is not None else 0
        cache_key = hierarchy_key or (f"id:{affaire_pk}" if affaire_pk else "")
        if hierarchy_key or affaire_pk:
            demandes = _base_find_demandes_by_affaire(
                conn,
                hierarchy_key,
                affaire_rst_id=affaire_pk,
            )
            campagnes_by_demande = {}
            interventions_by_campagne = {}
            for demande in demandes:
                campagnes = _base_find_campagnes_by_demande(conn, demande["id"])
                campagnes_by_demande[demande["id"]] = campagnes
                for campagne in campagnes:
                    interventions = _base_find_interventions_by_campagne(conn, campagne["id"])
                    interventions_by_campagne[campagne["id"]] = interventions
            existing_hierarchy[cache_key] = {
                "demandes": demandes,
                "campagnes_by_demande": campagnes_by_demande,
                "interventions_by_campagne": interventions_by_campagne,
            }

        # Build per-sheet SC preview identifiers:
        # - predicted_sc_reference: future feuille ref (...-SCxxxx)
        # - predicted_point_code: future point code (SCx) when sheet carries an SC number
        resolved_labo = "SP"
        hierarchy_demandes = (existing_hierarchy.get(cache_key or "", {}) or {}).get("demandes") or []
        if hierarchy_demandes:
            resolved_labo = _extract_labo_from_demande_reference(hierarchy_demandes[0].get("reference"))
        else:
            resolved_labo = _extract_labo_from_demande_reference(selected_affaire.get("reference"))

        seq_cache: dict[tuple[int, str], int] = {}

        def _next_preview_sc_reference(date_sondage: Any) -> str:
            parsed = _parse_iso_date(date_sondage)
            year = int(parsed.year) if parsed else datetime.now().year
            labo = resolved_labo or "SP"
            key = (year, labo)
            if key not in seq_cache:
                prefix = f"{year}-{labo}-SC"
                rows = conn.execute(
                    "SELECT reference FROM feuilles_terrain WHERE reference LIKE ?",
                    (f"{prefix}%",),
                ).fetchall()
                max_seq = 0
                for row in rows:
                    ref = _clean(row["reference"] if isinstance(row, sqlite3.Row) else row[0])
                    match = re.match(rf"^{re.escape(prefix)}(\d+)$", ref)
                    if match:
                        max_seq = max(max_seq, int(match.group(1)))
                seq_cache[key] = max_seq
            seq_cache[key] += 1
            return f"{year}-{labo}-SC{seq_cache[key]:04d}"

        for sheet_row in sheets_preview:
            if sheet_row.get("error"):
                continue
            sc_number = sheet_row.get("sc_number")
            if isinstance(sc_number, int) and sc_number > 0:
                sheet_row["predicted_point_code"] = f"SC{sc_number}"
            else:
                sheet_row["predicted_point_code"] = ""
            sheet_row["predicted_sc_reference"] = _next_preview_sc_reference(sheet_row.get("date_sondage"))

            # If sheet is already imported, resolve the REAL existing hierarchy for this sheet only.
            existing_source_id = sheet_row.get("existing_essai_id")
            if not (sheet_row.get("already_imported") and existing_source_id):
                continue

            feuille = conn.execute(
                """
                SELECT id, reference, demande_id, campagne_id, intervention_id, serie_id
                FROM feuilles_terrain
                WHERE id = ?
                LIMIT 1
                """,
                (int(existing_source_id),),
            ).fetchone()
            if not feuille:
                continue

            demande_ref = ""
            campagne_ref = ""
            intervention_ref = ""
            point_code = ""

            if feuille["demande_id"]:
                demande = conn.execute("SELECT reference FROM demandes WHERE id = ? LIMIT 1", (int(feuille["demande_id"]),)).fetchone()
                if demande:
                    demande_ref = _clean(demande["reference"])
            if feuille["campagne_id"]:
                campagne = conn.execute("SELECT reference FROM campagnes WHERE id = ? LIMIT 1", (int(feuille["campagne_id"]),)).fetchone()
                if campagne:
                    campagne_ref = _clean(campagne["reference"])
            if feuille["intervention_id"]:
                intervention = conn.execute("SELECT reference FROM interventions WHERE id = ? LIMIT 1", (int(feuille["intervention_id"]),)).fetchone()
                if intervention:
                    intervention_ref = _clean(intervention["reference"])
            if feuille["serie_id"]:
                point = conn.execute(
                    """
                    SELECT point_code
                    FROM points_terrain
                    WHERE serie_id = ?
                    ORDER BY id ASC
                    LIMIT 1
                    """,
                    (int(feuille["serie_id"]),),
                ).fetchone()
                if point:
                    point_code = _clean(point["point_code"])

            sheet_row["existing_binding"] = {
                "demande_reference": demande_ref,
                "campagne_reference": campagne_ref,
                "intervention_reference": intervention_ref,
                "sc_reference": _clean(feuille["reference"]),
                "sc_code": point_code,
            }

    valid_sheets = [row for row in sheets_preview if not row.get("error")]
    valid_sheets.sort(key=lambda r: ((r.get("date_sondage") or "9999-99-99"), _clean(r.get("sheet_name"))))
    # Keep imported sheets in proposals so preview still shows contextual matches
    # (demande/campagne/intervention) for already imported files.
    preview_candidate_sheets = list(valid_sheets)

    demande_groups = group_rows_by_temporal_gap(preview_candidate_sheets, 120, date_field="date_sondage")
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        predictions = _predict_sc_references_for_preview(conn, demande_groups, 7, labo_code="SP")

    demande_proposals: list[dict[str, Any]] = []
    for d_idx, (d_group, d_pred) in enumerate(zip(demande_groups, predictions), start=1):
        campagne_groups = group_rows_by_temporal_gap(d_group, 7, date_field="date_sondage")
        campagne_proposals: list[dict[str, Any]] = []
        for c_idx, (c_group, c_pred) in enumerate(zip(campagne_groups, d_pred.get("campagnes", [])), start=1):
            raw_interventions = c_pred.get("interventions", [])
            shared_intervention_ref = ""
            if raw_interventions:
                shared_intervention_ref = _clean(raw_interventions[0].get("predicted_intervention_reference"))
            i_pred_map = {
                _clean(r.get("sheet_name")): shared_intervention_ref
                for r in c_group
                if _clean(r.get("sheet_name"))
            }
            campagne_proposals.append(
                {
                    "proposal_index": c_idx,
                    "start_date": c_group[0].get("date_sondage") or "",
                    "end_date": c_group[-1].get("date_sondage") or "",
                    "interventions_count": 1 if c_group else 0,
                    "sheets": [r.get("sheet_name") for r in c_group],
                    "predicted_campagne_reference": c_pred.get("predicted_campagne_reference", ""),
                    "predicted_intervention_references": i_pred_map,
                }
            )
        demande_proposals.append(
            {
                "proposal_index": d_idx,
                "start_date": d_group[0].get("date_sondage") or "",
                "end_date": d_group[-1].get("date_sondage") or "",
                "interventions_count": sum(int(c.get("interventions_count") or 0) for c in campagne_proposals),
                "campagnes_count": len(campagne_proposals),
                "campagnes": campagne_proposals,
                "sheets": [r.get("sheet_name") for r in d_group],
                "predicted_demande_reference": d_pred.get("predicted_demande_reference", ""),
                "imported_count": sum(1 for r in d_group if r.get("already_imported")),
            }
        )

    existing_counts = {"demandes": 0, "campagnes": 0, "interventions": 0}
    for _, hierarchy in existing_hierarchy.items():
        demandes = hierarchy.get("demandes") or []
        existing_counts["demandes"] += len(demandes)
        for demande in demandes:
            did = demande.get("id")
            campagnes = (hierarchy.get("campagnes_by_demande") or {}).get(did) or []
            existing_counts["campagnes"] += len(campagnes)
            for campagne in campagnes:
                cid = campagne.get("id")
                interventions = (hierarchy.get("interventions_by_campagne") or {}).get(cid) or []
                existing_counts["interventions"] += len(interventions)

    return {
        "file_name": file_name,
        "file_hash": file_hash,
        "file_path": str(file_path),
        "sheets_count": len(wb.sheetnames),
        "already_imported_count": sum(1 for row in sheets_preview if row.get("already_imported")),
        "affaires_detected": list(affaires_detected),
        "affaire_nge_detected": normalized_from_sheets,
        "auto_defaults": {
            "affaire_nge_suggested": inferred_affaire_nge,
            "affaire_reference_suggested": (affaire_context.get("selected") or {}).get("reference", ""),
            "demande_gap_days_suggested": 120,
            "campagne_gap_days_suggested": 7,
        },
        "affaire_context": affaire_context,
        "proposals": {
            "demandes_count": len(demande_proposals),
            "campagnes_count": sum(len(d.get("campagnes") or []) for d in demande_proposals),
            "interventions_count": sum(int(d.get("interventions_count") or 0) for d in demande_proposals),
            "demandes": demande_proposals,
        },
        "existing_matches": existing_counts,
        "existing_hierarchy": existing_hierarchy,  # Demandes/campagnes/interventions per affaire_nge
        "sheets": sheets_preview,
    }


def _extract_sc_header(ws) -> dict[str, Any]:
    """
    Extract SC header fields from a worksheet.
    Returns: { affaire, date_sondage, type_ouvrage, partie_ouvrage, arret_sondage, ... }
    """
    # Cell mappings for SC header (EXACT from analysis)
    affaire = ws.cell(5, 12).value  # Row 5, Col L
    date_redaction = ws.cell(5, 16).value  # Row 5, Col P
    type_ouvrage = ws.cell(8, 8).value  # Row 8, Col H
    partie_ouvrage = ws.cell(9, 8).value  # Row 9, Col H
    procede = ws.cell(13, 8).value  # Row 13, Col H
    diametre = ws.cell(14, 8).value  # Row 14, Col H
    date_sondage = ws.cell(15, 8).value  # Row 15, Col H
    arret_sondage = ws.cell(17, 8).value  # Row 17, Col H
    photo_number = ws.cell(13, 20).value  # Row 13, Col T (Photo :)
    sc_number = _extract_sc_number(ws)

    return {
        "affaire": str(affaire or "").strip(),
        "date_redaction": _normalize_date(date_redaction),
        "date_sondage": _normalize_date(date_sondage),
        "type_ouvrage": str(type_ouvrage or "").strip(),
        "partie_ouvrage": str(partie_ouvrage or "").strip(),
        "procede": str(procede or "").strip(),
        "diametre": str(diametre or "").strip(),
        "arret_sondage": str(arret_sondage or "").strip(),
        "photo_number": str(photo_number or "").strip() if photo_number else None,
        "sc_number": sc_number,
    }


def _count_couches(ws) -> int:
    """Count non-empty descriptions in column 9 (couches data)."""
    count = 0
    for row_idx in range(25, ws.max_row + 1):
        desc = ws.cell(row_idx, 9).value
        if desc and str(desc).strip():
            count += 1
    return count


def _normalize_date(value: Any) -> Optional[str]:
    """Convert date value to ISO 8601 string (YYYY-MM-DD)."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        # Try strict ISO first
        try:
            dt = datetime.fromisoformat(text.split()[0])
            return dt.date().isoformat()
        except Exception:
            pass

        # Handle day ranges like "Nuit 07-08/03/2024": keep the end day (08/03/2024).
        range_match = re.search(r"(\d{1,2})\s*-\s*(\d{1,2})\s*[\-/\.]\s*(\d{1,2})\s*[\-/\.]\s*(\d{2,4})", text)
        if range_match:
            day = int(range_match.group(2))
            month = int(range_match.group(3))
            year = int(range_match.group(4))
            if year < 100:
                year += 2000
            try:
                return datetime(year, month, day).date().isoformat()
            except ValueError:
                return text

        # Accept common FR/legacy formats, including noisy strings like "Nuit 12/03/2024".
        date_match = re.search(r"(\d{1,2})[\-/\.](\d{1,2})[\-/\.](\d{2,4})", text)
        if date_match:
            day = int(date_match.group(1))
            month = int(date_match.group(2))
            year = int(date_match.group(3))
            if year < 100:
                year += 2000
            try:
                return datetime(year, month, day).date().isoformat()
            except ValueError:
                return text

        return text
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


def _predict_sc_references_for_preview(
    conn: sqlite3.Connection,
    demande_groups: list[list[dict[str, Any]]],
    campagne_gap_days: int,
    labo_code: str = "SP",
) -> list[dict[str, Any]]:
    return _base_predict_references(
        conn,
        demande_groups,
        campagne_gap_days,
        labo_code=labo_code,
    )


# ──────────────────────────────────────────────────────────────────────────────
# PHASE 2: EXTRACT - Read sheet → normalized payload
# ──────────────────────────────────────────────────────────────────────────────

def _extract_sc_payload(ws, file_name: str, file_hash: str) -> dict[str, Any]:
    """
    Extract full SC sheet data into normalized payload structure.
    """
    header = _extract_sc_header(ws)
    couches = _extract_couches(ws)

    # Build payload (similar structure to DE but adapted for SC)
    return {
        "source": "sc_excel_import",
        "version": 1,
        "file_name": file_name,
        "file_hash": file_hash,
        "sheet": ws.title,
        "meta": {
            "affaire_nge_raw": header.get("affaire"),
            "date_sondage": header.get("date_sondage"),
            "date_redaction": header.get("date_redaction"),
            "type_ouvrage": header.get("type_ouvrage"),
            "partie_ouvrage": header.get("partie_ouvrage"),
            "procede": header.get("procede"),
            "diametre": header.get("diametre"),
            "arret_sondage": header.get("arret_sondage"),
            "photo_number": header.get("photo_number"),
            "sc_number": header.get("sc_number"),
        },
        "couches": couches,
        "raw_cells": _extract_non_empty_cells(ws),
        "resume": {
            "couches_count": len(couches),
            "couches_with_lab": sum(1 for c in couches if c.get("d") or c.get("vide") or c.get("compacite")),
        },
    }


def _extract_couches(ws) -> list[dict[str, Any]]:
    """
    Extract couches (layers) from the SC spreadsheet.
    Returns list of { profondeur, description, d, vide, compacite, ... }
    """
    couches = []
    
    # Find all descriptions in column 9 and their associated lab data
    for row_idx in range(25, ws.max_row + 1):
        desc = ws.cell(row_idx, 9).value
        if not desc or not str(desc).strip():
            continue
        
        # Description found - extract associated data
        d_val = ws.cell(row_idx, 22).value  # Column V
        vide_val = ws.cell(row_idx, 23).value  # Column W
        compac_val = ws.cell(row_idx, 24).value  # Column X
        
        couche = {
            "row": row_idx,
            "description": str(desc).strip(),
            "d": _parse_float(d_val),
            "vide": _parse_float(vide_val),
            "compacite": _parse_float(compac_val),
        }
        couches.append(couche)
    
    return couches


def _extract_non_empty_cells(ws) -> dict[str, Any]:
    """
    Extract all non-empty worksheet cells for a full-fidelity import snapshot.
    Keys are Excel coordinates (A1, B2, ...).
    """
    cells: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str):
                value = value.strip()
                if not value:
                    continue
            cells[cell.coordinate] = value
    return cells


def _parse_float(value: Any) -> Optional[float]:
    """Parse numeric value, skip formulas."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s.startswith("="):
        return None  # Skip formulas
    try:
        return float(s.replace(",", "."))  # Handle French decimal separator
    except ValueError:
        return None


def _file_hash(file_path: Path) -> str:
    """Compute SHA256 hash of file."""
    sha = hashlib.sha256()
    with open(file_path, "rb") as f:
        sha.update(f.read())
    return sha.hexdigest()[:16]


def _get_storage_photos_dir(affaire_nge: str) -> Path:
    """
    Get the directory where photos for this affaire should be stored.
    Path: /storage/essais_photos/{affaire_nge}/
    """
    storage_root = Path(__file__).resolve().parents[3] / "storage"
    photos_dir = storage_root / "essais_photos" / affaire_nge
    photos_dir.mkdir(parents=True, exist_ok=True)
    return photos_dir


def _resolve_zip_target(base_path: str, target: str) -> str:
    """Resolve OPC relationship target path inside .xlsx zip."""
    import posixpath
    base = PurePosixPath(base_path)
    if target.startswith("/"):
        return target.lstrip("/")
    raw = str((base.parent / target).as_posix())
    return posixpath.normpath(raw)


def _find_sheet_drawing_media(zip_ref: zipfile.ZipFile, sheet_name: str) -> Optional[str]:
    """
    Resolve media file linked to a sheet via drawing oneCellAnchor.
    Returns zip path like xl/media/imageX.jpg when found.
    """
    ns_main = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    ns_rel = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    ns_drawing = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    workbook_xml = ET.fromstring(zip_ref.read("xl/workbook.xml"))
    target_sheet = None
    for sheet in workbook_xml.findall("main:sheets/main:sheet", ns_main):
        if sheet.get("name") == sheet_name:
            target_sheet = sheet
            break
    if target_sheet is None:
        return None

    wb_rel_id = target_sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
    if not wb_rel_id:
        return None

    workbook_rels = ET.fromstring(zip_ref.read("xl/_rels/workbook.xml.rels"))
    sheet_target = None
    for rel in workbook_rels.findall("rel:Relationship", ns_rel):
        if rel.get("Id") == wb_rel_id:
            sheet_target = rel.get("Target")
            break
    if not sheet_target:
        return None

    sheet_xml_path = _resolve_zip_target("xl/workbook.xml", sheet_target)
    sheet_xml = ET.fromstring(zip_ref.read(sheet_xml_path))
    drawing_el = sheet_xml.find("main:drawing", ns_main)
    if drawing_el is None:
        return None

    drawing_rel_id = drawing_el.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
    if not drawing_rel_id:
        return None

    sheet_rels_path = f"{PurePosixPath(sheet_xml_path).parent.as_posix()}/_rels/{PurePosixPath(sheet_xml_path).name}.rels"
    sheet_rels = ET.fromstring(zip_ref.read(sheet_rels_path))
    drawing_target = None
    for rel in sheet_rels.findall("rel:Relationship", ns_rel):
        if rel.get("Id") == drawing_rel_id:
            drawing_target = rel.get("Target")
            break
    if not drawing_target:
        return None

    drawing_xml_path = _resolve_zip_target(sheet_xml_path, drawing_target)
    drawing_xml = ET.fromstring(zip_ref.read(drawing_xml_path))

    drawing_rels_path = f"{PurePosixPath(drawing_xml_path).parent.as_posix()}/_rels/{PurePosixPath(drawing_xml_path).name}.rels"
    drawing_rels = ET.fromstring(zip_ref.read(drawing_rels_path))
    drawing_rel_targets = {
        rel.get("Id"): rel.get("Target")
        for rel in drawing_rels.findall("rel:Relationship", ns_rel)
        if rel.get("Id") and rel.get("Target")
    }

    def _anchor_media(anchor) -> Optional[str]:
        """Extract and resolve media path from an anchor element."""
        blip = anchor.find(".//a:blip", ns_drawing)
        if blip is None:
            return None
        embed_id = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
        if not embed_id:
            return None
        media_target = drawing_rel_targets.get(embed_id)
        if not media_target:
            return None
        media_path = _resolve_zip_target(drawing_xml_path, media_target)
        if media_path.startswith("xl/media/") and media_path in zip_ref.namelist():
            return media_path
        return None

    def _anchor_row(anchor) -> int:
        """Return the starting row index of an anchor (0 if unreadable)."""
        from_el = anchor.find("xdr:from", ns_drawing)
        if from_el is None:
            return 0
        try:
            return int(from_el.findtext("xdr:row", "0", namespaces=ns_drawing))
        except (ValueError, TypeError):
            return 0

    # 1. Prefer oneCellAnchor (modern files).
    for anchor in drawing_xml.findall("xdr:oneCellAnchor", ns_drawing):
        media_path = _anchor_media(anchor)
        if media_path:
            return media_path

    # 2. Fall back to twoCellAnchor: skip header-area images (row < 10 = logo zone).
    #    Sort by row so we pick the first non-header image (topmost carotte).
    MIN_CONTENT_ROW = 10
    two_cell = drawing_xml.findall("xdr:twoCellAnchor", ns_drawing)
    two_cell_sorted = sorted(two_cell, key=_anchor_row)
    for anchor in two_cell_sorted:
        if _anchor_row(anchor) < MIN_CONTENT_ROW:
            continue
        media_path = _anchor_media(anchor)
        if media_path:
            return media_path

    return None


def _extract_and_save_photo(
    excel_path: Path, 
    sheet_name: str,
    essai_id: int,
    affaire_nge: str,
    forced_filename: Optional[str] = None,
) -> Optional[str]:
    """
    Extract the first image from the Excel sheet and save it with essai_id as name.
    Returns: filename if saved, None otherwise
    
    The image is saved as: /storage/essais_photos/{affaire_nge}/essai_{essai_id}.jpg
    """
    try:
        # Excel files are ZIP archives - extract media files
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            media_file = _find_sheet_drawing_media(zip_ref, sheet_name)

            # Fallback for legacy files without any recognised drawing anchor.
            if not media_file:
                media_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
                if media_files:
                    media_file = media_files[0]

            if not media_file:
                return None

            image_data = zip_ref.read(media_file)
            
            # Determine extension from filename
            ext = Path(media_file).suffix.lower()  # .png, .jpg, .jpeg, etc
            
            # Save with deterministic name when provided; fallback to essai_id format.
            photos_dir = _get_storage_photos_dir(affaire_nge)
            if forced_filename:
                forced_base = Path(str(forced_filename)).name
                forced_ext = Path(forced_base).suffix.lower()
                if forced_ext in {".jpg", ".jpeg", ".png"}:
                    filename = forced_base
                else:
                    filename = f"{Path(forced_base).stem}{ext}"
            else:
                filename = f"essai_{essai_id}{ext}"
            file_path = photos_dir / filename
            
            with open(file_path, 'wb') as f:
                f.write(image_data)
            
            return filename
    except Exception as e:
        # Log but don't fail - photos are optional
        print(f"Warning: Could not extract photo from Excel: {e}")
        return None


# ──────────────────────────────────────────────────────────────────────────────
# PHASE 3: MATERIALIZE - Create hierarchy in DB
# ──────────────────────────────────────────────────────────────────────────────

def _materialize_sc_sheet(
    payload: dict[str, Any],
    file_path: Path,
    affaire_reference: str = "",
    affaire_nge_hint: str = "",
    demande_gap_days: int = 120,
    campagne_gap_days: int = 7,
    affaire_rst_id: Optional[int] = None,
    excel_path: Optional[Path] = None,
    demande_id: Optional[int] = None,
    campagne_id: Optional[int] = None,
    intervention_id: Optional[int] = None,
) -> dict[str, Any]:
    """
    Materialize SC payload into DB: create essai/feuille_terrain/points_terrain/couches.
    Can optionally link to existing demande/campagne/intervention hierarchy.
    """
    db_path = get_db_path()
    
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        result_col = _get_essais_result_column(conn)
        essais_cols = [row[1] for row in conn.execute("PRAGMA table_info(essais)").fetchall()]
        has_reference_col = "reference" in essais_cols
        
        # Resolve affaire context (affaire_rst_id forces the affaire row — same idea as PMT explicit context)
        affaire_context = _resolve_sc_affaire_context(
            conn,
            affaire_reference,
            affaire_nge_hint or _clean(payload.get("meta", {}).get("affaire_nge_raw")),
            affaire_rst_id=affaire_rst_id,
        )
        if not affaire_context.get("selected"):
            raise HTTPException(status_code=400, detail="Aucune affaire trouvée pour l'import SC")
        selected_affaire = affaire_context.get("selected") or {}
        affaire_nge_for_storage = (
            _clean(selected_affaire.get("affaire_nge"))
            or _clean((payload.get("meta") or {}).get("affaire_nge_raw"))
            or "UNKNOWN"
        )

        # Find or create hierarchy using unified base orchestrator
        anchor_date = _parse_iso_date(payload.get("meta", {}).get("date_sondage")) or datetime.now().date()
        hierarchy = ensure_hierarchy(
            conn,
            affaire_context,
            anchor_date,
            demande_gap_days=demande_gap_days,
            campagne_gap_days=campagne_gap_days,
            demande_id=demande_id,
            campagne_id=campagne_id,
            intervention_id=intervention_id,
            labo_code="SP",
            import_profile_label="Sondage carotté",
        )
        _enrich_hierarchy_references(conn, hierarchy)

        # Activate terrain + interventions modules for this demande (idempotent)
        ensure_modules_enabled(
            conn,
            hierarchy["demande_id"],
            ["interventions", "essais_terrain"],
        )

        terrain_series_intervention_id = hierarchy["intervention_id"]
        terrain_feuille_intervention_id = hierarchy["intervention_id"]
        terrain_point_intervention_id = hierarchy["intervention_id"]
        
        # SC import: no essai created here (essai only created when user creates prelevement → ensaio)
        sheet_name = payload["sheet"]
        sc_number_raw = (payload.get("meta") or {}).get("sc_number")
        if isinstance(sc_number_raw, int):
            sc_number = int(sc_number_raw)
        elif isinstance(sc_number_raw, float) and sc_number_raw.is_integer():
            sc_number = int(sc_number_raw)
        else:
            sc_number = None
        year_ref, labo_ref = _extract_year_labo_for_sc_reference(payload, hierarchy)
        reference = _next_sc_feuille_reference(conn, year_ref, labo_ref)
        now_iso = datetime.now().isoformat()

        # 2. Create series_essais_terrain (shared terrain axis root)
        series_cols = {row[1] for row in conn.execute("PRAGMA table_info(series_essais_terrain)").fetchall()}
        series_ref = f"SER-SC-{hierarchy['demande_id']}-{sheet_name}"
        series_values: dict[str, Any] = {
            "reference": series_ref,
            "demande_id": hierarchy["demande_id"],
            "campagne_id": hierarchy["campagne_id"],
            "intervention_id": terrain_series_intervention_id,
            "code_essai": "SC",
            "libelle_essai": "Coupe de sondage carotté",
            "source_file": _clean(payload.get("source_file")),
            "sheet_name": sheet_name,
            "group_signature": f"SC|{_clean(payload.get('file_hash'))}|{sheet_name}",
            "import_mode": "sc_excel_import",
            "statut": "Importée",
            "date_essai": payload["meta"]["date_sondage"] or "",
            "operateur": "import",
            "section_controlee": payload["meta"].get("partie_ouvrage") or "",
            "couche": "",
            "observations": "Import SC",
            "payload_json": json.dumps(payload),
            "created_at": now_iso,
            "updated_at": now_iso,
        }
        series_insert = {k: v for k, v in series_values.items() if k in series_cols}
        series_columns_sql = ", ".join(series_insert.keys())
        series_placeholders_sql = ", ".join(["?"] * len(series_insert))
        serie_id = conn.execute(
            f"INSERT INTO series_essais_terrain ({series_columns_sql}) VALUES ({series_placeholders_sql})",
            tuple(series_insert.values()),
        ).lastrowid

        # 3. Create feuille_terrain linked to hierarchy/series
        feuilles_cols = {row[1] for row in conn.execute("PRAGMA table_info(feuilles_terrain)").fetchall()}
        feuille_values: dict[str, Any] = {
            "reference": reference,
            "demande_id": hierarchy["demande_id"],
            "campagne_id": hierarchy["campagne_id"],
            "intervention_id": terrain_feuille_intervention_id,
            "serie_id": serie_id,
            "code_feuille": "SC",
            "label": payload["meta"].get("type_ouvrage") or "",
            "date_feuille": payload["meta"].get("date_sondage") or "",
            "operateur": "import",
            "statut": "Importée",
            "observations": "Import SC",
            "resultats_json": json.dumps(payload),
            "created_at": now_iso,
            "updated_at": now_iso,
        }
        feuille_insert = {k: v for k, v in feuille_values.items() if k in feuilles_cols}
        feuille_columns_sql = ", ".join(feuille_insert.keys())
        feuille_placeholders_sql = ", ".join(["?"] * len(feuille_insert))
        feuille_id = conn.execute(
            f"INSERT INTO feuilles_terrain ({feuille_columns_sql}) VALUES ({feuille_placeholders_sql})",
            tuple(feuille_insert.values()),
        ).lastrowid

        # 4. Create point_terrain (1 sheet => 1 point), schema-compatible
        points_cols = {row[1] for row in conn.execute("PRAGMA table_info(points_terrain)").fetchall()}
        point_reference = _next_sc_point_reference(conn, year_ref, labo_ref)
        point_code = allocate_next_point_code_for_scope(
            conn,
            'SC',
            intervention_id=int(terrain_point_intervention_id),
            preferred_number=sc_number,
        )
        photo_number = _clean((payload.get("meta") or {}).get("photo_number"))
        sc_wbs_short = build_sc_wbs_short(
            _clean(affaire_context.get("selected", {}).get("reference")),
            _clean(hierarchy.get("demande_reference")),
            _clean(hierarchy.get("campagne_reference")),
            _clean(hierarchy.get("intervention_reference")),
            _clean(reference),
            _clean(point_code),
        )
        forced_photo_name = f"{sc_wbs_short}__F01.jpg" if sc_wbs_short else (f"SC{photo_number}.jpg" if photo_number else None)
        extracted_photo_name = _extract_and_save_photo(
            file_path,
            sheet_name,
            int(point_code.replace("SC", "") or 1) if str(point_code).upper().startswith("SC") else 1,
            affaire_nge_for_storage,
            forced_filename=forced_photo_name,
        )
        photo_url = f"/api/photos/sc/{affaire_nge_for_storage}/{photo_number}" if photo_number else ""

        point_values: dict[str, Any] = {
            "serie_id": serie_id,
            "intervention_id": terrain_point_intervention_id,
            "campagne_id": hierarchy["campagne_id"],
            "demande_id": hierarchy["demande_id"],
            "reference": point_reference,
            "point_code": point_code,
            "point_type": "SONDAGE_CAROTTE",
            "ordre": 1,
            "localisation": payload["meta"].get("partie_ouvrage") or "",
            "position_label": payload["meta"].get("partie_ouvrage") or "",
            "profil": "",
            "profondeur_haut": 0,
            "profondeur_bas": _parse_depth(payload["meta"].get("arret_sondage") or ""),
            "payload_json": json.dumps(
                build_sc_point_payload(
                    reference=point_reference,
                    point_code=point_code,
                    source="SC_IMPORT",
                    meta=payload.get("meta") or {},
                    couches=payload.get("couches") or [],
                    legacy_flat={
                        **(payload.get("meta") or {}),
                        "photo_number": photo_number,
                        "photo_stored_name": extracted_photo_name or "",
                        "photo_url": photo_url,
                        "wbs_full": join_wbs_display(
                            _clean(affaire_context.get("selected", {}).get("reference")),
                            _clean(hierarchy.get("campagne_reference")),
                            _clean(hierarchy.get("intervention_reference")),
                            _clean(reference),
                            _clean(point_code),
                        ),
                        "wbs_short": sc_wbs_short,
                        "source_payload": payload,
                        "carotte_coupes": [
                            {
                                "id": "coupe-1",
                                "title": "Coupe 1",
                                "photo_stored_name": extracted_photo_name or "",
                                "photo_url": photo_url,
                                "notes": "",
                                "couches": payload.get("couches") or [],
                            }
                        ],
                    },
                    status="imported",
                ),
                ensure_ascii=False,
            ),
            "created_at": now_iso,
        }
        point_insert = {k: v for k, v in point_values.items() if k in points_cols}
        point_columns_sql = ", ".join(point_insert.keys())
        point_placeholders_sql = ", ".join(["?"] * len(point_insert))
        point_id = conn.execute(
            f"INSERT INTO points_terrain ({point_columns_sql}) VALUES ({point_placeholders_sql})",
            tuple(point_insert.values()),
        ).lastrowid
        
        # 5. Create couches from payload
        for idx, couche_data in enumerate(payload.get("couches", []), 1):
            conn.execute(
                """
                INSERT INTO sondage_couches (
                    point_terrain_id,
                    ordre,
                    description_libre,
                    profondeur_eau,
                    payload_json,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    point_id,
                    idx,
                    couche_data.get("description", ""),
                    None,
                    json.dumps(couche_data),
                    datetime.now().isoformat(), datetime.now().isoformat(),
                ),
            )
        
        conn.commit()
        
        return {
            "status": "success",
            "demande_id": hierarchy["demande_id"],
            "campagne_id": hierarchy["campagne_id"],
            "intervention_id": hierarchy["intervention_id"],
            "hierarchy_created": hierarchy["created"],
            "feuille_id": feuille_id,
            "point_id": point_id,
            "reference": reference,
            "point_reference": point_reference,
            "point_code": point_code,
            "couches_created": len(payload.get("couches", [])),
        }


def _materialize_sc_payloads(
    payloads: list[dict[str, Any]],
    file_path: Path,
    affaire_reference: str = "",
    affaire_nge_hint: str = "",
    demande_gap_days: int = 120,
    campagne_gap_days: int = 7,
    affaire_rst_id: Optional[int] = None,
    demande_id: Optional[int] = None,
    campagne_id: Optional[int] = None,
    intervention_id: Optional[int] = None,
) -> dict[str, Any]:
    """
    Materialize a workbook import as ONE feuille terrain with multiple SC points.
    Each sheet becomes one point (SC1, SC2, ...), preserving couche content.
    """
    if not payloads:
        raise HTTPException(status_code=400, detail="Aucune feuille SC à materialiser")

    first_payload = payloads[0]
    db_path = get_db_path()

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")

        affaire_context = _resolve_sc_affaire_context(
            conn,
            affaire_reference,
            affaire_nge_hint or _clean((first_payload.get("meta") or {}).get("affaire_nge_raw")),
            affaire_rst_id=affaire_rst_id,
        )
        if not affaire_context.get("selected"):
            raise HTTPException(status_code=400, detail="Aucune affaire trouvée pour l'import SC")

        # Workbook-level hierarchy: one demande/campagne/intervention anchor for all sheets.
        # Anchor = earliest sheet date so 120-day matching favours joining an existing window that covers the lot.
        sheet_dates: list[Any] = []
        for p in payloads:
            sheet_dates.append((p.get("meta") or {}).get("date_sondage"))
        parsed_dates = [d for d in (_parse_iso_date(x) for x in sheet_dates) if d is not None]
        anchor_date = min(parsed_dates) if parsed_dates else (
            _parse_iso_date((first_payload.get("meta") or {}).get("date_sondage")) or datetime.now().date()
        )
        hierarchy = ensure_hierarchy(
            conn,
            affaire_context,
            anchor_date,
            demande_gap_days=demande_gap_days,
            campagne_gap_days=campagne_gap_days,
            demande_id=demande_id,
            campagne_id=campagne_id,
            intervention_id=intervention_id,
            labo_code="SP",
            import_profile_label="Sondage carotté",
        )
        _enrich_hierarchy_references(conn, hierarchy)

        ensure_modules_enabled(
            conn,
            hierarchy["demande_id"],
            ["interventions", "essais_terrain"],
        )

        terrain_series_intervention_id = hierarchy["intervention_id"]
        terrain_feuille_intervention_id = hierarchy["intervention_id"]
        terrain_point_intervention_id = hierarchy["intervention_id"]

        year_ref, labo_ref = _extract_year_labo_for_sc_reference(first_payload, hierarchy)
        reference = _next_sc_feuille_reference(conn, year_ref, labo_ref)
        now_iso = datetime.now().isoformat()

        workbook_payload = {
            "source": "sc_excel_import_workbook",
            "version": 1,
            "file_name": first_payload.get("file_name"),
            "file_hash": first_payload.get("file_hash"),
            "sheets": payloads,
            "sheet_count": len(payloads),
        }

        # Create one shared terrain series for the whole workbook import.
        series_cols = {row[1] for row in conn.execute("PRAGMA table_info(series_essais_terrain)").fetchall()}
        file_hash = _clean(first_payload.get("file_hash"))
        sheet_names = [_clean(p.get("sheet")) for p in payloads]
        first_sheet = sheet_names[0] if sheet_names else "SC"
        series_ref = f"SER-SC-{hierarchy['demande_id']}-{(file_hash[:8] if file_hash else first_sheet)}"
        series_values: dict[str, Any] = {
            "reference": series_ref,
            "demande_id": hierarchy["demande_id"],
            "campagne_id": hierarchy["campagne_id"],
            "intervention_id": terrain_series_intervention_id,
            "code_essai": "SC",
            "libelle_essai": "Coupe de sondage carotté",
            "source_file": _clean(first_payload.get("source_file")),
            "sheet_name": ", ".join([name for name in sheet_names if name]),
            "group_signature": f"SC|{file_hash}|workbook",
            "import_mode": "sc_excel_import",
            "statut": "Importée",
            "date_essai": (first_payload.get("meta") or {}).get("date_sondage") or "",
            "operateur": "import",
            "section_controlee": (first_payload.get("meta") or {}).get("partie_ouvrage") or "",
            "couche": "",
            "observations": "Import SC (workbook)",
            "payload_json": json.dumps(workbook_payload),
            "created_at": now_iso,
            "updated_at": now_iso,
        }
        series_insert = {k: v for k, v in series_values.items() if k in series_cols}
        series_columns_sql = ", ".join(series_insert.keys())
        series_placeholders_sql = ", ".join(["?"] * len(series_insert))
        serie_id = conn.execute(
            f"INSERT INTO series_essais_terrain ({series_columns_sql}) VALUES ({series_placeholders_sql})",
            tuple(series_insert.values()),
        ).lastrowid

        feuilles_cols = {row[1] for row in conn.execute("PRAGMA table_info(feuilles_terrain)").fetchall()}
        feuille_values: dict[str, Any] = {
            "reference": reference,
            "demande_id": hierarchy["demande_id"],
            "campagne_id": hierarchy["campagne_id"],
            "intervention_id": terrain_feuille_intervention_id,
            "serie_id": serie_id,
            "code_feuille": "SC",
            "label": (first_payload.get("meta") or {}).get("type_ouvrage") or "",
            "date_feuille": (first_payload.get("meta") or {}).get("date_sondage") or "",
            "operateur": "import",
            "statut": "Importée",
            "observations": "Import SC (workbook)",
            "resultats_json": json.dumps(workbook_payload),
            "created_at": now_iso,
            "updated_at": now_iso,
        }
        feuille_insert = {k: v for k, v in feuille_values.items() if k in feuilles_cols}
        feuille_columns_sql = ", ".join(feuille_insert.keys())
        feuille_placeholders_sql = ", ".join(["?"] * len(feuille_insert))
        feuille_id = conn.execute(
            f"INSERT INTO feuilles_terrain ({feuille_columns_sql}) VALUES ({feuille_placeholders_sql})",
            tuple(feuille_insert.values()),
        ).lastrowid

        points_cols = {row[1] for row in conn.execute("PRAGMA table_info(points_terrain)").fetchall()}
        point_results: list[dict[str, Any]] = []

        def _point_sort_key(item: dict[str, Any]) -> tuple[int, str]:
            meta = item.get("meta") or {}
            raw = meta.get("sc_number")
            if isinstance(raw, int):
                return (raw, _clean(item.get("sheet")))
            if isinstance(raw, float) and raw.is_integer():
                return (int(raw), _clean(item.get("sheet")))
            return (10_000_000, _clean(item.get("sheet")))

        ordered_payloads = sorted(payloads, key=_point_sort_key)
        year_ref, labo_ref = _extract_year_labo_for_sc_reference(ordered_payloads[0] if ordered_payloads else {}, hierarchy)
        reserved_sc_numbers: set[int] = set()

        for point_order, payload in enumerate(ordered_payloads, start=1):
            meta = payload.get("meta") or {}
            sc_number_raw = meta.get("sc_number")
            if isinstance(sc_number_raw, int):
                sc_number = int(sc_number_raw)
            elif isinstance(sc_number_raw, float) and sc_number_raw.is_integer():
                sc_number = int(sc_number_raw)
            else:
                sc_number = None

            point_code = allocate_next_point_code_for_scope(
                conn,
                'SC',
                intervention_id=int(terrain_point_intervention_id),
                preferred_number=sc_number,
                reserved_numbers=reserved_sc_numbers,
            )
            point_reference = _next_sc_point_reference(conn, year_ref, labo_ref)

            photo_number = _clean(meta.get("photo_number"))
            sc_wbs_short = build_sc_wbs_short(
                _clean(affaire_context.get("selected", {}).get("reference")),
                _clean(hierarchy.get("demande_reference")),
                _clean(hierarchy.get("campagne_reference")),
                _clean(hierarchy.get("intervention_reference")),
                _clean(reference),
                _clean(point_code),
            )
            forced_photo_name = f"{sc_wbs_short}__F01.jpg" if sc_wbs_short else (f"SC{photo_number}.jpg" if photo_number else None)
            extracted_photo_name = _extract_and_save_photo(
                file_path,
                _clean(payload.get("sheet")) or f"SC{point_order}",
                point_order,
                affaire_nge_for_storage,
                forced_filename=forced_photo_name,
            )
            photo_url = f"/api/photos/sc/{affaire_nge_for_storage}/{photo_number}" if photo_number else ""

            point_values: dict[str, Any] = {
                "serie_id": serie_id,
                "intervention_id": terrain_point_intervention_id,
                "campagne_id": hierarchy["campagne_id"],
                "demande_id": hierarchy["demande_id"],
                "reference": point_reference,
                "point_code": point_code,
                "point_type": "SONDAGE_CAROTTE",
                "ordre": len(point_results) + 1,
                "localisation": meta.get("partie_ouvrage") or "",
                "position_label": meta.get("partie_ouvrage") or "",
                "profil": "",
                "profondeur_haut": 0,
                "profondeur_bas": _parse_depth(meta.get("arret_sondage") or ""),
                "payload_json": json.dumps(
                    build_sc_point_payload(
                        reference=point_reference,
                        point_code=point_code,
                        source="SC_IMPORT",
                        meta=meta,
                        couches=payload.get("couches") or [],
                        legacy_flat={
                            **meta,
                            "photo_number": photo_number,
                            "photo_stored_name": extracted_photo_name or "",
                            "photo_url": photo_url,
                            "wbs_full": join_wbs_display(
                                _clean(affaire_context.get("selected", {}).get("reference")),
                                _clean(hierarchy.get("campagne_reference")),
                                _clean(hierarchy.get("intervention_reference")),
                                _clean(reference),
                                _clean(point_code),
                            ),
                            "wbs_short": sc_wbs_short,
                            "source_payload": payload,
                            "carotte_coupes": [
                                {
                                    "id": "coupe-1",
                                    "title": "Coupe 1",
                                    "photo_stored_name": extracted_photo_name or "",
                                    "photo_url": photo_url,
                                    "notes": "",
                                    "couches": payload.get("couches") or [],
                                }
                            ],
                        },
                        status="imported",
                    ),
                    ensure_ascii=False,
                ),
                "created_at": now_iso,
            }
            point_insert = {k: v for k, v in point_values.items() if k in points_cols}
            point_columns_sql = ", ".join(point_insert.keys())
            point_placeholders_sql = ", ".join(["?"] * len(point_insert))
            point_id = conn.execute(
                f"INSERT INTO points_terrain ({point_columns_sql}) VALUES ({point_placeholders_sql})",
                tuple(point_insert.values()),
            ).lastrowid

            for idx, couche_data in enumerate(payload.get("couches", []), 1):
                conn.execute(
                    """
                    INSERT INTO sondage_couches (
                        point_terrain_id,
                        ordre,
                        description_libre,
                        profondeur_eau,
                        payload_json,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        point_id,
                        idx,
                        couche_data.get("description", ""),
                        None,
                        json.dumps(couche_data),
                        datetime.now().isoformat(), datetime.now().isoformat(),
                    ),
                )

            point_results.append(
                {
                    "sheet": _clean(payload.get("sheet")),
                    "point_id": int(point_id),
                    "point_reference": point_reference,
                    "point_code": point_code,
                    "couches_created": len(payload.get("couches", [])),
                }
            )

        conn.commit()

        # Backward-compatible convenience fields for clients expecting single-point shape.
        top_point_id: Optional[int] = None
        top_point_reference: Optional[str] = None
        top_point_code: Optional[str] = None
        if len(point_results) == 1:
            top_point_id = int(point_results[0].get("point_id"))
            top_point_reference = _clean(point_results[0].get("point_reference"))
            top_point_code = _clean(point_results[0].get("point_code"))

        return {
            "status": "success",
            "demande_id": hierarchy["demande_id"],
            "campagne_id": hierarchy["campagne_id"],
            "intervention_id": hierarchy["intervention_id"],
            "hierarchy_created": hierarchy["created"],
            "feuille_id": int(feuille_id),
            "reference": reference,
            "point_id": top_point_id,
            "point_reference": top_point_reference,
            "point_code": top_point_code,
            "points_created": point_results,
            "sheets_imported": len(payloads),
        }


def _parse_depth(depth_str: str) -> Optional[float]:
    """Parse depth from string like '25,5 cm' → 25.5"""
    if not depth_str:
        return None
    s = depth_str.strip().lower()
    s = s.replace(" cm", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_depth_m(depth_str: str) -> Optional[float]:
    """Parse depth in meters."""
    cm = _parse_depth(depth_str)
    return cm / 100 if cm else None


# ──────────────────────────────────────────────────────────────────────────────
# API ENDPOINTS
# ──────────────────────────────────────────────────────────────────────────────

@router.post("/preview")
async def preview_sc_import(
    file: UploadFile = File(...),
    affaire_reference: str = Form(""),
    affaire_nge: str = Form(""),
    affaire_rst_id: Optional[int] = Form(None),
):
    """
    Preview SC Excel file: show sheets, headers, estimated structure.
    """
    try:
        # Save temp file (cross-platform, including Windows)
        tmp_dir = Path(tempfile.gettempdir())
        temp_path = tmp_dir / file.filename
        with open(temp_path, "wb") as f:
            f.write(await file.read())

        preview = _preview_sc_workbook(
            temp_path,
            affaire_reference=affaire_reference,
            affaire_nge_hint=affaire_nge,
            affaire_rst_id=affaire_rst_id,
        )
        if temp_path.exists():
            temp_path.unlink()
        
        return preview
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/materialize")
async def materialize_sc_import(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    affaire_reference: str = Form(""),
    affaire_nge: str = Form(""),
    affaire_rst_id: Optional[int] = Form(None),
    demande_gap_days: int = Form(120),
    campagne_gap_days: int = Form(7),
    demande_id: Optional[int] = Form(None),
    campagne_id: Optional[int] = Form(None),
    intervention_id: Optional[int] = Form(None),
):
    """
    Upload SC Excel file and materialize selected sheet(s).

    Behavior:
    - `sheet_name` provided: imports one sheet as one feuille with one point.
    - no `sheet_name`: imports workbook as ONE feuille with one point per sheet (SC1, SC2, ...).
    """
    try:
        # Save temp file (cross-platform, including Windows)
        tmp_dir = Path(tempfile.gettempdir())
        temp_path = tmp_dir / file.filename
        with open(temp_path, "wb") as f:
            f.write(await file.read())
        
        # Materialize selected sheet(s)
        wb = openpyxl.load_workbook(temp_path, data_only=True)
        file_hash = _file_hash(temp_path)
        target_sheets = [sheet_name] if sheet_name else list(wb.sheetnames)

        for target in target_sheets:
            if target not in wb.sheetnames:
                raise HTTPException(status_code=400, detail=f"Feuille introuvable: {target}")

        payloads = []
        for target in target_sheets:
            ws = wb[target]
            payloads.append(_extract_sc_payload(ws, file.filename, file_hash))

        if sheet_name:
            result = _materialize_sc_sheet(
                payloads[0],
                temp_path,
                affaire_reference=affaire_reference,
                affaire_nge_hint=affaire_nge,
                demande_gap_days=demande_gap_days,
                campagne_gap_days=campagne_gap_days,
                affaire_rst_id=affaire_rst_id,
                demande_id=demande_id,
                campagne_id=campagne_id,
                intervention_id=intervention_id,
                excel_path=temp_path,
            )
        else:
            result = _materialize_sc_payloads(
                payloads,
                temp_path,
                affaire_reference=affaire_reference,
                affaire_nge_hint=affaire_nge,
                demande_gap_days=demande_gap_days,
                campagne_gap_days=campagne_gap_days,
                affaire_rst_id=affaire_rst_id,
                demande_id=demande_id,
                campagne_id=campagne_id,
                intervention_id=intervention_id,
            )
        
        if temp_path.exists():
            temp_path.unlink()
        
        if sheet_name:
            return {
                "status": result.get("status", "success"),
                "file_name": file.filename,
                "sheet_name": sheet_name,
                **result,
            }

        return {
            "status": "success",
            "file_name": file.filename,
            "sheets_imported": len(payloads),
            "results": [result],
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
